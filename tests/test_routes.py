import io
import json
import shutil
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import yaml
from bs4 import BeautifulSoup
from bs4.element import Tag
from flask import Flask
from openpyxl import Workbook, load_workbook
from werkzeug.datastructures import FileStorage

from app.auth import SAML_USER_SESSION_KEY
from app.config import CONFIG_FILENAME
from app.db import get_db, get_ip_username, get_setting, init_db, seed_defaults, set_setting
from app.documents import DocumentReadError
from app.formatting import render_markdown
from app.routes import (
    UPLOAD_PATH_SAFE_CHARS,
    _consistency_task_title,
    _find_enabled_model,
    _parse_result_json,
    _prepare_task_results,
    _upload_destination,
    get_enabled_models,
    register_routes,
)
from app.task_types import CONSISTENCY_TASK_TYPE, DOCUMENT_TASK_TYPE, IMAGE_TASK_TYPE, LANGUAGE_CONSISTENCY_TASK_TYPE, VIDEO_TASK_TYPE


_TINY_PNG = (
    b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01"
    b"\x08\x06\x00\x00\x00\x1f\x15\xc4\x89\x00\x00\x00\nIDATx\x9cc\x00\x01"
    b"\x00\x00\x05\x00\x01\r\n-\xb4\x00\x00\x00\x00IEND\xaeB`\x82"
)


def _required_tag(value: object) -> Tag:
    if not isinstance(value, Tag):
        raise AssertionError("expected HTML tag")
    return value


def _xlsx_bytes(rows, *, title: str = "Sheet1") -> io.BytesIO:
    output = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = title
    for row in rows:
        sheet.append(row)
    workbook.save(output)
    workbook.close()
    output.seek(0)
    return output


def _pdf_with_image_bytes() -> io.BytesIO:
    import fitz

    document = fitz.open()
    page = document.new_page(width=240, height=180)
    page.insert_text((24, 32), "图 1 是电源接线图。")
    page.insert_image(fitz.Rect(24, 52, 84, 112), stream=_TINY_PNG)
    output = io.BytesIO(document.tobytes())
    document.close()
    output.seek(0)
    return output


def _saml_auth_config() -> dict:
    return {
        "mode": "saml",
        "saml": {
            "sp_entity_id": "https://doc.example.com/auth/saml/metadata",
            "acs_url": "https://doc.example.com/auth/saml/acs",
            "idp_entity_id": "https://sso.example.com/idp",
            "idp_sso_url": "https://sso.example.com/login",
            "idp_x509_cert": "test-cert",
            "user_id_attribute": "uid",
            "username_attribute": "displayName",
        },
    }


class _FakeSamlAuth:
    def __init__(self):
        self.processed_request_id = None

    def login(self, return_to=None):
        self.return_to = return_to
        return "https://sso.example.com/login?SAMLRequest=test"

    def process_response(self, request_id=None):
        self.processed_request_id = request_id

    def get_last_request_id(self):
        return "REQ-1"

    def get_errors(self):
        return []

    def is_authenticated(self):
        return True

    def get_nameid(self):
        return "nameid-1"

    def get_attributes(self):
        return {"uid": ["100086"], "displayName": ["张三"]}

    def get_friendlyname_attributes(self):
        return {}


class AdminSettingsRouteTest(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        root_dir = Path(self.temp_dir.name)
        project_root = Path(__file__).resolve().parents[1]
        self.app = Flask(
            __name__,
            template_folder=str(project_root / "app" / "templates"),
            static_folder=str(project_root / "app" / "static"),
        )
        self.app.add_template_filter(render_markdown, "markdown")
        self.app.config.update(
            SECRET_KEY="test-secret",
            ADMIN_URL="/admin",
            ROOT_DIR=root_dir,
            DATABASE=str(root_dir / "test.sqlite3"),
            UPLOAD_FOLDER=str(root_dir / "uploads"),
            MAX_UPLOAD_MB=1024,
            MAX_CONTENT_LENGTH=1024 * 1024 * 1024,
            NETWORK={"proxy_mode": "direct", "proxy": "", "ssl_verify": False},
        )
        Path(self.app.config["UPLOAD_FOLDER"]).mkdir(parents=True, exist_ok=True)
        with self.app.app_context():
            init_db()
            seed_defaults()
        register_routes(self.app)
        self.client = self.app.test_client()
        with self.client.session_transaction() as session:
            session["admin_logged_in"] = True

    def tearDown(self):
        self.temp_dir.cleanup()

    def _logout_test_client(self):
        with self.client.session_transaction() as session:
            session.clear()

    def _configure_provider(self, owner_subject: str = "ip:127.0.0.1") -> str:
        with self.app.app_context():
            now = "2026-05-01 09:00:00"
            cursor = get_db().execute(
                """
                INSERT INTO user_model_providers(
                    owner_subject, name, api_base, api_key,
                    request_timeout, max_input_chars, is_active, created_at, updated_at
                )
                VALUES (?, '测试提供商', 'https://example.test/v1/chat/completions', '',
                        30, 80000, 1, ?, ?)
                """,
                (owner_subject, now, now),
            )
            provider_id = cursor.lastrowid
            get_db().execute(
                """
                INSERT INTO user_model_configs(provider_id, model_name, force_disable_thinking, sort_order, created_at, updated_at)
                VALUES (?, 'model-a', 0, 10, ?, ?)
                """,
                (provider_id, now, now),
            )
            get_db().commit()
        return f"{provider_id}:0:model-a"

    def _reject_task_inserts(self):
        with self.app.app_context():
            db = get_db()
            db.execute(
                """
                CREATE TRIGGER reject_task_insert
                BEFORE INSERT ON tasks
                BEGIN
                    SELECT RAISE(ABORT, 'forced task insert failure');
                END
                """
            )
            db.commit()

    def _insert_task(
        self,
        *,
        task_type: str = DOCUMENT_TASK_TYPE,
        ip: str = "127.0.0.1",
        status: str = "completed",
        created_at: str = "2026-05-01 10:00:00",
        username_snapshot: str | None = None,
        owner_subject: str | None = None,
        owner_name_snapshot: str | None = None,
        owner_source: str | None = None,
    ):
        with self.app.app_context():
            cursor = get_db().execute(
                """
                INSERT INTO tasks(
                    task_type, ip, username_snapshot, owner_subject, owner_name_snapshot, owner_source,
                    original_filename, stored_filename, file_type,
                    file_size, checks_json, model_name, api_base, status, progress, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, '测试文档.txt', 'stored.txt', 'txt', 12, '[]', 'model-a', 'https://example.test/v1/chat/completions', ?, 100, ?, ?)
                """,
                (
                    task_type,
                    ip,
                    username_snapshot,
                    owner_subject,
                    owner_name_snapshot,
                    owner_source,
                    status,
                    created_at,
                    created_at,
                ),
            )
            get_db().commit()
            return cursor.lastrowid

    def test_admin_delete_task_reports_locked_file_without_removing_task(self):
        self._insert_task()
        upload_path = Path(self.app.config["UPLOAD_FOLDER"]) / "stored.txt"
        upload_path.write_text("content", encoding="utf-8")
        with self.app.app_context():
            task_id = get_db().execute("SELECT id FROM tasks WHERE stored_filename = 'stored.txt'").fetchone()["id"]

        with patch("app.routes.remove_file", return_value=(False, "[WinError 32] 文件正被占用")):
            response = self.client.post(f"/admin/tasks/{task_id}/delete", follow_redirects=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("任务文件正被其他程序使用", response.get_data(as_text=True))
        with self.app.app_context():
            task = get_db().execute("SELECT * FROM tasks WHERE id = ?", (task_id,)).fetchone()
        self.assertIsNotNone(task)
        self.assertTrue(upload_path.exists())

    def test_admin_delete_task_removes_report_suppression_hits(self):
        task_id = self._insert_task()
        with self.app.app_context():
            db = get_db()
            now = "2026-05-01 10:05:00"
            rule_id = db.execute(
                """
                INSERT INTO report_suppression_rules(
                    task_type, check_code, fingerprint, item_json, created_at, updated_at
                )
                VALUES (?, 'typo', 'fingerprint', '{}', ?, ?)
                """,
                (DOCUMENT_TASK_TYPE, now, now),
            ).lastrowid
            db.execute(
                """
                INSERT INTO report_suppression_hits(
                    rule_id, task_id, result_code, item_id, item_json, created_at
                )
                VALUES (?, ?, 'typo', 'item-1', '{}', ?)
                """,
                (rule_id, task_id, now),
            )
            db.commit()

        response = self.client.post(f"/admin/tasks/{task_id}/delete")

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            hit = get_db().execute(
                "SELECT id FROM report_suppression_hits WHERE task_id = ?",
                (task_id,),
            ).fetchone()
        self.assertIsNone(hit)

    def test_user_bulk_delete_removes_selected_history_tasks(self):
        deletable_task_ids = [
            self._insert_task(task_type=IMAGE_TASK_TYPE, status="failed"),
            self._insert_task(task_type=IMAGE_TASK_TYPE, status="canceled", created_at="2026-05-01 10:01:00"),
            self._insert_task(task_type=IMAGE_TASK_TYPE, status="completed", created_at="2026-05-01 10:02:00"),
        ]
        queued_task_id = self._insert_task(
            task_type=IMAGE_TASK_TYPE,
            status="queued",
            created_at="2026-05-01 10:03:00",
        )

        response = self.client.post(
            "/tasks/bulk-delete",
            data={"task_ids": [*deletable_task_ids, queued_task_id], "next": "/images?page=2"},
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers["Location"], "/images?page=2")
        with self.app.app_context():
            remaining_ids = {
                row["id"]
                for row in get_db().execute(
                    "SELECT id FROM tasks WHERE id IN (?, ?, ?, ?)",
                    (*deletable_task_ids, queued_task_id),
                ).fetchall()
            }
        self.assertEqual(remaining_ids, {queued_task_id})
        with self.client.session_transaction() as session:
            messages = [message for _, message in session.get("_flashes", [])]
        self.assertIn("已批量删除 3 个任务。", messages)
        self.assertIn("已跳过 1 个排队中或运行中的任务。", messages)

    def test_user_bulk_delete_rejects_another_users_task(self):
        task_id = self._insert_task(
            status="completed",
            ip="10.0.0.8",
            owner_subject="ip:10.0.0.8",
            owner_source="ip",
        )

        response = self.client.post(
            "/tasks/bulk-delete",
            data={"task_ids": [task_id], "next": "/"},
        )

        self.assertEqual(response.status_code, 404)
        with self.app.app_context():
            task = get_db().execute("SELECT id FROM tasks WHERE id = ?", (task_id,)).fetchone()
        self.assertIsNotNone(task)

    def test_admin_bulk_delete_removes_selected_canceled_task(self):
        canceled_task_id = self._insert_task(
            task_type=VIDEO_TASK_TYPE,
            status="canceled",
            ip="10.0.0.8",
            owner_subject="ip:10.0.0.8",
            owner_source="ip",
        )

        response = self.client.post(
            "/admin/tasks/bulk-delete",
            data={"task_ids": [canceled_task_id], "next": "/admin/videos?status=canceled"},
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers["Location"], "/admin/videos?status=canceled")
        with self.app.app_context():
            task = get_db().execute("SELECT id FROM tasks WHERE id = ?", (canceled_task_id,)).fetchone()
        self.assertIsNone(task)

    def test_all_task_lists_expose_bulk_selection_for_history_tasks(self):
        task_routes = (
            (DOCUMENT_TASK_TYPE, "/", "/admin/tasks"),
            (CONSISTENCY_TASK_TYPE, "/consistency", "/admin/consistency"),
            (LANGUAGE_CONSISTENCY_TASK_TYPE, "/language-consistency", "/admin/language-consistency"),
            (IMAGE_TASK_TYPE, "/images", "/admin/images"),
            (VIDEO_TASK_TYPE, "/videos", "/admin/videos"),
        )

        for task_type, user_list_url, admin_list_url in task_routes:
            deletable_task_ids = {
                self._insert_task(task_type=task_type, status="failed"),
                self._insert_task(task_type=task_type, status="completed", created_at="2026-05-01 10:01:00"),
                self._insert_task(task_type=task_type, status="canceled", created_at="2026-05-01 10:02:00"),
            }
            self._insert_task(task_type=task_type, status="queued", created_at="2026-05-01 10:03:00")
            self._insert_task(task_type=task_type, status="running", created_at="2026-05-01 10:04:00")
            for list_url, action in (
                (user_list_url, "/tasks/bulk-delete"),
                (admin_list_url, "/admin/tasks/bulk-delete"),
            ):
                with self.subTest(task_type=task_type, list_url=list_url):
                    response = self.client.get(list_url)
                    self.assertEqual(response.status_code, 200)
                    soup = BeautifulSoup(response.get_data(as_text=True), "html.parser")
                    form = _required_tag(soup.select_one("[data-bulk-delete-form]"))
                    checkboxes = soup.select("[data-bulk-task]")
                    toggle = _required_tag(soup.select_one("[data-bulk-task-toggle]"))
                    button = _required_tag(soup.select_one("[data-bulk-delete-button]"))
                    self.assertEqual(form.get("action"), action)
                    self.assertEqual({int(checkbox.get("value")) for checkbox in checkboxes}, deletable_task_ids)
                    self.assertTrue(all(checkbox.get("form") == "bulk-delete-tasks" for checkbox in checkboxes))
                    self.assertIsNone(toggle.get("disabled"))
                    self.assertIsNotNone(button.get("disabled"))
                    self.assertEqual(button.get_text(" ", strip=True), "批量删除")

    def test_diagnostics_fetch_returns_saved_state(self):
        response = self.client.post(
            "/admin/settings",
            data={"action": "diagnostics", "llm_stream_trace_enabled": "on"},
            headers={"X-Requested-With": "fetch"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json(), {"llm_stream_trace_enabled": True})
        with self.app.app_context():
            self.assertTrue(get_setting("llm_stream_trace_enabled"))

    def test_diagnostics_toggle_is_unchecked_by_default(self):
        response = self.client.get("/admin/settings")

        self.assertEqual(response.status_code, 200)
        soup = BeautifulSoup(response.get_data(as_text=True), "html.parser")
        toggle = _required_tag(soup.find("input", {"name": "llm_stream_trace_enabled"}))
        self.assertEqual(toggle.get("data-saved-checked"), "false")
        self.assertIsNone(toggle.get("checked"))

    def test_diagnostics_toggle_treats_text_false_as_unchecked(self):
        with self.app.app_context():
            set_setting("llm_stream_trace_enabled", "false")

        response = self.client.get("/admin/settings")

        self.assertEqual(response.status_code, 200)
        soup = BeautifulSoup(response.get_data(as_text=True), "html.parser")
        toggle = _required_tag(soup.find("input", {"name": "llm_stream_trace_enabled"}))
        self.assertEqual(toggle.get("data-saved-checked"), "false")
        self.assertIsNone(toggle.get("checked"))

    def test_diagnostics_fetch_can_disable_setting(self):
        self.client.post(
            "/admin/settings",
            data={"action": "diagnostics", "llm_stream_trace_enabled": "on"},
            headers={"X-Requested-With": "fetch"},
        )

        response = self.client.post(
            "/admin/settings",
            data={"action": "diagnostics", "llm_stream_trace_enabled": "off"},
            headers={"X-Requested-With": "fetch"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json(), {"llm_stream_trace_enabled": False})
        with self.app.app_context():
            self.assertFalse(get_setting("llm_stream_trace_enabled"))

    def test_diagnostics_accept_json_returns_saved_state(self):
        response = self.client.post(
            "/admin/settings",
            data={"action": "diagnostics", "llm_stream_trace_enabled": "on"},
            headers={"Accept": "application/json"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json(), {"llm_stream_trace_enabled": True})

    def test_admin_settings_shows_network_config(self):
        self.app.config["NETWORK"] = {
            "proxy_mode": "custom",
            "proxy": "http://127.0.0.1:7890",
            "ssl_verify": True,
        }

        response = self.client.get("/admin/settings")

        self.assertEqual(response.status_code, 200)
        soup = BeautifulSoup(response.get_data(as_text=True), "html.parser")
        form = _required_tag(soup.find("form", {"class": "settings-network-form"}))
        proxy_mode_select = _required_tag(soup.find("select", {"name": "proxy_mode"}))
        selected_option = _required_tag(proxy_mode_select.find("option", selected=True))
        self.assertEqual(form.get("data-network-proxy-mode"), "custom")
        self.assertEqual(selected_option.get("value"), "custom")
        proxy_input = _required_tag(soup.find("input", {"name": "proxy"}))
        self.assertEqual(proxy_input.get("value"), "http://127.0.0.1:7890")
        self.assertIsNotNone(proxy_input.get("required"))
        ssl_verify_input = _required_tag(soup.find("input", {"name": "ssl_verify"}))
        self.assertIsNotNone(ssl_verify_input.get("checked"))

    def test_admin_settings_marks_proxy_field_hidden_by_default(self):
        response = self.client.get("/admin/settings")

        self.assertEqual(response.status_code, 200)
        soup = BeautifulSoup(response.get_data(as_text=True), "html.parser")
        form = _required_tag(soup.find("form", {"class": "settings-network-form"}))
        proxy_field = soup.select_one(".settings-network-proxy-field")
        proxy_input = _required_tag(soup.find("input", {"name": "proxy"}))
        self.assertEqual(form.get("data-network-proxy-mode"), "direct")
        self.assertIsNotNone(proxy_field)
        self.assertIsNone(proxy_input.get("required"))

    def test_admin_settings_saves_task_limits(self):
        response = self.client.post(
            "/admin/settings",
            data={
                "action": "concurrency",
                "global_concurrency": "4",
                "user_concurrency": "2",
                "check_item_concurrency": "3",
                "image_page_check_max_pages": "36",
                "issue_output_limit": "100",
                "report_retention_days": "14",
            },
        )

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            self.assertEqual(get_setting("global_concurrency"), 4)
            self.assertEqual(get_setting("user_concurrency"), 2)
            self.assertEqual(get_setting("check_item_concurrency"), 3)
            self.assertEqual(get_setting("image_page_check_max_pages"), 36)
            self.assertEqual(get_setting("issue_output_limit"), 30)
            self.assertEqual(get_setting("report_retention_days"), 14)

    def test_admin_settings_saves_network_to_yaml_config(self):
        response = self.client.post(
            "/admin/settings",
            data={
                "action": "network",
                "proxy_mode": "custom",
                "proxy": " http://127.0.0.1:7890 ",
                "ssl_verify": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        expected = {"proxy_mode": "custom", "proxy": "http://127.0.0.1:7890", "ssl_verify": True}
        self.assertEqual(self.app.config["NETWORK"], expected)
        config = yaml.safe_load((self.app.config["ROOT_DIR"] / CONFIG_FILENAME).read_text(encoding="utf-8"))
        self.assertEqual(config["network"], expected)
        with self.app.app_context():
            self.assertIsNone(get_setting("network"))

    def test_admin_settings_saves_ip_username_mapping_in_ip_mode(self):
        self._insert_task(ip="10.0.0.8")

        response = self.client.post(
            "/admin/settings",
            data={"action": "ip_username", "ip": "10.0.0.8", "username": "张三"},
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(response.headers["Location"].endswith("/admin/settings?tab=ip_users"))
        with self.app.app_context():
            self.assertEqual(get_ip_username("10.0.0.8"), "张三")

        settings_response = self.client.get("/admin/settings")
        settings_html = settings_response.get_data(as_text=True)
        self.assertIn("IP 用户标记", settings_html)
        self.assertNotIn("张三", settings_html)

        ip_tab_response = self.client.get("/admin/settings?tab=ip_users")
        ip_tab_html = ip_tab_response.get_data(as_text=True)
        ip_tab_soup = BeautifulSoup(ip_tab_html, "html.parser")
        row_form = _required_tag(ip_tab_soup.select_one(".settings-ip-row-form"))
        self.assertIn("张三", ip_tab_html)
        self.assertNotIn("系统出站网络", ip_tab_html)
        self.assertIsNone(row_form.find("button"))
        self.assertIsNotNone(row_form.find("input", {"data-ip-username-input": ""}))

        json_response = self.client.post(
            "/admin/settings",
            data={"action": "ip_username", "ip": "10.0.0.8", "username": "李四"},
            headers={"X-Requested-With": "fetch"},
        )
        self.assertEqual(json_response.status_code, 200)
        self.assertEqual(json_response.get_json(), {"ok": True, "ip": "10.0.0.8", "username": "李四"})

    def test_admin_settings_shows_ip_username_mapping_in_local_ip_mode(self):
        self.app.config["PLATFORM"] = False
        self._insert_task(ip="10.0.0.8")

        settings_response = self.client.get("/admin/settings")
        ip_tab_response = self.client.get("/admin/settings?tab=ip_users")

        self.assertEqual(settings_response.status_code, 200)
        self.assertIn("IP 用户标记", settings_response.get_data(as_text=True))
        self.assertEqual(ip_tab_response.status_code, 200)
        ip_tab_html = ip_tab_response.get_data(as_text=True)
        self.assertIn("IP 用户标记", ip_tab_html)
        self.assertIn("10.0.0.8", ip_tab_html)
        self.assertNotIn("系统出站网络", ip_tab_html)

    def test_admin_settings_hides_ip_username_mapping_outside_ip_mode(self):
        self.app.config["AUTH"] = {
            "mode": "trusted_header",
            "trusted_header": {
                "user_id": "X-SSO-User-Id",
                "username": "X-SSO-User-Name",
            },
        }

        response = self.client.get("/admin/settings")
        blocked = self.client.post(
            "/admin/settings",
            data={"action": "ip_username", "ip": "10.0.0.8", "username": "张三"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertNotIn("IP 用户标记", response.get_data(as_text=True))
        self.assertEqual(blocked.status_code, 404)
        with self.app.app_context():
            self.assertEqual(get_ip_username("10.0.0.8"), "")

        self.app.config["AUTH"] = _saml_auth_config()
        saml_response = self.client.get("/admin/settings")
        self.assertEqual(saml_response.status_code, 200)
        self.assertNotIn("IP 用户标记", saml_response.get_data(as_text=True))

    def test_admin_settings_shows_document_and_consistency_prompt_groups(self):
        response = self.client.get("/admin/settings")

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        soup = BeautifulSoup(html, "html.parser")
        self.assertIn("单文档检查-提示词设置", html)
        self.assertIn("多文档对照检查-提示词设置", html)
        self.assertIn("跨语种文档一致性检查-提示词设置", html)
        self.assertIn("图片检查-提示词设置", html)
        self.assertIn("视频检查-提示词设置", html)
        self.assertIn("consistency_check", html)
        self.assertIn("language_consistency_check", html)
        self.assertIn("image_check", html)
        self.assertIn("video_check", html)
        document_tip = _required_tag(soup.find("button", {"aria-label": "单文档检查-提示词设置说明"}))
        consistency_tip = _required_tag(soup.find("button", {"aria-label": "多文档对照检查-提示词设置说明"}))
        language_tip = _required_tag(soup.find("button", {"aria-label": "跨语种文档一致性检查-提示词设置说明"}))
        image_tip = _required_tag(soup.find("button", {"aria-label": "图片检查-提示词设置说明"}))
        video_tip = _required_tag(soup.find("button", {"aria-label": "视频检查-提示词设置说明"}))
        self.assertEqual(document_tip.get("data-tip"), "内置检查项不可删除；扩展检查项可新增、停用或删除。")
        self.assertEqual(consistency_tip.get("data-tip"), "内置检查项不可删除；扩展检查项可新增、停用或删除，提交多文档对照任务时可多选。")
        self.assertEqual(language_tip.get("data-tip"), "内置检查项不可删除；扩展检查项可新增、停用或删除，提交跨语种检查任务时可多选。")
        self.assertEqual(image_tip.get("data-tip"), "内置检查项不可删除；扩展检查项可新增、停用或删除，提交图片检查任务时可多选。")
        self.assertEqual(video_tip.get("data-tip"), "内置检查项不可删除；扩展检查项可新增、停用或删除，提交视频检查任务时可多选。")
        visible_descriptions = [item.get_text(strip=True) for item in soup.select(".settings-section-head p")]
        self.assertNotIn("内置检查项不可删除；扩展检查项可新增、停用或删除。", visible_descriptions)
        self.assertNotIn("内置检查项不可删除；扩展检查项可新增、停用或删除，提交多文档对照任务时可多选。", visible_descriptions)
        self.assertNotIn("内置检查项不可删除；扩展检查项可新增、停用或删除，提交跨语种检查任务时可多选。", visible_descriptions)
        self.assertNotIn("内置检查项不可删除；扩展检查项可新增、停用或删除，提交图片检查任务时可多选。", visible_descriptions)
        self.assertNotIn("内置检查项不可删除；扩展检查项可新增、停用或删除，提交视频检查任务时可多选。", visible_descriptions)

    def test_admin_overview_counts_tasks_in_selected_range(self):
        self._insert_task(ip="10.0.0.1", username_snapshot="测试用户A", created_at="2026-05-01 10:00:00")
        self._insert_task(
            task_type=CONSISTENCY_TASK_TYPE,
            ip="10.0.0.1",
            username_snapshot="测试用户A",
            status="failed",
            created_at="2026-05-01 11:00:00",
        )
        self._insert_task(ip="10.0.0.2", status="queued", created_at="2026-05-02 08:00:00")
        self._insert_task(task_type=LANGUAGE_CONSISTENCY_TASK_TYPE, ip="10.0.0.2", created_at="2026-05-02 09:00:00")
        self._insert_task(task_type=VIDEO_TASK_TYPE, ip="10.0.0.2", created_at="2026-05-02 10:00:00")
        self._insert_task(ip="10.0.0.3", created_at="2026-04-30 23:59:59")

        response = self.client.get("/admin?start_date=2026-05-01&end_date=2026-05-02")

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn("统计概览", html)
        self.assertNotIn("平台统计", html)
        self.assertNotIn("2026-05-01 至 2026-05-02", html)
        self.assertIn("<span>活跃用户</span><strong>2</strong>", html)
        self.assertIn("<span>提交任务</span><strong>5</strong>", html)
        self.assertIn("<span>单文档检查任务</span><strong>2</strong>", html)
        self.assertIn("<span>多文档对照任务</span><strong>1</strong>", html)
        self.assertIn("<span>跨语种检查任务</span><strong>1</strong>", html)
        self.assertIn("<span>视频检查任务</span><strong>1</strong>", html)
        self.assertIn("<span>排队</span><strong>1</strong>", html)
        self.assertIn("<span>失败</span><strong>1</strong>", html)
        self.assertIn("测试用户A", html)
        self.assertIn("10.0.0.2", html)
        self.assertNotIn("10.0.0.3", html)

    def test_admin_overview_uses_ip_username_mapping(self):
        self._insert_task(ip="10.0.0.8", created_at="2026-05-01 10:00:00")
        self.client.post(
            "/admin/settings",
            data={"action": "ip_username", "ip": "10.0.0.8", "username": "张三"},
        )

        response = self.client.get("/admin?start_date=2026-05-01&end_date=2026-05-01")

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn("张三", html)
        self.assertIn("IP 10.0.0.8", html)
        self.assertNotIn("ip:10.0.0.8", html)

    def test_admin_task_owner_cell_avoids_duplicate_ip_metadata(self):
        self._insert_task(ip="127.0.0.1")

        response = self.client.get("/admin/tasks")

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        soup = BeautifulSoup(html, "html.parser")
        owner_cell = _required_tag(soup.select_one(".task-owner-cell"))
        self.assertEqual(owner_cell.get_text(" ", strip=True), "127.0.0.1")
        self.assertNotIn("ip:127.0.0.1 · IP 127.0.0.1", html)

    def test_admin_task_report_totals_are_cached_until_result_changes(self):
        task_id = self._insert_task()
        report = [
            {
                "code": "typo",
                "name": "错别字检查",
                "result": json.dumps(
                    {
                        "summary": "发现一项问题",
                        "items": [
                            {
                                "status": "issue",
                                "severity": "low",
                                "confidence": "high",
                                "category": "拼写错误",
                                "location": "第1页",
                                "excerpt": "示例",
                                "description": "存在错字",
                                "impact": "影响阅读",
                                "suggestion": "修改",
                            }
                        ],
                    },
                    ensure_ascii=False,
                ),
            }
        ]
        with self.app.app_context():
            get_db().execute(
                "UPDATE tasks SET result_json = ?, updated_at = '2026-05-01 10:01:00' WHERE id = ?",
                (json.dumps(report, ensure_ascii=False), task_id),
            )
            get_db().commit()

        with patch("app.routes._parse_result_json", wraps=_parse_result_json) as parser:
            first = self.client.get("/admin/tasks")
            second = self.client.get("/admin/tasks")

        self.assertEqual(first.status_code, 200)
        self.assertEqual(second.status_code, 200)
        self.assertEqual(parser.call_count, 1)
        with self.app.app_context():
            cached = get_db().execute(
                "SELECT issue_count, source_updated_at FROM task_report_stats WHERE task_id = ?",
                (task_id,),
            ).fetchone()
        self.assertEqual(cached["issue_count"], 1)
        self.assertEqual(cached["source_updated_at"], "2026-05-01 10:01:00")

    def test_task_result_update_invalidates_cached_report_totals(self):
        task_id = self._insert_task()
        with self.app.app_context():
            get_db().execute(
                """
                INSERT INTO task_report_stats(task_id, source_updated_at, suppression_version, updated_at)
                VALUES (?, '2026-05-01 10:00:00', '0:0:', '2026-05-01 10:00:00')
                """,
                (task_id,),
            )
            get_db().execute("UPDATE tasks SET result_json = '[]' WHERE id = ?", (task_id,))
            get_db().commit()
            cached = get_db().execute(
                "SELECT task_id FROM task_report_stats WHERE task_id = ?",
                (task_id,),
            ).fetchone()
        self.assertIsNone(cached)

    def test_suppression_rule_change_refreshes_cached_report_totals(self):
        task_id = self._insert_task()
        report = [
            {
                "code": "typo",
                "name": "错别字检查",
                "result": json.dumps(
                    {
                        "summary": "发现一项问题",
                        "items": [
                            {
                                "status": "issue",
                                "severity": "low",
                                "confidence": "high",
                                "category": "拼写错误",
                                "location": "第1页",
                                "excerpt": "示例",
                                "description": "存在错字",
                                "impact": "影响阅读",
                                "suggestion": "修改",
                            }
                        ],
                    },
                    ensure_ascii=False,
                ),
            }
        ]
        with self.app.app_context():
            get_db().execute(
                "UPDATE tasks SET result_json = ?, updated_at = '2026-05-01 10:01:00' WHERE id = ?",
                (json.dumps(report, ensure_ascii=False), task_id),
            )
            get_db().commit()
        self.client.get("/admin/tasks")

        with self.app.app_context():
            get_db().execute(
                """
                INSERT INTO report_suppression_rules(
                    task_type, check_code, fingerprint, item_json, reason,
                    enabled, created_at, updated_at
                )
                VALUES (?, 'typo', 'fingerprint-1', ?, '模型误报', 1, ?, ?)
                """,
                (
                    DOCUMENT_TASK_TYPE,
                    json.dumps({"description": "存在错字"}, ensure_ascii=False),
                    "2026-05-01 10:02:00",
                    "2026-05-01 10:02:00",
                ),
            )
            get_db().commit()
        self.client.get("/admin/tasks")

        with self.app.app_context():
            cached = get_db().execute(
                "SELECT issue_count, suppressed_count FROM task_report_stats WHERE task_id = ?",
                (task_id,),
            ).fetchone()
        self.assertEqual(cached["issue_count"], 0)
        self.assertEqual(cached["suppressed_count"], 1)

    def test_admin_task_status_endpoint_returns_lightweight_progress(self):
        task_id = self._insert_task(status="running")
        response = self.client.get(f"/admin/task-statuses?task_type={DOCUMENT_TASK_TYPE}&ids={task_id}")

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["active"])
        self.assertEqual(payload["tasks"], [{"id": task_id, "progress": 100, "status": "running", "status_label": "检查中"}])

    def test_admin_task_page_exposes_lightweight_refresh_metadata(self):
        task_id = self._insert_task(status="running")
        response = self.client.get("/admin/tasks")
        soup = BeautifulSoup(response.get_data(as_text=True), "html.parser")

        stats = _required_tag(soup.select_one('[data-refresh-region="stats"]'))
        task_row = _required_tag(soup.select_one(f'[data-task-id="{task_id}"]'))
        self.assertIn("/admin/task-statuses?task_type=document_check", str(stats.get("data-refresh-url")))
        self.assertEqual(task_row.get("data-task-status"), "running")

    def test_admin_overview_filters_tasks_by_auth_mode(self):
        self._insert_task(
            ip="10.0.0.1",
            owner_subject="ip:10.0.0.1",
            owner_source="ip",
            created_at="2026-05-01 10:00:00",
        )
        self._insert_task(
            ip="10.0.0.2",
            owner_subject="trusted_header:100086",
            owner_name_snapshot="张三",
            owner_source="trusted_header",
            created_at="2026-05-01 11:00:00",
        )
        self._insert_task(
            ip="10.0.0.3",
            owner_subject="saml:100086",
            owner_name_snapshot="李四",
            owner_source="saml",
            created_at="2026-05-01 12:00:00",
        )
        self.app.config["AUTH"] = {
            "mode": "trusted_header",
            "trusted_header": {
                "user_id": "X-SSO-User-Id",
                "username": "X-SSO-User-Name",
            },
        }

        response = self.client.get("/admin?start_date=2026-05-01&end_date=2026-05-01")

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn("<span>活跃用户</span><strong>1</strong>", html)
        self.assertIn("<span>提交任务</span><strong>1</strong>", html)
        self.assertIn("张三", html)
        self.assertNotIn("李四", html)
        self.assertNotIn("ip:10.0.0.1", html)

    def test_local_mode_admin_root_redirects_to_management_view(self):
        self.app.config["PLATFORM"] = False
        self._logout_test_client()

        response = self.client.get("/admin")

        self.assertEqual(response.status_code, 302)
        self.assertTrue(response.headers["Location"].endswith("/"))

    def test_admin_model_route_requires_admin_login(self):
        self._logout_test_client()

        response = self.client.get("/admin/models")

        self.assertEqual(response.status_code, 302)
        self.assertIn("/admin/login", response.headers["Location"])

    def test_admin_model_page_uses_same_identity_as_user_model_page(self):
        self._configure_provider("ip:127.0.0.1")

        admin_response = self.client.get("/admin/models")
        user_response = self.client.get("/models")

        self.assertEqual(admin_response.status_code, 200)
        self.assertEqual(user_response.status_code, 200)
        self.assertIn("测试提供商", admin_response.get_data(as_text=True))
        self.assertIn("测试提供商", user_response.get_data(as_text=True))
        self.assertIn("模型管理", admin_response.get_data(as_text=True))

    def test_admin_model_page_saves_same_user_model_config(self):
        response = self.client.post(
            "/admin/models",
            data={
                "name": "Console 提供商",
                "api_base": "https://example.test/v1/chat/completions",
                "api_key": "",
                "request_timeout": "30",
                "max_input_chars": "80000",
                "is_active": "on",
                "model_configs": json.dumps(
                    [{"model_name": "console-model", "force_disable_thinking": False}],
                    ensure_ascii=False,
                ),
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(response.headers["Location"].endswith("/admin/models"))
        user_response = self.client.get("/models")
        self.assertIn("Console 提供商", user_response.get_data(as_text=True))
        self.assertIn("console-model", user_response.get_data(as_text=True))

    def test_ip_identity_uses_configured_real_ip_header(self):
        self.app.config["REAL_IP_HEADER"] = "X-Real-IP"
        model_id = self._configure_provider("ip:10.20.30.40")
        with self.app.app_context():
            item = get_db().execute("SELECT id FROM check_items WHERE code = 'typo'").fetchone()

        response = self.client.post(
            "/",
            data={
                "document": (io.BytesIO("测试文档".encode("utf-8")), "doc.txt"),
                "checks": [str(item["id"])],
                "model_id": model_id,
            },
            headers={"X-Real-IP": "10.20.30.40"},
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            task = get_db().execute("SELECT owner_subject, ip FROM tasks").fetchone()
        self.assertEqual(task["owner_subject"], "ip:10.20.30.40")
        self.assertEqual(task["ip"], "10.20.30.40")

    def test_invalid_real_ip_header_falls_back_to_remote_addr(self):
        self.app.config["REAL_IP_HEADER"] = "X-Real-IP"
        self._configure_provider("ip:127.0.0.1")

        response = self.client.get("/models", headers={"X-Real-IP": "not-an-ip"})

        self.assertEqual(response.status_code, 200)
        self.assertIn("测试提供商", response.get_data(as_text=True))

    def test_local_mode_root_shows_admin_view_without_login(self):
        self.app.config["PLATFORM"] = False
        self._logout_test_client()

        response = self.client.get("/")

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn("单文档检查任务", html)
        self.assertIn("模型管理", html)
        self.assertNotIn("用户管理", html)
        self.assertNotIn("退出", html)

    def test_local_mode_user_model_page_does_not_require_login(self):
        self.app.config["PLATFORM"] = False
        self._logout_test_client()

        response = self.client.get("/models")

        self.assertEqual(response.status_code, 200)
        self.assertIn("我的模型", response.get_data(as_text=True))

    def test_user_management_route_is_not_registered(self):
        platform_response = self.client.get("/admin/users")

        self.app.config["PLATFORM"] = False
        self._logout_test_client()
        local_response = self.client.get("/admin/users")

        self.assertEqual(platform_response.status_code, 404)
        self.assertEqual(local_response.status_code, 404)

    def test_user_models_saves_model_force_disable_thinking(self):
        response = self.client.post(
            "/models",
            data={
                "name": "测试提供商",
                "api_base": "https://example.test/v1/chat/completions",
                "api_key": "",
                "request_timeout": "30",
                "max_input_chars": "80000",
                "is_active": "on",
                "model_configs": json.dumps(
                    [
                        {"model_name": "model-a", "force_disable_thinking": True},
                        {"model_name": "model-b", "force_disable_thinking": False},
                    ],
                    ensure_ascii=False,
                ),
            },
        )

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            models = get_db().execute(
                """
                SELECT m.model_name, m.force_disable_thinking
                FROM user_model_configs m
                JOIN user_model_providers p ON p.id = m.provider_id
                WHERE p.owner_subject = ?
                ORDER BY m.sort_order
                """,
                ("ip:127.0.0.1",),
            ).fetchall()
        self.assertEqual(
            [(row["model_name"], bool(row["force_disable_thinking"])) for row in models],
            [
                ("model-a", True),
                ("model-b", False),
            ],
        )

    def test_user_models_accepts_million_char_input_limit(self):
        response = self.client.post(
            "/models",
            data={
                "name": "测试提供商",
                "api_base": "https://example.test/v1/chat/completions",
                "api_key": "",
                "request_timeout": "30",
                "max_input_chars": "1000000",
                "is_active": "on",
                "model_configs": json.dumps(
                    [{"model_name": "model-a", "force_disable_thinking": False}],
                    ensure_ascii=False,
                ),
            },
        )

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            row = get_db().execute("SELECT max_input_chars FROM user_model_providers").fetchone()
        self.assertEqual(row["max_input_chars"], 1000000)

    def test_user_models_defaults_input_limit_to_five_hundred_thousand(self):
        response = self.client.post(
            "/models",
            data={
                "name": "默认上限提供商",
                "api_base": "https://example.test/v1/chat/completions",
                "api_key": "",
                "request_timeout": "30",
                "is_active": "on",
                "model_configs": json.dumps(
                    [{"model_name": "model-a", "force_disable_thinking": True}],
                    ensure_ascii=False,
                ),
            },
        )

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            row = get_db().execute("SELECT max_input_chars FROM user_model_providers").fetchone()
        self.assertEqual(row["max_input_chars"], 500000)

    def test_user_models_defaults_legacy_model_config_to_disable_thinking(self):
        response = self.client.post(
            "/models",
            data={
                "name": "旧格式提供商",
                "api_base": "https://example.test/v1/chat/completions",
                "api_key": "",
                "request_timeout": "30",
                "is_active": "on",
                "model_configs": json.dumps([{"model_name": "legacy-model"}], ensure_ascii=False),
            },
        )

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            row = get_db().execute(
                "SELECT force_disable_thinking FROM user_model_configs WHERE model_name = 'legacy-model'"
            ).fetchone()
        self.assertEqual(row["force_disable_thinking"], 1)

    def test_user_models_allows_same_name_for_distinct_thinking_modes(self):
        response = self.client.post(
            "/models",
            data={
                "name": "测试提供商",
                "api_base": "https://example.test/v1/chat/completions",
                "api_key": "",
                "request_timeout": "30",
                "max_input_chars": "80000",
                "is_active": "on",
                "model_configs": json.dumps(
                    [
                        {"model_name": "same-model", "force_disable_thinking": False},
                        {"model_name": "same-model", "force_disable_thinking": True},
                        {"model_name": "same-model", "force_disable_thinking": False},
                    ],
                    ensure_ascii=False,
                ),
            },
        )

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            saved_models = get_db().execute(
                """
                SELECT m.model_name, m.force_disable_thinking
                FROM user_model_configs m
                ORDER BY m.sort_order
                """
            ).fetchall()
        self.assertEqual(
            [(row["model_name"], bool(row["force_disable_thinking"])) for row in saved_models],
            [
                ("same-model", False),
                ("same-model", True),
            ],
        )

        with self.app.app_context():
            models = get_enabled_models("ip:127.0.0.1")
            self.assertEqual(len(models), 2)
            self.assertEqual(len({model["id"] for model in models}), 2)
            by_mode = {model["force_disable_thinking"]: model for model in models}
            thinking_enabled_model = _find_enabled_model(by_mode[False]["id"], "ip:127.0.0.1")
            thinking_disabled_model = _find_enabled_model(by_mode[True]["id"], "ip:127.0.0.1")
            assert thinking_enabled_model is not None
            assert thinking_disabled_model is not None
            self.assertFalse(thinking_enabled_model["force_disable_thinking"])
            self.assertTrue(thinking_disabled_model["force_disable_thinking"])

    def test_user_model_test_endpoint_uses_submitted_model_config(self):
        self.app.config["NETWORK"] = {
            "proxy_mode": "custom",
            "proxy": "http://127.0.0.1:7890",
            "ssl_verify": True,
        }
        with patch("app.routes.test_model_connection", return_value="模型连通性测试通过。") as mocked_test:
            response = self.client.post(
                "/models/test",
                json={
                    "api_base": "https://example.test/v1/chat/completions",
                    "api_key": "sk-test",
                    "request_timeout": "30",
                    "model_name": "model-a",
                    "force_disable_thinking": True,
                },
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json(), {"ok": True, "message": "模型连通性测试通过。"})
        mocked_test.assert_called_once_with(
            api_base="https://example.test/v1/chat/completions",
            api_key="sk-test",
            proxy_mode="custom",
            proxy="http://127.0.0.1:7890",
            ssl_verify=True,
            request_timeout=30,
            model_name="model-a",
            force_disable_thinking=True,
        )

    def test_user_fetch_models_uses_system_network_config(self):
        self.app.config["NETWORK"] = {
            "proxy_mode": "system",
            "proxy": "",
            "ssl_verify": True,
        }
        with patch("app.routes.fetch_models", return_value=["model-a"]) as mocked_fetch:
            response = self.client.post(
                "/models/fetch",
                json={
                    "api_base": "https://example.test/v1/chat/completions",
                    "api_key": "sk-test",
                    "request_timeout": "30",
                },
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json(), {"fetched_models": ["model-a"], "fetched_count": 1})
        mocked_fetch.assert_called_once_with(
            api_base="https://example.test/v1/chat/completions",
            api_key="sk-test",
            proxy_mode="system",
            proxy="",
            ssl_verify=True,
            request_timeout=30,
        )

    def test_user_fetch_models_rejects_get_query_parameters(self):
        response = self.client.get(
            "/models/fetch",
            query_string={"api_base": "https://example.test/v1/chat/completions", "api_key": "secret"},
        )

        self.assertEqual(response.status_code, 405)

    def test_admin_settings_creates_consistency_check_item(self):
        response = self.client.post(
            "/admin/settings",
            data={
                "action": "create_check_item",
                "task_type": CONSISTENCY_TASK_TYPE,
                "name": "遗漏内容检查",
                "description": "检查资料是否遗漏素材关键内容",
                "prompt": "只检查资料遗漏。",
                "enabled": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            item = get_db().execute(
                """
                SELECT task_type, code, name, description, prompt, enabled
                FROM check_items
                WHERE name = ?
                """,
                ("遗漏内容检查",),
            ).fetchone()
        self.assertEqual(item["task_type"], CONSISTENCY_TASK_TYPE)
        self.assertTrue(item["code"].startswith("custom-consistency-"))
        self.assertEqual(item["description"], "检查资料是否遗漏素材关键内容")
        self.assertEqual(item["prompt"], "只检查资料遗漏。")
        self.assertEqual(item["enabled"], 1)

    def test_admin_settings_creates_image_check_item(self):
        response = self.client.post(
            "/admin/settings",
            data={
                "action": "create_check_item",
                "task_type": IMAGE_TASK_TYPE,
                "name": "接线颜色检查",
                "description": "检查线缆颜色是否符合图纸要求",
                "prompt": "只检查接线颜色。",
                "enabled": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            item = get_db().execute(
                """
                SELECT task_type, code, name, description, prompt, enabled
                FROM check_items
                WHERE name = ?
                """,
                ("接线颜色检查",),
            ).fetchone()
        self.assertEqual(item["task_type"], IMAGE_TASK_TYPE)
        self.assertTrue(item["code"].startswith("custom-image-"))
        self.assertEqual(item["description"], "检查线缆颜色是否符合图纸要求")
        self.assertEqual(item["prompt"], "只检查接线颜色。")
        self.assertEqual(item["enabled"], 1)

    def test_admin_settings_creates_video_check_item(self):
        response = self.client.post(
            "/admin/settings",
            data={
                "action": "create_check_item",
                "task_type": VIDEO_TASK_TYPE,
                "name": "铭牌信息检查",
                "description": "检查视频中设备铭牌是否清晰",
                "prompt": "只检查铭牌清晰度。",
                "enabled": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            item = get_db().execute(
                """
                SELECT task_type, code, name, description, prompt, enabled
                FROM check_items
                WHERE name = ?
                """,
                ("铭牌信息检查",),
            ).fetchone()
        self.assertEqual(item["task_type"], VIDEO_TASK_TYPE)
        self.assertTrue(item["code"].startswith("custom-video-"))
        self.assertEqual(item["description"], "检查视频中设备铭牌是否清晰")
        self.assertEqual(item["prompt"], "只检查铭牌清晰度。")
        self.assertEqual(item["enabled"], 1)

    def test_upload_destination_uses_unique_name_for_same_second_uploads(self):
        with self.app.app_context():
            first_name, _ = _upload_destination("报告.txt", "127.0.0.1", "2026-05-22 12:00:00", "txt")
            second_name, _ = _upload_destination("报告.txt", "127.0.0.1", "2026-05-22 12:00:00", "txt")

        self.assertNotEqual(first_name, second_name)
        self.assertTrue(first_name.endswith(".txt"))
        self.assertTrue(second_name.endswith(".txt"))

    def test_create_task_rejects_disabled_check_item_before_saving_file(self):
        model_id = self._configure_provider()
        with self.app.app_context():
            item = get_db().execute("SELECT id FROM check_items WHERE code = 'typo'").fetchone()
            get_db().execute("UPDATE check_items SET enabled = 0 WHERE id = ?", (item["id"],))
            get_db().commit()

        response = self.client.post(
            "/",
            data={
                "document": (io.BytesIO("测试文档".encode("utf-8")), "doc.txt"),
                "checks": [str(item["id"])],
                "model_id": model_id,
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            total = get_db().execute("SELECT COUNT(*) AS total FROM tasks").fetchone()["total"]
        self.assertEqual(total, 0)
        self.assertEqual(list(Path(self.app.config["UPLOAD_FOLDER"]).iterdir()), [])

    def test_document_task_form_allows_multiple_uploads(self):
        self._configure_provider()

        response = self.client.get("/")

        self.assertEqual(response.status_code, 200)
        soup = BeautifulSoup(response.get_data(as_text=True), "html.parser")
        form = _required_tag(soup.find("form", {"data-prevent-double-submit": "true"}))
        self.assertEqual(form.get("data-submitting-label"), "提交中...")
        self.assertIn("请勿重复提交", form.get("data-submitting-message", ""))
        progress = _required_tag(form.select_one("[data-submit-progress]"))
        self.assertTrue(progress.has_attr("hidden"))
        self.assertIn("请勿重复提交", progress.get_text(strip=True))
        upload = _required_tag(soup.find("input", {"name": "document"}))
        self.assertTrue(upload.has_attr("multiple"))
        self.assertIsNone(upload.get("data-file-limit"))
        field = _required_tag(upload.find_parent(class_="multi-file-field"))
        self.assertIsNotNone(field.select_one("[data-file-list]"))

    def test_create_task_saves_check_snapshot_and_extracted_text(self):
        model_id = self._configure_provider()
        with self.app.app_context():
            item = get_db().execute("SELECT id, code, name, prompt FROM check_items WHERE code = 'typo'").fetchone()

        response = self.client.post(
            "/",
            data={
                "document": (io.BytesIO("测试文档".encode("utf-8")), "doc.txt"),
                "checks": [str(item["id"])],
                "model_id": model_id,
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            task = get_db().execute("SELECT * FROM tasks").fetchone()
        snapshots = json.loads(task["checks_snapshot_json"])
        self.assertEqual(task["document_text"], "file: doc.txt\n\n测试文档")
        self.assertEqual(
            snapshots,
            [
                {
                    "id": item["id"],
                    "code": item["code"],
                    "name": item["name"],
                    "prompt": item["prompt"],
                }
            ],
        )

    def test_create_task_saves_long_filename_when_upload_folder_is_missing(self):
        model_id = self._configure_provider()
        with self.app.app_context():
            item = get_db().execute("SELECT id FROM check_items WHERE code = 'typo'").fetchone()

        upload_folder = Path(self.app.config["ROOT_DIR"]) / ("deep-" + "x" * 50) / "uploads"
        self.app.config["UPLOAD_FOLDER"] = str(upload_folder)
        if upload_folder.exists():
            shutil.rmtree(upload_folder)
        filename = (
            "PowerCube 1000 V300R008C10 Installation Guide(Site Construction,"
            "ICC330-HA1-C11,ICC330-HD1-C6,ICC360-HA1-C2,ICC800-A1-C2)_运营商专用.pdf"
        )

        response = self.client.post(
            "/",
            data={
                "document": (_pdf_with_image_bytes(), filename),
                "checks": [str(item["id"])],
                "model_id": model_id,
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            task = get_db().execute("SELECT original_filename, stored_filename, document_text FROM tasks").fetchone()
        stored_path = upload_folder / task["stored_filename"]
        self.assertTrue(stored_path.is_file())
        self.assertLessEqual(len(str(stored_path.resolve())), UPLOAD_PATH_SAFE_CHARS)
        self.assertEqual(task["original_filename"], filename)
        self.assertTrue(task["document_text"].startswith(f"file: {filename}\n\n"))
        self.assertIn("[第1页]", task["document_text"])

    def test_create_task_creates_one_task_per_uploaded_document(self):
        model_id = self._configure_provider()
        with self.app.app_context():
            item = get_db().execute(
                "SELECT id, code, name, prompt FROM check_items WHERE code = 'typo'"
            ).fetchone()

        response = self.client.post(
            "/",
            data={
                "document": [
                    (io.BytesIO(f"document {index}".encode("utf-8")), f"doc-{index:02d}.txt")
                    for index in range(21)
                ],
                "checks": [str(item["id"])],
                "model_id": model_id,
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            tasks = get_db().execute(
                """
                SELECT original_filename, document_text, status, checks_snapshot_json
                FROM tasks
                ORDER BY original_filename ASC
                """
            ).fetchall()
            uploaded_files = list(Path(self.app.config["UPLOAD_FOLDER"]).iterdir())

        self.assertEqual(len(tasks), 21)
        self.assertEqual(tasks[0]["original_filename"], "doc-00.txt")
        self.assertEqual(tasks[-1]["original_filename"], "doc-20.txt")
        self.assertTrue(all(task["status"] == "queued" for task in tasks))
        self.assertEqual(tasks[0]["document_text"], "file: doc-00.txt\n\ndocument 0")
        self.assertEqual(tasks[-1]["document_text"], "file: doc-20.txt\n\ndocument 20")
        self.assertEqual(len(uploaded_files), 21)
        snapshots = [json.loads(task["checks_snapshot_json"]) for task in tasks]
        self.assertTrue(all(snapshot[0]["code"] == item["code"] for snapshot in snapshots))

    def test_create_task_rejects_entire_batch_when_one_document_has_no_text(self):
        model_id = self._configure_provider()
        with self.app.app_context():
            item = get_db().execute("SELECT id FROM check_items WHERE code = 'typo'").fetchone()

        response = self.client.post(
            "/",
            data={
                "document": [
                    (io.BytesIO(b"valid document"), "valid.txt"),
                    (io.BytesIO(b"   "), "blank.txt"),
                ],
                "checks": [str(item["id"])],
                "model_id": model_id,
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            total = get_db().execute("SELECT COUNT(*) AS total FROM tasks").fetchone()["total"]
            uploaded_files = list(Path(self.app.config["UPLOAD_FOLDER"]).iterdir())
        self.assertEqual(total, 0)
        self.assertEqual(uploaded_files, [])

    def test_create_task_handles_unexpected_upload_preparation_error(self):
        model_id = self._configure_provider()
        with self.app.app_context():
            item = get_db().execute("SELECT id FROM check_items WHERE code = 'typo'").fetchone()

        with patch("app.routes.extract_text", side_effect=ValueError("company parser failed")):
            response = self.client.post(
                "/",
                data={
                    "document": (io.BytesIO("测试文档".encode("utf-8")), "doc.txt"),
                    "checks": [str(item["id"])],
                    "model_id": model_id,
                },
                content_type="multipart/form-data",
                follow_redirects=True,
            )

        self.assertEqual(response.status_code, 200)
        self.assertIn("文档上传或读取失败", response.get_data(as_text=True))
        self.assertIn("company parser failed", response.get_data(as_text=True))
        with self.app.app_context():
            total = get_db().execute("SELECT COUNT(*) AS total FROM tasks").fetchone()["total"]
            uploaded_files = list(Path(self.app.config["UPLOAD_FOLDER"]).iterdir())
        self.assertEqual(total, 0)
        self.assertEqual(uploaded_files, [])

    def test_create_task_removes_partial_file_when_upload_save_fails(self):
        model_id = self._configure_provider()
        with self.app.app_context():
            item = get_db().execute("SELECT id FROM check_items WHERE code = 'typo'").fetchone()

        def fail_after_partial_write(_upload, destination, buffer_size=16384):
            Path(destination).write_bytes(b"partial")
            raise OSError("disk write failed")

        with patch.object(FileStorage, "save", autospec=True, side_effect=fail_after_partial_write):
            response = self.client.post(
                "/",
                data={
                    "document": (io.BytesIO("测试文档".encode("utf-8")), "doc.txt"),
                    "checks": [str(item["id"])],
                    "model_id": model_id,
                },
                content_type="multipart/form-data",
                follow_redirects=True,
            )

        self.assertEqual(response.status_code, 200)
        self.assertIn("文档上传或读取失败", response.get_data(as_text=True))
        with self.app.app_context():
            total = get_db().execute("SELECT COUNT(*) AS total FROM tasks").fetchone()["total"]
        self.assertEqual(total, 0)
        self.assertEqual(list(Path(self.app.config["UPLOAD_FOLDER"]).iterdir()), [])

    def test_create_image_task_saves_extracted_image_metadata(self):
        model_id = self._configure_provider()
        with self.app.app_context():
            item = get_db().execute(
                "SELECT id, code, name, prompt FROM check_items WHERE code = 'image-small-language-text'"
            ).fetchone()

        response = self.client.post(
            "/images",
            data={
                "document": (_pdf_with_image_bytes(), "diagram.pdf"),
                "checks": [str(item["id"])],
                "model_id": model_id,
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            task = get_db().execute("SELECT * FROM tasks").fetchone()
            image_root = Path(self.app.config["UPLOAD_FOLDER"]).parent / "extracted_images"
        meta = json.loads(task["document_meta_json"])
        snapshots = json.loads(task["checks_snapshot_json"])
        self.assertEqual(task["task_type"], IMAGE_TASK_TYPE)
        self.assertEqual(meta["source_document"]["file_type"], "pdf")
        self.assertEqual(len(meta["page_images"]), 1)
        self.assertIn("page001-screenshot", meta["page_images"][0]["filename"])
        self.assertTrue((image_root / meta["page_images"][0]["relative_path"]).is_file())
        self.assertIn("document_text:", task["document_text"])
        self.assertIn("extracted_images:", task["document_text"])
        self.assertIn("page_screenshots: 1", task["document_text"])
        self.assertEqual(
            snapshots,
            [
                {
                    "id": item["id"],
                    "code": item["code"],
                    "name": item["name"],
                    "prompt": item["prompt"],
                }
            ],
        )

    def test_create_image_task_removes_files_when_database_insert_fails(self):
        model_id = self._configure_provider()
        with self.app.app_context():
            item = get_db().execute(
                "SELECT id FROM check_items WHERE code = 'image-small-language-text'"
            ).fetchone()
        self._reject_task_inserts()

        response = self.client.post(
            "/images",
            data={
                "document": (_pdf_with_image_bytes(), "diagram.pdf"),
                "checks": [str(item["id"])],
                "model_id": model_id,
            },
            content_type="multipart/form-data",
            follow_redirects=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn("创建图片检查任务失败", response.get_data(as_text=True))
        with self.app.app_context():
            total = get_db().execute("SELECT COUNT(*) AS total FROM tasks").fetchone()["total"]
        image_root = Path(self.app.config["UPLOAD_FOLDER"]).parent / "extracted_images"
        self.assertEqual(total, 0)
        self.assertEqual(list(Path(self.app.config["UPLOAD_FOLDER"]).iterdir()), [])
        self.assertFalse(image_root.exists() and any(image_root.rglob("*")))

    def test_create_image_task_removes_partial_embedded_images_before_page_fallback(self):
        model_id = self._configure_provider()
        with self.app.app_context():
            item = get_db().execute(
                "SELECT id FROM check_items WHERE code = 'image-small-language-text'"
            ).fetchone()

        def fail_after_partial_image(_document_path, _file_type, output_dir, *, source_filename=""):
            output_dir.mkdir(parents=True, exist_ok=True)
            (output_dir / "partial-embedded.png").write_bytes(_TINY_PNG)
            raise DocumentReadError("embedded image stream is damaged")

        with patch("app.routes.extract_images", side_effect=fail_after_partial_image):
            response = self.client.post(
                "/images",
                data={
                    "document": (_pdf_with_image_bytes(), "diagram.pdf"),
                    "checks": [str(item["id"])],
                    "model_id": model_id,
                },
                content_type="multipart/form-data",
            )

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            task = get_db().execute("SELECT document_meta_json FROM tasks").fetchone()
        image_root = Path(self.app.config["UPLOAD_FOLDER"]).parent / "extracted_images"
        meta = json.loads(task["document_meta_json"])
        generated_files = [path for path in image_root.rglob("*") if path.is_file()]
        self.assertEqual(meta["images"], [])
        self.assertEqual(len(meta["page_images"]), 1)
        self.assertEqual([path.name for path in generated_files], [meta["page_images"][0]["filename"]])

    def test_create_image_task_rejects_non_pdf_document(self):
        model_id = self._configure_provider()
        with self.app.app_context():
            item = get_db().execute(
                "SELECT id FROM check_items WHERE code = 'image-small-language-text'"
            ).fetchone()

        response = self.client.post(
            "/images",
            data={
                "document": (io.BytesIO(b"<html></html>"), "diagram.html"),
                "checks": [str(item["id"])],
                "model_id": model_id,
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            total = get_db().execute("SELECT COUNT(*) AS total FROM tasks").fetchone()["total"]
        self.assertEqual(total, 0)

    def test_create_video_task_saves_extracted_frame_metadata(self):
        model_id = self._configure_provider()
        with self.app.app_context():
            item = get_db().execute(
                "SELECT id, code, name, prompt FROM check_items WHERE code = 'video-installation-sequence'"
            ).fetchone()

        def fake_extract_video_frames(video_path, output_dir, *, source_filename="", max_frames=16):
            output_dir.mkdir(parents=True, exist_ok=True)
            frame_path = output_dir / "0001_t000001000.jpg"
            frame_path.write_bytes(_TINY_PNG)
            return (
                [
                    {
                        "id": "frame-0001",
                        "filename": "0001_t000001000.jpg",
                        "stored_filename": "0001_t000001000.jpg",
                        "relative_path": "0001_t000001000.jpg",
                        "mime_type": "image/jpeg",
                        "position": "00:01.000",
                        "source": source_filename,
                        "size_bytes": frame_path.stat().st_size,
                        "kind": "video_frame",
                        "timestamp_seconds": 1.0,
                    }
                ],
                {
                    "duration_seconds": 8.0,
                    "selected_timestamps": [1.0],
                    "max_frames": max_frames,
                    "frame_count": 1,
                    "strategy": "uniform-sampling",
                },
            )

        with patch("app.routes.extract_video_frames", side_effect=fake_extract_video_frames):
            response = self.client.post(
                "/videos",
                data={
                    "video": (io.BytesIO(b"video-bytes"), "install.mp4"),
                    "checks": [str(item["id"])],
                    "model_id": model_id,
                },
                content_type="multipart/form-data",
            )

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            task = get_db().execute("SELECT * FROM tasks").fetchone()
            image_root = Path(self.app.config["UPLOAD_FOLDER"]).parent / "extracted_images"
        meta = json.loads(task["document_meta_json"])
        snapshots = json.loads(task["checks_snapshot_json"])
        self.assertEqual(task["task_type"], VIDEO_TASK_TYPE)
        self.assertEqual(task["file_type"], "mp4")
        self.assertEqual(meta["source_video"]["file_type"], "mp4")
        self.assertEqual(meta["frame_selection"]["frame_count"], 1)
        self.assertEqual(len(meta["frames"]), 1)
        self.assertEqual(meta["frames"][0]["position"], "00:01.000")
        self.assertTrue((image_root / meta["frames"][0]["relative_path"]).is_file())
        self.assertIn("video_context:", task["document_text"])
        self.assertIn("video_frames:", task["document_text"])
        self.assertIn("00:01.000", task["document_text"])
        self.assertEqual(
            snapshots,
            [
                {
                    "id": item["id"],
                    "code": item["code"],
                    "name": item["name"],
                    "prompt": item["prompt"],
                }
            ],
        )

    def test_create_video_task_removes_files_when_database_insert_fails(self):
        model_id = self._configure_provider()
        with self.app.app_context():
            item = get_db().execute(
                "SELECT id FROM check_items WHERE code = 'video-installation-sequence'"
            ).fetchone()
        self._reject_task_inserts()

        def fake_extract_video_frames(video_path, output_dir, *, source_filename="", max_frames=16):
            output_dir.mkdir(parents=True, exist_ok=True)
            frame_path = output_dir / "0001_t000001000.jpg"
            frame_path.write_bytes(_TINY_PNG)
            return (
                [
                    {
                        "filename": frame_path.name,
                        "mime_type": "image/jpeg",
                        "position": "00:01.000",
                        "size_bytes": frame_path.stat().st_size,
                    }
                ],
                {"frame_count": 1, "max_frames": max_frames},
            )

        with patch("app.routes.extract_video_frames", side_effect=fake_extract_video_frames):
            response = self.client.post(
                "/videos",
                data={
                    "video": (io.BytesIO(b"video-bytes"), "install.mp4"),
                    "checks": [str(item["id"])],
                    "model_id": model_id,
                },
                content_type="multipart/form-data",
                follow_redirects=True,
            )

        self.assertEqual(response.status_code, 200)
        self.assertIn("创建视频检查任务失败", response.get_data(as_text=True))
        with self.app.app_context():
            total = get_db().execute("SELECT COUNT(*) AS total FROM tasks").fetchone()["total"]
        image_root = Path(self.app.config["UPLOAD_FOLDER"]).parent / "extracted_images"
        self.assertEqual(total, 0)
        self.assertEqual(list(Path(self.app.config["UPLOAD_FOLDER"]).iterdir()), [])
        self.assertFalse(image_root.exists() and any(image_root.rglob("*")))

    def test_create_video_task_rejects_unsupported_file(self):
        model_id = self._configure_provider()
        with self.app.app_context():
            item = get_db().execute(
                "SELECT id FROM check_items WHERE code = 'video-installation-sequence'"
            ).fetchone()

        response = self.client.post(
            "/videos",
            data={
                "video": (io.BytesIO(b"<html></html>"), "install.html"),
                "checks": [str(item["id"])],
                "model_id": model_id,
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            total = get_db().execute("SELECT COUNT(*) AS total FROM tasks").fetchone()["total"]
        self.assertEqual(total, 0)

    def test_oversized_upload_shows_chinese_limit_message(self):
        self.app.config["MAX_UPLOAD_MB"] = 1
        self.app.config["MAX_CONTENT_LENGTH"] = 1

        response = self.client.post(
            "/videos",
            data={"video": (io.BytesIO(b"video-bytes"), "install.mp4")},
            content_type="multipart/form-data",
            follow_redirects=True,
        )

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn("上传文件过大，当前上传上限为 1MB", html)
        self.assertIn("创建视频检查", html)

    def test_image_task_detail_hides_extracted_image_list(self):
        with self.app.app_context():
            now = "2026-05-22 12:00:00"
            meta = {
                "images": [
                    {
                        "filename": "0001_page001-image001.png",
                        "relative_path": "task/0001_page001-image001.png",
                        "mime_type": "image/png",
                        "position": "page001-image001",
                        "size_bytes": 1024,
                    }
                ],
                "page_images": [
                    {
                        "filename": "0001_page001-screenshot.png",
                        "relative_path": "task/0001_page001-screenshot.png",
                        "mime_type": "image/png",
                        "position": "page001-screenshot",
                        "size_bytes": 2048,
                    }
                ],
            }
            result_json = [
                {
                    "code": "image-figure-table-title-standard",
                    "name": "图表标题规范检查",
                    "result": "检查结果正文：page001 表格缺少表标题。",
                }
            ]
            cursor = get_db().execute(
                """
                INSERT INTO tasks(
                    task_type, ip, original_filename, stored_filename, file_type,
                    file_size, document_meta_json, result_json, checks_json, model_name, api_base,
                    status, progress, created_at, updated_at
                )
                VALUES (?, '127.0.0.1', '图纸.pdf', 'stored.pdf', 'pdf',
                        4096, ?, ?, '[]', 'model-a', 'https://example.test/v1/chat/completions',
                        'completed', 100, ?, ?)
                """,
                (
                    IMAGE_TASK_TYPE,
                    json.dumps(meta, ensure_ascii=False),
                    json.dumps(result_json, ensure_ascii=False),
                    now,
                    now,
                ),
            )
            get_db().commit()
            task_id = cursor.lastrowid

        detail = self.client.get(f"/admin/tasks/{task_id}")
        exported = self.client.get(f"/admin/tasks/{task_id}/export")

        self.assertEqual(detail.status_code, 200)
        self.assertEqual(exported.status_code, 200)
        detail_html = detail.get_data(as_text=True)
        exported_html = exported.get_data(as_text=True)
        self.assertNotIn("提取图片", detail_html)
        self.assertNotIn("0001_page001-image001.png", detail_html)
        self.assertNotIn("0001_page001-screenshot.png", detail_html)
        self.assertIn("检查结果正文：page001 表格缺少表标题。", detail_html)
        self.assertNotIn("提取图片", exported_html)
        self.assertNotIn("0001_page001-image001.png", exported_html)
        self.assertNotIn("0001_page001-screenshot.png", exported_html)

    def test_task_detail_renders_report_items_and_counts(self):
        with self.app.app_context():
            now = "2026-05-23 12:00:00"
            result_json = [
                {
                    "code": "consistency",
                    "name": "全文一致性检查",
                    "result": (
                        "总体结论：存在一致性风险。\n\n"
                        "1. 问题类型：参数不一致\n"
                        "位置：第1章、第2章\n"
                        "原文摘录：A 为 10；A 为 20\n"
                        "问题描述：同一参数前后不一致\n"
                        "影响说明：客户可能按错误参数配置\n"
                        "修改建议：统一参数值。\n\n"
                        "2. 建议：补充适用范围\n"
                        "位置：第3章\n"
                        "原文摘录：安装前检查环境\n"
                        "修改建议：补充温度范围。"
                    ),
                }
            ]
            cursor = get_db().execute(
                """
                INSERT INTO tasks(
                    task_type, ip, original_filename, stored_filename, file_type,
                    file_size, result_json, checks_json, model_name, api_base,
                    status, progress, created_at, updated_at
                )
                VALUES (?, '127.0.0.1', 'report.txt', 'stored.txt', 'txt',
                        1024, ?, '[]', 'model-a', 'https://example.test/v1/chat/completions',
                        'completed', 100, ?, ?)
                """,
                (DOCUMENT_TASK_TYPE, json.dumps(result_json, ensure_ascii=False), now, now),
            )
            get_db().commit()
            task_id = cursor.lastrowid

        detail = self.client.get(f"/admin/tasks/{task_id}")
        exported = self.client.get(f"/admin/tasks/{task_id}/export")

        self.assertEqual(detail.status_code, 200)
        self.assertEqual(exported.status_code, 200)
        soup = BeautifulSoup(detail.get_data(as_text=True), "html.parser")
        items = soup.select("[data-report-item]")
        self.assertEqual(len(items), 2)
        self.assertEqual(_required_tag(soup.select_one('[data-report-count="issue"]')).get_text(strip=True), "1")
        self.assertEqual(_required_tag(soup.select_one('[data-report-count="suggestion"]')).get_text(strip=True), "1")
        self.assertEqual(_required_tag(soup.select_one('[data-report-count="non_issue"]')).get_text(strip=True), "0")
        self.assertEqual(_required_tag(soup.select_one('[data-report-count="pending_issue_acceptance"]')).get_text(strip=True), "1")
        self.assertEqual(_required_tag(soup.select_one('[data-report-count="issue_detection_rate"]')).get_text(strip=True), "50.0%")
        self.assertEqual(_required_tag(soup.select_one('[data-report-count="issue_acceptance_rate"]')).get_text(strip=True), "-")
        self.assertIn("AI 检查条目统计", exported.get_data(as_text=True))
        self.assertIn("条目 1", exported.get_data(as_text=True))

    def test_task_report_exports_excel_for_statistics(self):
        with self.app.app_context():
            now = "2026-05-23 12:10:00"
            structured_report = {
                "summary": "发现 1 个明确问题，1 个建议。",
                "items": [
                    {
                        "status": "issue",
                        "category": "参数不一致",
                        "location": "第1章、第2章",
                        "excerpt": "A 为 10；A 为 20",
                        "description": "同一参数前后不一致",
                        "impact": "客户可能按错误参数配置",
                        "suggestion": "统一参数值。",
                    },
                    {
                        "status": "suggestion",
                        "category": "补充说明",
                        "location": "第3章",
                        "excerpt": "安装前检查环境",
                        "description": "未说明温度范围",
                        "impact": "",
                        "suggestion": "补充适用范围。",
                    },
                ],
            }
            result_json = [
                {
                    "code": "consistency",
                    "name": "全文一致性检查",
                    "result": json.dumps(structured_report, ensure_ascii=False),
                }
            ]
            cursor = get_db().execute(
                """
                INSERT INTO tasks(
                    task_type, ip, original_filename, stored_filename, file_type,
                    file_size, result_json, checks_json, model_name, api_base,
                    status, progress, created_at, updated_at
                )
                VALUES (?, '127.0.0.1', 'report.txt', 'stored.txt', 'txt',
                        1024, ?, '[]', 'model-a', 'https://example.test/v1/chat/completions',
                        'completed', 100, ?, ?)
                """,
                (DOCUMENT_TASK_TYPE, json.dumps(result_json, ensure_ascii=False), now, now),
            )
            get_db().commit()
            task_id = cursor.lastrowid

        detail = self.client.get(f"/admin/tasks/{task_id}")
        soup = BeautifulSoup(detail.get_data(as_text=True), "html.parser")
        item_id = _required_tag(soup.select_one("[data-report-item]"))["data-item-id"]
        review_response = self.client.post(
            f"/admin/tasks/{task_id}/report-items",
            json={
                "result_code": "consistency",
                "item_id": item_id,
                "item_type": "issue",
                "acceptance_status": "rejected",
                "rejection_reason": "false_positive",
                "rejection_note": "上下文可解释",
            },
        )
        self.assertEqual(review_response.status_code, 200)

        response = self.client.get(f"/admin/tasks/{task_id}/export.xlsx")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.mimetype, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertIn(f'document-check-report-{task_id}.xlsx', response.headers["Content-Disposition"])
        workbook = load_workbook(io.BytesIO(response.data), read_only=True, data_only=True)
        try:
            self.assertEqual(workbook.sheetnames, ["报告条目", "统计"])
            report_rows = list(workbook["报告条目"].iter_rows(values_only=True))
            self.assertEqual(
                report_rows[0],
                (
                    "任务ID",
                    "任务类型",
                    "文件名称",
                    "检查项",
                    "条目",
                    "严重程度",
                    "证据可信度",
                    "问题类型",
                    "位置",
                    "原文/证据",
                    "问题描述",
                    "影响",
                    "修改建议",
                    "条目判定",
                    "是否接纳",
                    "不接纳原因",
                    "人工原因",
                ),
            )
            self.assertEqual(len(report_rows), 3)
            self.assertEqual(report_rows[1][0], task_id)
            self.assertEqual(report_rows[1][2], "report.txt")
            self.assertEqual(report_rows[1][3], "全文一致性检查")
            self.assertEqual(report_rows[1][4], "条目 1")
            self.assertEqual(report_rows[1][7], "参数不一致")
            self.assertEqual(report_rows[1][13], "问题")
            self.assertEqual(report_rows[1][14], "不接纳")
            self.assertEqual(report_rows[1][15], "模型误报")
            self.assertEqual(report_rows[1][16], "上下文可解释")
            self.assertEqual(report_rows[2][12], "补充适用范围。")
            self.assertEqual(report_rows[2][13], "建议")
            stats = dict(workbook["统计"].iter_rows(min_row=2, values_only=True))
            self.assertEqual(stats["问题"], 1)
            self.assertEqual(stats["建议"], 1)
            self.assertEqual(stats["不接纳问题"], 1)
            self.assertEqual(stats["问题检出率"], "50.0%")
            self.assertEqual(stats["问题接纳率"], "0.0%")
        finally:
            workbook.close()

    def test_video_task_report_exports_excel_with_compact_media_columns(self):
        with self.app.app_context():
            now = "2026-05-23 12:20:00"
            structured_report = {
                "summary": "发现 1 条安装顺序问题。",
                "items": [
                    {
                        "status": "issue",
                        "category": "安装顺序",
                        "location": "视频时间 00:01.000",
                        "excerpt": "未确认接地线后直接上电",
                        "description": "画面显示未确认接地线状态即执行上电步骤。",
                        "impact": "存在安全风险",
                        "suggestion": "补充接地线确认动作",
                    }
                ],
            }
            result_json = [
                {
                    "code": "video-installation-sequence",
                    "name": "安装顺序检查",
                    "result": json.dumps(structured_report, ensure_ascii=False),
                }
            ]
            cursor = get_db().execute(
                """
                INSERT INTO tasks(
                    task_type, ip, original_filename, stored_filename, file_type,
                    file_size, result_json, checks_json, model_name, api_base,
                    status, progress, created_at, updated_at
                )
                VALUES (?, '127.0.0.1', 'install.mp4', 'stored.mp4', 'mp4',
                        2048, ?, '[]', 'model-a', 'https://example.test/v1/chat/completions',
                        'completed', 100, ?, ?)
                """,
                (VIDEO_TASK_TYPE, json.dumps(result_json, ensure_ascii=False), now, now),
            )
            get_db().commit()
            task_id = cursor.lastrowid

        response = self.client.get(f"/admin/tasks/{task_id}/export.xlsx")

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.data), read_only=True, data_only=True)
        try:
            rows = list(workbook["报告条目"].iter_rows(values_only=True))
            self.assertEqual(
                rows[0],
                (
                    "任务ID",
                    "任务类型",
                    "文件名称",
                    "检查项",
                    "条目",
                    "AI检查结论",
                    "条目判定",
                    "是否接纳",
                    "不接纳原因",
                    "人工原因",
                ),
            )
            self.assertEqual(len(rows[1]), 10)
            self.assertIn("画面显示未确认接地线状态", rows[1][5])
            self.assertIn("位置/画面：视频时间 00:01.000", rows[1][5])
            self.assertIn("建议：补充接地线确认动作", rows[1][5])
            self.assertEqual(rows[1][6], "问题")
        finally:
            workbook.close()

    def test_task_detail_renders_structured_json_report_table(self):
        with self.app.app_context():
            now = "2026-05-23 12:30:00"
            structured_report = {
                "summary": "发现 1 个明确问题，1 个需人工确认项。",
                "items": [
                    {
                        "status": "issue",
                        "category": "参数不一致",
                        "location": "第1章、第2章",
                        "excerpt": "A 为 10；A 为 20",
                        "description": "同一参数前后不一致",
                        "impact": "客户可能按错误参数配置",
                        "suggestion": "统一参数值。",
                    },
                    {
                        "status": "suggestion",
                        "category": "需人工确认",
                        "location": "第3章",
                        "excerpt": "安装前检查环境",
                        "description": "未说明温度范围，证据不足需人工确认。",
                        "impact": "",
                        "suggestion": "确认后补充适用范围。",
                    },
                ],
            }
            result_json = [
                {
                    "code": "consistency",
                    "name": "全文一致性检查",
                    "result": json.dumps(structured_report, ensure_ascii=False),
                    "issue_output_limit": 1,
                }
            ]
            cursor = get_db().execute(
                """
                INSERT INTO tasks(
                    task_type, ip, original_filename, stored_filename, file_type,
                    file_size, result_json, checks_json, model_name, api_base,
                    status, progress, created_at, updated_at
                )
                VALUES (?, '127.0.0.1', 'report.txt', 'stored.txt', 'txt',
                        1024, ?, '[]', 'model-a', 'https://example.test/v1/chat/completions',
                        'completed', 100, ?, ?)
                """,
                (DOCUMENT_TASK_TYPE, json.dumps(result_json, ensure_ascii=False), now, now),
            )
            get_db().commit()
            task_id = cursor.lastrowid

        detail = self.client.get(f"/admin/tasks/{task_id}")

        self.assertEqual(detail.status_code, 200)
        soup = BeautifulSoup(detail.get_data(as_text=True), "html.parser")
        headers = [node.get_text(strip=True) for node in soup.select(".report-table th")]
        self.assertEqual(
            headers,
            ["条目", "严重程度", "证据可信度", "问题类型", "位置", "原文/证据", "问题描述", "影响", "修改建议", "条目判定", "是否接纳", "不接纳原因"],
        )
        rows = soup.select("tr[data-report-item]")
        self.assertEqual(len(rows), 1)
        self.assertEqual(_required_tag(rows[0].select_one("[data-report-item-type]")).get("data-saved-value"), "issue")
        acceptance = _required_tag(rows[0].select_one("[data-report-acceptance-status]"))
        reason = _required_tag(rows[0].select_one("[data-report-rejection-reason]"))
        note = _required_tag(rows[0].select_one("[data-report-rejection-note]"))
        self.assertEqual(acceptance.get("data-saved-value"), "pending")
        self.assertTrue(reason.has_attr("disabled"))
        self.assertTrue(note.has_attr("disabled"))
        self.assertIn("同一参数前后不一致", rows[0].get_text(" ", strip=True))
        self.assertIn("报告硬限制保留前 1 条，省略 1 条", soup.get_text(" ", strip=True))
        self.assertEqual(_required_tag(soup.select_one('[data-report-count="issue"]')).get_text(strip=True), "1")
        self.assertEqual(_required_tag(soup.select_one('[data-report-count="suggestion"]')).get_text(strip=True), "0")

    def test_media_task_detail_uses_compact_report_table(self):
        with self.app.app_context():
            now = "2026-05-23 12:40:00"
            structured_report = {
                "summary": "发现 1 条图文界面问题。",
                "items": [
                    {
                        "status": "issue",
                        "category": "界面步骤一致性",
                        "location": "第25页 / 图3-1",
                        "excerpt": "系统配置表格与文字描述一致",
                        "description": "图中配置项与步骤文字一致，但按钮标注需要人工复核。",
                        "impact": "",
                        "suggestion": "",
                    }
                ],
            }
            result_json = [
                {
                    "code": "image-ui-steps",
                    "name": "图文与界面步骤一致性检查",
                    "result": json.dumps(structured_report, ensure_ascii=False),
                }
            ]
            cursor = get_db().execute(
                """
                INSERT INTO tasks(
                    task_type, ip, original_filename, stored_filename, file_type,
                    file_size, result_json, checks_json, model_name, api_base,
                    status, progress, created_at, updated_at
                )
                VALUES (?, '127.0.0.1', 'guide.pdf', 'stored.pdf', 'pdf',
                        1024, ?, '[]', 'model-a', 'https://example.test/v1/chat/completions',
                        'completed', 100, ?, ?)
                """,
                (IMAGE_TASK_TYPE, json.dumps(result_json, ensure_ascii=False), now, now),
            )
            get_db().commit()
            task_id = cursor.lastrowid

        detail = self.client.get(f"/admin/tasks/{task_id}")
        exported = self.client.get(f"/admin/tasks/{task_id}/export")

        self.assertEqual(detail.status_code, 200)
        self.assertEqual(exported.status_code, 200)
        soup = BeautifulSoup(detail.get_data(as_text=True), "html.parser")
        headers = [node.get_text(strip=True) for node in soup.select(".report-table th")]
        self.assertEqual(headers, ["条目", "AI检查结论", "条目判定", "是否接纳", "不接纳原因"])
        table = _required_tag(soup.select_one(".report-table-media"))
        self.assertIn("report-table-media", table.get("class", []))
        row_text = _required_tag(soup.select_one("tr[data-report-item]")).get_text(" ", strip=True)
        self.assertIn("图中配置项与步骤文字一致", row_text)
        self.assertIn("问题类型：界面步骤一致性", row_text)
        self.assertIn("位置/画面：第25页 / 图3-1", row_text)
        self.assertNotIn("原文/证据", detail.get_data(as_text=True))
        self.assertNotIn("修改建议", exported.get_data(as_text=True))

    def test_image_report_items_exclude_page_level_summary_blocks(self):
        with self.app.app_context():
            now = "2026-05-23 12:42:00"
            result_text = (
                "4. 第28页**：图3-3接线端子的实物图与内部信号示意图结合紧密，"
                "表3-2和表3-3中的脚编号与信号名称匹配。\n"
                "页面级检查结果（批次 8/27）\n"
                "覆盖图片：\n"
                "PDF第29页（page029-screenshot）：0029_page029-screenshot.png\n"
                "PDF第30页（page030-screenshot）：0030_page030-screenshot.png\n"
                "检查项：image-text-correspondence | 图文与界面步骤一致性检查\n"
                "总体判断\n"
                "未发现明确问题。本次提供的截图内容与对应文档文本描述保持一致。\n"
                "明确问题\n"
                "未发现明确问题\n"
                "需人工确认\n"
                "未发现需人工确认项\n\n"
                "5. 第29页**：“图3-4 COM口引脚”展示的 RJ45 母头示意图与"
                "“表3-5 COM引脚定义”的文字描述对应一致。"
            )
            result_json = [
                {
                    "code": "image-text-correspondence",
                    "name": "图文与界面步骤一致性检查",
                    "result": result_text,
                }
            ]
            cursor = get_db().execute(
                """
                INSERT INTO tasks(
                    task_type, ip, original_filename, stored_filename, file_type,
                    file_size, result_json, checks_json, model_name, api_base,
                    status, progress, created_at, updated_at
                )
                VALUES (?, '127.0.0.1', 'guide.pdf', 'stored.pdf', 'pdf',
                        1024, ?, '[]', 'model-a', 'https://example.test/v1/chat/completions',
                        'completed', 100, ?, ?)
                """,
                (IMAGE_TASK_TYPE, json.dumps(result_json, ensure_ascii=False), now, now),
            )
            get_db().commit()
            task_id = cursor.lastrowid

        response = self.client.get(f"/admin/tasks/{task_id}")

        self.assertEqual(response.status_code, 200)
        soup = BeautifulSoup(response.get_data(as_text=True), "html.parser")
        rows = soup.select("tr[data-report-item]")
        self.assertEqual(len(rows), 2)
        first_row_text = rows[0].get_text(" ", strip=True)
        second_row_text = rows[1].get_text(" ", strip=True)
        self.assertIn("第28页", first_row_text)
        self.assertIn("第29页", second_row_text)
        self.assertNotIn("页面级检查结果", first_row_text)
        self.assertNotIn("覆盖图片", first_row_text)
        self.assertNotIn("总体判断", first_row_text)

    def test_language_consistency_no_action_items_are_non_issues(self):
        with self.app.app_context():
            now = "2026-05-23 12:45:00"
            structured_report = {
                "summary": "发现 1 个实质差异，1 个无须修改差异。",
                "items": [
                    {
                        "status": "issue",
                        "category": "缺失与增补",
                        "location": "文档B 第10页目录",
                        "excerpt": "Measurement Methods of PV Optimizers",
                        "description": "中文版目录缺少英文标题中的冠词，但不影响用户理解。",
                        "impact": "无实质影响。",
                        "suggestion": "无需修改。",
                    },
                    {
                        "status": "issue",
                        "category": "关键事实差异",
                        "location": "文档A 第3页 / 文档B 第4页",
                        "excerpt": "额定功率 500W / Rated power 550W",
                        "description": "同一型号的额定功率不一致。",
                        "impact": "客户可能按错误参数配置。",
                        "suggestion": "核实并统一额定功率。",
                    },
                ],
            }
            result_json = [
                {
                    "code": "language-consistency-cross-lingual",
                    "name": "跨语种内容一致性检查",
                    "result": json.dumps(structured_report, ensure_ascii=False),
                }
            ]
            cursor = get_db().execute(
                """
                INSERT INTO tasks(
                    task_type, ip, original_filename, stored_filename, file_type,
                    file_size, result_json, checks_json, model_name, api_base,
                    status, progress, created_at, updated_at
                )
                VALUES (?, '127.0.0.1', 'cross-language.txt', 'stored.txt', 'txt',
                        1024, ?, '[]', 'model-a', 'https://example.test/v1/chat/completions',
                        'completed', 100, ?, ?)
                """,
                (LANGUAGE_CONSISTENCY_TASK_TYPE, json.dumps(result_json, ensure_ascii=False), now, now),
            )
            get_db().commit()
            task_id = cursor.lastrowid

        response = self.client.get(f"/admin/tasks/{task_id}")

        self.assertEqual(response.status_code, 200)
        soup = BeautifulSoup(response.get_data(as_text=True), "html.parser")
        rows = soup.select("tr[data-report-item]")
        self.assertEqual(len(rows), 2)
        self.assertIn("额定功率不一致", rows[0].get_text(" ", strip=True))
        self.assertEqual(_required_tag(rows[0].select_one("[data-report-item-type]")).get("data-saved-value"), "issue")
        self.assertIn("无需修改", rows[1].get_text(" ", strip=True))
        self.assertEqual(_required_tag(rows[1].select_one("[data-report-item-type]")).get("data-saved-value"), "non_issue")
        self.assertEqual(_required_tag(soup.select_one('[data-report-count="issue"]')).get_text(strip=True), "1")
        self.assertEqual(_required_tag(soup.select_one('[data-report-count="non_issue"]')).get_text(strip=True), "1")
        self.assertEqual(_required_tag(soup.select_one('[data-report-count="issue_detection_rate"]')).get_text(strip=True), "50.0%")

    def test_task_detail_parses_double_encoded_structured_json_report(self):
        with self.app.app_context():
            now = "2026-05-23 13:00:00"
            structured_report = {
                "summary": "发现错别字和标点问题。",
                "items": [
                    {
                        "status": "issue",
                        "category": "标点误用",
                        "location": "[第9页] 章节：1 安全注意事项",
                        "excerpt": "或/和",
                        "description": "“或/和”中斜杠前后存在多余空格。",
                        "impact": "影响阅读体验和文档规范性。",
                        "suggestion": "改为“或和”或按规范统一表达。",
                    }
                ],
            }
            result_json = [
                {
                    "code": "typo",
                    "name": "错别字检查",
                    "result": json.dumps(json.dumps(structured_report, ensure_ascii=False), ensure_ascii=False),
                }
            ]
            cursor = get_db().execute(
                """
                INSERT INTO tasks(
                    task_type, ip, original_filename, stored_filename, file_type,
                    file_size, result_json, checks_json, model_name, api_base,
                    status, progress, created_at, updated_at
                )
                VALUES (?, '127.0.0.1', 'report.txt', 'stored.txt', 'txt',
                        1024, ?, '[]', 'model-a', 'https://example.test/v1/chat/completions',
                        'completed', 100, ?, ?)
                """,
                (DOCUMENT_TASK_TYPE, json.dumps(result_json, ensure_ascii=False), now, now),
            )
            get_db().commit()
            task_id = cursor.lastrowid

        response = self.client.get(f"/admin/tasks/{task_id}")

        self.assertEqual(response.status_code, 200)
        soup = BeautifulSoup(response.get_data(as_text=True), "html.parser")
        rows = soup.select("tr[data-report-item]")
        self.assertEqual(len(rows), 1)
        row_text = rows[0].get_text(" ", strip=True)
        self.assertIn("标点误用", row_text)
        self.assertIn("[第9页] 章节：1 安全注意事项", row_text)
        self.assertIn("“或/和”中斜杠前后存在多余空格。", row_text)
        self.assertNotIn('"items"', row_text)
        self.assertEqual(_required_tag(soup.select_one('[data-report-count="issue"]')).get_text(strip=True), "1")

    def test_task_detail_parses_structured_json_with_raw_newline_in_string(self):
        with self.app.app_context():
            now = "2026-05-23 13:30:00"
            raw_json = (
                '{"summary":"发现 1 个问题","items":[{"status":"issue","category":"格式问题",'
                '"location":"第1页","excerpt":"第一行\n第二行","description":"描述包含原文换行",'
                '"impact":"影响阅读","suggestion":"删除多余换行"}]}'
            )
            result_json = [
                {
                    "code": "typo",
                    "name": "错别字检查",
                    "result": raw_json,
                }
            ]
            cursor = get_db().execute(
                """
                INSERT INTO tasks(
                    task_type, ip, original_filename, stored_filename, file_type,
                    file_size, result_json, checks_json, model_name, api_base,
                    status, progress, created_at, updated_at
                )
                VALUES (?, '127.0.0.1', 'report.txt', 'stored.txt', 'txt',
                        1024, ?, '[]', 'model-a', 'https://example.test/v1/chat/completions',
                        'completed', 100, ?, ?)
                """,
                (DOCUMENT_TASK_TYPE, json.dumps(result_json, ensure_ascii=False), now, now),
            )
            get_db().commit()
            task_id = cursor.lastrowid

        response = self.client.get(f"/admin/tasks/{task_id}")

        self.assertEqual(response.status_code, 200)
        soup = BeautifulSoup(response.get_data(as_text=True), "html.parser")
        row = _required_tag(soup.select_one("tr[data-report-item]"))
        row_text = row.get_text(" ", strip=True)
        self.assertIn("格式问题", row_text)
        self.assertIn("第一行", row_text)
        self.assertIn("第二行", row_text)
        self.assertNotIn('"items"', row_text)

    def test_task_detail_salvages_complete_items_from_truncated_structured_json(self):
        with self.app.app_context():
            now = "2026-05-23 13:35:00"
            raw_json = (
                '{"summary":"总体风险等级为中","items":['
                '{"status":"issue","category":"参数不一致","location":"第10页","excerpt":"10A",'
                '"description":"电流参数不一致。","impact":"可能误导实施","suggestion":"统一为10A"},'
                '{"status":"suggestion","category":"术语一致性","location":"第20页","excerpt":"控制器",'
                '"description":"名称表述不统一。","impact":"影响理解","suggestion":"统一名称"},'
                '{"status":"issue","category":"未完成条目","description":"响应在这里被截断'
            )
            result_json = [
                {
                    "code": "consistency",
                    "name": "全文一致性检查",
                    "result": raw_json,
                }
            ]
            cursor = get_db().execute(
                """
                INSERT INTO tasks(
                    task_type, ip, original_filename, stored_filename, file_type,
                    file_size, result_json, checks_json, model_name, api_base,
                    status, progress, created_at, updated_at
                )
                VALUES (?, '127.0.0.1', 'report.txt', 'stored.txt', 'txt',
                        1024, ?, '[]', 'model-a', 'https://example.test/v1/chat/completions',
                        'completed', 100, ?, ?)
                """,
                (DOCUMENT_TASK_TYPE, json.dumps(result_json, ensure_ascii=False), now, now),
            )
            get_db().commit()
            task_id = cursor.lastrowid

        response = self.client.get(f"/admin/tasks/{task_id}")

        self.assertEqual(response.status_code, 200)
        soup = BeautifulSoup(response.get_data(as_text=True), "html.parser")
        rows = soup.select("tr[data-report-item]")
        self.assertEqual(len(rows), 2)
        page_text = soup.get_text(" ", strip=True)
        rows_text = " ".join(row.get_text(" ", strip=True) for row in rows)
        self.assertIn("总体风险等级为中", page_text)
        self.assertIn("电流参数不一致", rows_text)
        self.assertIn("名称表述不统一", rows_text)
        self.assertNotIn("未完成条目", rows_text)
        self.assertNotIn('"items"', rows_text)

    def test_task_detail_deduplicates_ranks_and_limits_structured_items(self):
        with self.app.app_context():
            now = "2026-05-23 13:37:00"
            structured_result = {
                "summary": "发现多个不同优先级的问题",
                "items": [
                    {
                        "status": "issue",
                        "severity": "low",
                        "confidence": "high",
                        "category": "低优先级",
                        "location": "第1页",
                        "excerpt": "低风险原文",
                        "description": "低风险问题",
                        "impact": "影响较小",
                        "suggestion": "后续修改",
                    },
                    {
                        "status": "issue",
                        "severity": "critical",
                        "confidence": "medium",
                        "category": "安全问题",
                        "location": "第20页",
                        "excerpt": "禁止带电操作",
                        "description": "安全约束前后矛盾",
                        "impact": "可能造成人身安全风险",
                        "suggestion": "立即统一安全要求",
                    },
                    {
                        "status": "issue",
                        "severity": "high",
                        "confidence": "high",
                        "category": "参数问题",
                        "location": "第10页",
                        "excerpt": "额定电流10A",
                        "description": "关键参数不一致",
                        "impact": "可能导致配置错误",
                        "suggestion": "统一关键参数",
                    },
                    {
                        "status": "issue",
                        "severity": "high",
                        "confidence": "high",
                        "category": "参数问题",
                        "location": "第30页",
                        "excerpt": "额定电流10A",
                        "description": "关键参数不一致",
                        "impact": "可能导致配置错误",
                        "suggestion": "统一关键参数",
                    },
                    {
                        "status": "non_issue",
                        "severity": "critical",
                        "confidence": "high",
                        "category": "无需修改",
                        "location": "第40页",
                        "excerpt": "当前表述正确",
                        "description": "该项不是问题",
                        "impact": "无实质影响",
                        "suggestion": "无需修改",
                    },
                ],
            }
            result_json = [
                {
                    "code": "consistency",
                    "name": "全文一致性检查",
                    "result": json.dumps(structured_result, ensure_ascii=False),
                    "issue_output_limit": 2,
                }
            ]
            cursor = get_db().execute(
                """
                INSERT INTO tasks(
                    task_type, ip, original_filename, stored_filename, file_type,
                    file_size, result_json, checks_json, model_name, api_base,
                    status, progress, created_at, updated_at
                )
                VALUES (?, '127.0.0.1', 'ranked.txt', 'stored.txt', 'txt',
                        1024, ?, '[]', 'model-a', 'https://example.test/v1/chat/completions',
                        'completed', 100, ?, ?)
                """,
                (DOCUMENT_TASK_TYPE, json.dumps(result_json, ensure_ascii=False), now, now),
            )
            get_db().commit()
            task_id = cursor.lastrowid

        response = self.client.get(f"/admin/tasks/{task_id}")

        self.assertEqual(response.status_code, 200)
        soup = BeautifulSoup(response.get_data(as_text=True), "html.parser")
        rows = soup.select("tr[data-report-item]")
        self.assertEqual(len(rows), 2)
        first_row = rows[0].get_text(" ", strip=True)
        second_row = rows[1].get_text(" ", strip=True)
        self.assertIn("高", first_row)
        self.assertIn("关键参数不一致", first_row)
        self.assertIn("第10页；第30页", first_row)
        self.assertIn("高", second_row)
        self.assertIn("低风险问题", second_row)
        page_text = soup.get_text(" ", strip=True)
        self.assertIn("已合并 1 条重复问题", page_text)
        self.assertIn("硬限制保留前 2 条，省略 2 条", page_text)
        self.assertNotIn("安全约束前后矛盾", " ".join(row.get_text(" ", strip=True) for row in rows))
        self.assertNotIn("该项不是问题", " ".join(row.get_text(" ", strip=True) for row in rows))

    def test_report_hard_limit_keeps_top_thirty_likely_issues(self):
        raw_items = [
            {
                "status": "issue",
                "category": "明确问题",
                "description": f"明确问题 {index}",
            }
            for index in range(1, 32)
        ]
        raw_items.extend(
            [
                {
                    "status": "suggestion",
                    "severity": "critical",
                    "confidence": "high",
                    "category": "建议",
                    "description": "高置信度建议",
                },
                {
                    "status": "non_issue",
                    "severity": "critical",
                    "confidence": "high",
                    "category": "非问题",
                    "description": "明确非问题",
                },
            ]
        )
        prepared = _prepare_task_results(
            [
                {
                    "code": "compliance",
                    "name": "文档规范性检查",
                    "result": json.dumps({"summary": "检查完成", "items": raw_items}, ensure_ascii=False),
                    "issue_output_limit": 100,
                }
            ]
        )

        report_items = prepared[0]["report_items"]
        self.assertEqual(len(report_items), 30)
        self.assertTrue(all(item["type"] == "issue" for item in report_items))
        self.assertEqual(prepared[0]["report_limit"]["limit"], 30)
        self.assertEqual(prepared[0]["report_limit"]["omitted_count"], 3)

    def test_task_detail_repairs_duplicate_status_json_value(self):
        with self.app.app_context():
            now = "2026-05-23 13:40:00"
            raw_json = (
                '{"summary":"发现 1 个建议","items":[{"status":"issue":"suggestion",'
                '"category":"结构层级","location":"第4页","excerpt":"用户指南",'
                '"description":"文档名称和前言称谓不一致。","impact":"影响客户认知",'
                '"suggestion":"统一文档称谓。"}]}'
            )
            result_json = [
                {
                    "code": "compliance",
                    "name": "文档规范性检查",
                    "result": raw_json,
                }
            ]
            cursor = get_db().execute(
                """
                INSERT INTO tasks(
                    task_type, ip, original_filename, stored_filename, file_type,
                    file_size, result_json, checks_json, model_name, api_base,
                    status, progress, created_at, updated_at
                )
                VALUES (?, '127.0.0.1', 'report.txt', 'stored.txt', 'txt',
                        1024, ?, '[]', 'model-a', 'https://example.test/v1/chat/completions',
                        'completed', 100, ?, ?)
                """,
                (DOCUMENT_TASK_TYPE, json.dumps(result_json, ensure_ascii=False), now, now),
            )
            get_db().commit()
            task_id = cursor.lastrowid

        response = self.client.get(f"/admin/tasks/{task_id}")

        self.assertEqual(response.status_code, 200)
        soup = BeautifulSoup(response.get_data(as_text=True), "html.parser")
        row = _required_tag(soup.select_one("tr[data-report-item]"))
        row_text = row.get_text(" ", strip=True)
        self.assertIn("结构层级", row_text)
        self.assertIn("文档名称和前言称谓不一致。", row_text)
        self.assertNotIn('"items"', row_text)
        self.assertEqual(_required_tag(row.select_one("[data-report-item-type]")).get("data-saved-value"), "suggestion")

    def test_task_detail_splits_bold_numbered_compliance_items(self):
        with self.app.app_context():
            now = "2026-05-24 09:00:00"
            result_json = [
                {
                    "code": "compliance",
                    "name": "文档规范性检查",
                    "result": (
                        "总体规范性结论：该资料存在面向客户表达风险。\n\n"
                        "---\n\n"
                        "## 问题逐条列表\n\n"
                        "**1. 问题类型：技术信息呈现（严重错误）**\n"
                        "- 位置：第29页 / 3.2 参数说明\n"
                        "- 原文摘录：支持 220V 输入。\n"
                        "- 问题描述：参数呈现与客户资料规范不一致。\n"
                        "- 客户影响：客户可能按错误信息配置。\n"
                        "- 修改建议：核实并修正文档参数。\n\n"
                        "**2. 问题类型：客户资料定位（内部口吻）**\n"
                        "- 位置：第3页 / 注意事项\n"
                        "- 原文摘录：研发确认后再发布。\n"
                        "- 问题描述：面向客户资料出现内部流程口吻。\n"
                        "- 客户影响：影响客户对资料正式性的判断。\n"
                        "- 修改建议：改为客户可理解的正式表述。"
                    ),
                }
            ]
            cursor = get_db().execute(
                """
                INSERT INTO tasks(
                    task_type, ip, original_filename, stored_filename, file_type,
                    file_size, result_json, checks_json, model_name, api_base,
                    status, progress, created_at, updated_at
                )
                VALUES (?, '127.0.0.1', 'report.txt', 'stored.txt', 'txt',
                        1024, ?, '[]', 'model-a', 'https://example.test/v1/chat/completions',
                        'completed', 100, ?, ?)
                """,
                (DOCUMENT_TASK_TYPE, json.dumps(result_json, ensure_ascii=False), now, now),
            )
            get_db().commit()
            task_id = cursor.lastrowid

        detail = self.client.get(f"/admin/tasks/{task_id}")

        self.assertEqual(detail.status_code, 200)
        soup = BeautifulSoup(detail.get_data(as_text=True), "html.parser")
        items = soup.select("[data-report-item]")
        self.assertEqual(len(items), 2)
        self.assertIn("技术信息呈现", items[0].get_text(" ", strip=True))
        self.assertIn("客户资料定位", items[1].get_text(" ", strip=True))
        self.assertNotIn("总体规范性结论", items[0].get_text(" ", strip=True))
        self.assertEqual(_required_tag(soup.select_one('[data-report-count="total"]')).get_text(strip=True), "2")

    def test_admin_task_list_shows_report_item_totals(self):
        with self.app.app_context():
            now = "2026-05-24 10:00:00"
            result_json = [
                {
                    "code": "compliance",
                    "name": "文档规范性检查",
                    "result": (
                        "1. 问题类型：参数错误\n"
                        "位置：第1页\n"
                        "问题描述：参数前后不一致。\n\n"
                        "2. 建议：补充适用范围\n"
                        "修改建议：增加适用范围说明。\n\n"
                        "3. 非问题：未发现客户风险\n"
                        "问题描述：该表述无需修改。"
                    ),
                }
            ]
            get_db().execute(
                """
                INSERT INTO tasks(
                    task_type, ip, original_filename, stored_filename, file_type,
                    file_size, result_json, checks_json, model_name, api_base,
                    status, progress, created_at, updated_at
                )
                VALUES (?, '127.0.0.1', 'report.txt', 'stored.txt', 'txt',
                        1024, ?, '[]', 'model-a', 'https://example.test/v1/chat/completions',
                        'completed', 100, ?, ?)
                """,
                (DOCUMENT_TASK_TYPE, json.dumps(result_json, ensure_ascii=False), now, now),
            )
            get_db().commit()

        response = self.client.get("/admin/tasks")

        self.assertEqual(response.status_code, 200)
        soup = BeautifulSoup(response.get_data(as_text=True), "html.parser")
        self.assertEqual(_required_tag(soup.select_one('[data-admin-report-count="issue"]')).get_text(strip=True), "1")
        self.assertEqual(_required_tag(soup.select_one('[data-admin-report-count="suggestion"]')).get_text(strip=True), "1")
        self.assertEqual(_required_tag(soup.select_one('[data-admin-report-count="non_issue"]')).get_text(strip=True), "1")
        self.assertEqual(_required_tag(soup.select_one('[data-admin-report-count="total"]')).get_text(strip=True), "3")

    def test_user_task_list_pagination_allows_page_jump(self):
        for index in range(21):
            self._insert_task(created_at=f"2026-05-01 10:{index:02d}:00")

        response = self.client.get("/?page=2")

        self.assertEqual(response.status_code, 200)
        soup = BeautifulSoup(response.get_data(as_text=True), "html.parser")
        form = _required_tag(soup.select_one(".pagination .page-jump-form"))
        page_input = _required_tag(form.select_one('input[name="page"]'))
        self.assertEqual(form.get("method"), "get")
        self.assertEqual(form.get("action"), "/")
        self.assertEqual(page_input.get("type"), "number")
        self.assertEqual(page_input.get("min"), "1")
        self.assertEqual(page_input.get("max"), "2")
        self.assertEqual(page_input.get("value"), "2")
        self.assertIsNotNone(form.select_one('button[type="submit"]'))

    def test_user_cancel_preserves_proxy_prefix_and_page(self):
        oldest_task_id = None
        for index in range(21):
            task_id = self._insert_task(status="running", created_at=f"2026-05-01 10:{index:02d}:00")
            if index == 0:
                oldest_task_id = task_id

        response = self.client.get(
            "/?page=2",
            environ_overrides={"SCRIPT_NAME": "/infoCheck"},
        )

        self.assertEqual(response.status_code, 200)
        soup = BeautifulSoup(response.get_data(as_text=True), "html.parser")
        form = _required_tag(soup.select_one('form[action$="/cancel"]'))
        next_input = _required_tag(form.select_one('input[name="next"]'))
        self.assertEqual(form.get("action"), f"/infoCheck/tasks/{oldest_task_id}/cancel")
        self.assertEqual(next_input.get("value"), "/infoCheck/?page=2")

        cancel_response = self.client.post(
            f"/tasks/{oldest_task_id}/cancel",
            data={"next": next_input.get("value")},
            environ_overrides={"SCRIPT_NAME": "/infoCheck"},
        )

        self.assertEqual(cancel_response.status_code, 302)
        self.assertEqual(cancel_response.headers["Location"], "/infoCheck/?page=2")

    def test_user_task_report_link_has_clean_url_and_returns_to_task_list(self):
        for index in range(21):
            self._insert_task(created_at=f"2026-05-01 10:{index:02d}:00")

        response = self.client.get("/?page=2")

        self.assertEqual(response.status_code, 200)
        soup = BeautifulSoup(response.get_data(as_text=True), "html.parser")
        detail_link = _required_tag(soup.select_one("a.task-report-link"))
        self.assertRegex(str(detail_link.get("href")), r"^/tasks/\d+$")
        detail_response = self.client.get(str(detail_link.get("href")))

        self.assertEqual(detail_response.status_code, 200)
        detail_soup = BeautifulSoup(detail_response.get_data(as_text=True), "html.parser")
        back_link = _required_tag(detail_soup.select_one(".report-toolbar > a"))
        self.assertEqual(back_link.get("href"), "/")

    def test_all_task_report_links_open_in_new_tabs(self):
        task_routes = (
            (DOCUMENT_TASK_TYPE, "/", "/admin/tasks"),
            (CONSISTENCY_TASK_TYPE, "/consistency", "/admin/consistency"),
            (LANGUAGE_CONSISTENCY_TASK_TYPE, "/language-consistency", "/admin/language-consistency"),
            (IMAGE_TASK_TYPE, "/images", "/admin/images"),
            (VIDEO_TASK_TYPE, "/videos", "/admin/videos"),
        )

        for task_type, user_list_url, admin_list_url in task_routes:
            task_id = self._insert_task(task_type=task_type)
            for list_url, report_path in (
                (user_list_url, f"/tasks/{task_id}"),
                (admin_list_url, f"/admin/tasks/{task_id}"),
            ):
                with self.subTest(task_type=task_type, list_url=list_url):
                    response = self.client.get(list_url)
                    self.assertEqual(response.status_code, 200)
                    soup = BeautifulSoup(response.get_data(as_text=True), "html.parser")
                    report_links = soup.select(f'a[href="{report_path}"]')
                    self.assertEqual(len(report_links), 2)
                    for report_link in report_links:
                        self.assertEqual(report_link.get("target"), "_blank")
                        self.assertIn("noopener", report_link.get("rel", []))

    def test_admin_task_list_page_jump_preserves_filters(self):
        for index in range(21):
            self._insert_task(status="completed", created_at=f"2026-05-01 10:{index:02d}:00")

        response = self.client.get("/admin/tasks?status=completed&owner=127.0.0.1&page=2")

        self.assertEqual(response.status_code, 200)
        soup = BeautifulSoup(response.get_data(as_text=True), "html.parser")
        form = _required_tag(soup.select_one(".pagination .page-jump-form"))
        page_input = _required_tag(form.select_one('input[name="page"]'))
        status_input = _required_tag(form.select_one('input[name="status"]'))
        owner_input = _required_tag(form.select_one('input[name="owner"]'))
        self.assertEqual(form.get("method"), "get")
        self.assertEqual(form.get("action"), "/admin/tasks")
        self.assertEqual(page_input.get("value"), "2")
        self.assertEqual(page_input.get("max"), "2")
        self.assertEqual(status_input.get("value"), "completed")
        self.assertEqual(owner_input.get("value"), "127.0.0.1")

    def test_admin_task_report_link_has_clean_url_and_returns_to_task_list(self):
        for index in range(21):
            self._insert_task(status="completed", created_at=f"2026-05-01 10:{index:02d}:00")

        list_url = "/admin/tasks?status=completed&owner=127.0.0.1&page=2"
        response = self.client.get(list_url)

        self.assertEqual(response.status_code, 200)
        soup = BeautifulSoup(response.get_data(as_text=True), "html.parser")
        detail_link = _required_tag(soup.select_one("a.task-report-link"))
        self.assertRegex(str(detail_link.get("href")), r"^/admin/tasks/\d+$")
        detail_response = self.client.get(str(detail_link.get("href")))

        self.assertEqual(detail_response.status_code, 200)
        detail_soup = BeautifulSoup(detail_response.get_data(as_text=True), "html.parser")
        back_link = _required_tag(detail_soup.select_one(".report-toolbar > a"))
        self.assertEqual(back_link.get("href"), "/admin/tasks")

    def test_report_item_type_update_persists_classification(self):
        with self.app.app_context():
            now = "2026-05-24 12:00:00"
            result_json = [
                {
                    "code": "compliance",
                    "name": "文档规范性检查",
                    "result": (
                        "1. 问题类型：内部备注残留\n"
                        "位置：第1章\n"
                        "原文摘录：TODO：研发确认\n"
                        "问题描述：面向客户资料中残留内部备注\n"
                        "影响说明：影响客户信任\n"
                        "修改建议：删除内部备注。"
                    ),
                }
            ]
            cursor = get_db().execute(
                """
                INSERT INTO tasks(
                    task_type, ip, original_filename, stored_filename, file_type,
                    file_size, result_json, checks_json, model_name, api_base,
                    status, progress, created_at, updated_at
                )
                VALUES (?, '127.0.0.1', 'report.txt', 'stored.txt', 'txt',
                        1024, ?, '[]', 'model-a', 'https://example.test/v1/chat/completions',
                        'completed', 100, ?, ?)
                """,
                (DOCUMENT_TASK_TYPE, json.dumps(result_json, ensure_ascii=False), now, now),
            )
            get_db().commit()
            task_id = cursor.lastrowid

        detail = self.client.get(f"/admin/tasks/{task_id}")
        soup = BeautifulSoup(detail.get_data(as_text=True), "html.parser")
        item = _required_tag(soup.select_one("[data-report-item]"))
        item_id = item["data-item-id"]

        response = self.client.post(
            f"/admin/tasks/{task_id}/report-items",
            json={
                "result_code": "compliance",
                "item_id": item_id,
                "item_type": "non_issue",
                "acceptance_status": "rejected",
                "rejection_reason": "false_positive",
                "rejection_note": "原文上下文可解释",
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["totals"]["issue"], 0)
        self.assertEqual(payload["totals"]["non_issue"], 1)
        self.assertEqual(payload["acceptance_status"], "rejected")
        self.assertEqual(payload["rejection_reason"], "false_positive")
        self.assertEqual(payload["rejection_note"], "原文上下文可解释")
        self.assertEqual(payload["totals"]["issue_detection_rate"], "0.0%")
        self.assertEqual(payload["totals"]["issue_acceptance_rate"], "-")
        with self.app.app_context():
            task = get_db().execute("SELECT result_json FROM tasks WHERE id = ?", (task_id,)).fetchone()
            stored = json.loads(task["result_json"])
        self.assertEqual(stored[0]["item_classifications"][item_id], "non_issue")
        self.assertEqual(
            stored[0]["item_acceptances"][item_id],
            {
                "status": "rejected",
                "rejection_reason": "false_positive",
                "rejection_note": "原文上下文可解释",
            },
        )

        accepted_response = self.client.post(
            f"/admin/tasks/{task_id}/report-items",
            json={
                "result_code": "compliance",
                "item_id": item_id,
                "item_type": "issue",
                "acceptance_status": "accepted",
            },
        )

        self.assertEqual(accepted_response.status_code, 200)
        accepted_payload = accepted_response.get_json()
        self.assertEqual(accepted_payload["totals"]["issue"], 1)
        self.assertEqual(accepted_payload["totals"]["accepted_issue"], 1)
        self.assertEqual(accepted_payload["totals"]["pending_issue_acceptance"], 0)
        self.assertEqual(accepted_payload["totals"]["issue_detection_rate"], "100.0%")
        self.assertEqual(accepted_payload["totals"]["issue_acceptance_rate"], "100.0%")

    def test_report_suppression_candidate_can_hide_future_similar_description(self):
        result_json = [
            {
                "code": "compliance",
                "name": "文档规范性检查",
                "result": (
                    "1. 问题类型：内部备注残留\n"
                    "位置：第1章\n"
                    "原文摘录：TODO：研发确认\n"
                    "问题描述：面向客户资料中残留内部备注\n"
                    "影响说明：影响客户信任\n"
                    "修改建议：删除内部备注。"
                ),
            }
        ]
        with self.app.app_context():
            now = "2026-05-24 12:20:00"
            cursor = get_db().execute(
                """
                INSERT INTO tasks(
                    task_type, ip, original_filename, stored_filename, file_type,
                    file_size, result_json, checks_json, model_name, api_base,
                    status, progress, created_at, updated_at
                )
                VALUES (?, '127.0.0.1', 'source.txt', 'source.txt', 'txt',
                        1024, ?, '[]', 'model-a', 'https://example.test/v1/chat/completions',
                        'completed', 100, ?, ?)
                """,
                (DOCUMENT_TASK_TYPE, json.dumps(result_json, ensure_ascii=False), now, now),
            )
            get_db().commit()
            source_task_id = cursor.lastrowid

        detail = self.client.get(f"/admin/tasks/{source_task_id}")
        soup = BeautifulSoup(detail.get_data(as_text=True), "html.parser")
        item_id = _required_tag(soup.select_one("[data-report-item]"))["data-item-id"]

        response = self.client.post(
            f"/admin/tasks/{source_task_id}/report-items",
            json={
                "result_code": "compliance",
                "item_id": item_id,
                "item_type": "non_issue",
                "acceptance_status": "rejected",
                "rejection_reason": "false_positive",
                "rejection_note": "公司规范允许该表述",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.get_json()["suppression_candidate_created"])
        with self.app.app_context():
            rule = get_db().execute(
                "SELECT * FROM report_suppression_rules WHERE check_code = 'compliance'"
            ).fetchone()
            self.assertIsNotNone(rule)
            self.assertEqual(rule["enabled"], 0)
            self.assertEqual(rule["reason"], "模型误报")
            rule_id = rule["id"]

        settings = self.client.get("/admin/settings")
        settings_html = settings.get_data(as_text=True)
        self.assertIn("误报忽略规则", settings_html)
        self.assertIn("候选", settings_html)
        self.assertIn("面向客户资料中残留内部备注", settings_html)

        future_result_json = [
            {
                "code": "compliance",
                "name": "文档规范性检查",
                "result": json.dumps(
                    {
                        "summary": "发现一个相似误报和一个无关问题。",
                        "items": [
                            {
                                "status": "issue",
                                "category": "交付内容",
                                "location": "第8章",
                                "excerpt": "研发内部备注",
                                "description": "客户交付文档仍保留研发内部备注",
                                "impact": "可能影响交付观感",
                                "suggestion": "清理备注",
                            },
                            {
                                "status": "issue",
                                "category": "安全信息",
                                "location": "第9章",
                                "excerpt": "操作前确认断电",
                                "description": "文档中缺少安全警告",
                                "impact": "可能影响安全操作",
                                "suggestion": "补充安全警告",
                            },
                        ],
                    },
                    ensure_ascii=False,
                ),
            }
        ]

        with self.app.app_context():
            now = "2026-05-24 12:25:00"
            cursor = get_db().execute(
                """
                INSERT INTO tasks(
                    task_type, ip, original_filename, stored_filename, file_type,
                    file_size, result_json, checks_json, model_name, api_base,
                    status, progress, created_at, updated_at
                )
                VALUES (?, '127.0.0.1', 'future.txt', 'future.txt', 'txt',
                        1024, ?, '[]', 'model-a', 'https://example.test/v1/chat/completions',
                        'completed', 100, ?, ?)
                """,
                (DOCUMENT_TASK_TYPE, json.dumps(future_result_json, ensure_ascii=False), now, now),
            )
            get_db().commit()
            future_task_id = cursor.lastrowid

        detail_before_enable = self.client.get(f"/admin/tasks/{future_task_id}")
        self.assertEqual(len(BeautifulSoup(detail_before_enable.get_data(as_text=True), "html.parser").select("[data-report-item]")), 2)

        enable_response = self.client.post(
            "/admin/settings",
            data={
                "action": "report_suppression_rule",
                "rule_id": str(rule_id),
                "operation": "enable",
            },
            follow_redirects=True,
        )
        self.assertEqual(enable_response.status_code, 200)
        self.assertIn("已启用", enable_response.get_data(as_text=True))

        detail_after_enable = self.client.get(f"/admin/tasks/{future_task_id}")
        soup = BeautifulSoup(detail_after_enable.get_data(as_text=True), "html.parser")
        visible_items = soup.select("[data-report-item]")
        self.assertEqual(len(visible_items), 1)
        self.assertIn("文档中缺少安全警告", visible_items[0].get_text(" ", strip=True))
        self.assertEqual(_required_tag(soup.select_one('[data-report-count="suppressed"]')).get_text(strip=True), "1")
        self.assertIn("已忽略误报 1 条", detail_after_enable.get_data(as_text=True))
        self.assertIn("客户交付文档仍保留研发内部备注", detail_after_enable.get_data(as_text=True))
        self.assertIn("描述相似度 71%", detail_after_enable.get_data(as_text=True))
        with self.app.app_context():
            updated_rule = get_db().execute("SELECT hit_count FROM report_suppression_rules WHERE id = ?", (rule_id,)).fetchone()
            self.assertEqual(updated_rule["hit_count"], 1)

    def test_report_item_reject_requires_reason(self):
        with self.app.app_context():
            now = "2026-05-24 12:30:00"
            result_json = [
                {
                    "code": "compliance",
                    "name": "文档规范性检查",
                    "result": (
                        "1. 问题类型：参数错误\n"
                        "位置：第1章\n"
                        "问题描述：参数错误\n"
                        "修改建议：修正参数。"
                    ),
                }
            ]
            cursor = get_db().execute(
                """
                INSERT INTO tasks(
                    task_type, ip, original_filename, stored_filename, file_type,
                    file_size, result_json, checks_json, model_name, api_base,
                    status, progress, created_at, updated_at
                )
                VALUES (?, '127.0.0.1', 'report.txt', 'stored.txt', 'txt',
                        1024, ?, '[]', 'model-a', 'https://example.test/v1/chat/completions',
                        'completed', 100, ?, ?)
                """,
                (DOCUMENT_TASK_TYPE, json.dumps(result_json, ensure_ascii=False), now, now),
            )
            get_db().commit()
            task_id = cursor.lastrowid

        detail = self.client.get(f"/admin/tasks/{task_id}")
        soup = BeautifulSoup(detail.get_data(as_text=True), "html.parser")
        item_id = _required_tag(soup.select_one("[data-report-item]"))["data-item-id"]

        response = self.client.post(
            f"/admin/tasks/{task_id}/report-items",
            json={
                "result_code": "compliance",
                "item_id": item_id,
                "item_type": "issue",
                "acceptance_status": "rejected",
                "rejection_reason": "",
                "rejection_note": "",
            },
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.get_json()["error"], "不接纳时必须选择或填写原因。")

    def test_create_task_uses_trusted_header_identity(self):
        model_id = self._configure_provider("trusted_header:100086")
        self.app.config["AUTH"] = {
            "mode": "trusted_header",
            "trusted_header": {
                "user_id": "X-SSO-User-Id",
                "username": "X-SSO-User-Name",
            },
        }
        with self.app.app_context():
            item = get_db().execute("SELECT id FROM check_items WHERE code = 'typo'").fetchone()

        response = self.client.post(
            "/",
            data={
                "document": (io.BytesIO("测试文档".encode("utf-8")), "doc.txt"),
                "checks": [str(item["id"])],
                "model_id": model_id,
            },
            headers={"X-SSO-User-Id": "100086", "X-SSO-User-Name": "张三"},
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            task = get_db().execute("SELECT owner_subject, owner_name_snapshot, owner_source, ip FROM tasks").fetchone()
        self.assertEqual(task["owner_subject"], "trusted_header:100086")
        self.assertEqual(task["owner_name_snapshot"], "张三")
        self.assertEqual(task["owner_source"], "trusted_header")
        self.assertEqual(task["ip"], "127.0.0.1")

    def test_trusted_header_user_page_requires_sso_header(self):
        self.app.config["AUTH"] = {
            "mode": "trusted_header",
            "trusted_header": {
                "user_id": "X-SSO-User-Id",
                "username": "X-SSO-User-Name",
            },
        }

        response = self.client.get("/")

        self.assertEqual(response.status_code, 401)
        self.assertIn("未收到 SSO 用户信息", response.get_data(as_text=True))

    def test_trusted_header_admin_settings_still_uses_local_admin_login(self):
        self.app.config["AUTH"] = {
            "mode": "trusted_header",
            "trusted_header": {
                "user_id": "X-SSO-User-Id",
                "username": "X-SSO-User-Name",
            },
        }

        response = self.client.get("/admin/settings")

        self.assertEqual(response.status_code, 200)
        self.assertIn("系统设置", response.get_data(as_text=True))

    def test_trusted_header_admin_task_page_requires_same_sso_user(self):
        self.app.config["AUTH"] = {
            "mode": "trusted_header",
            "trusted_header": {
                "user_id": "X-SSO-User-Id",
                "username": "X-SSO-User-Name",
            },
        }

        response = self.client.get("/admin/tasks")

        self.assertEqual(response.status_code, 401)
        self.assertIn("未收到 SSO 用户信息", response.get_data(as_text=True))

    def test_trusted_header_admin_task_page_uses_sso_user_models(self):
        model_id = self._configure_provider("trusted_header:100086")
        self.app.config["AUTH"] = {
            "mode": "trusted_header",
            "trusted_header": {
                "user_id": "X-SSO-User-Id",
                "username": "X-SSO-User-Name",
            },
        }

        response = self.client.get(
            "/admin/tasks",
            headers={"X-SSO-User-Id": "100086", "X-SSO-User-Name": "张三"},
        )

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn("model-a", html)
        self.assertIn(f'value="{model_id}"', html)

    def test_consistency_check_items_are_unchecked_by_default(self):
        self._configure_provider()

        response = self.client.get("/consistency")

        self.assertEqual(response.status_code, 200)
        soup = BeautifulSoup(response.get_data(as_text=True), "html.parser")
        form = _required_tag(soup.find("form", {"data-require-checks": "true"}))
        self.assertEqual(form.get("autocomplete"), "off")
        self.assertEqual(form.get("data-default-unchecked-checks"), "true")
        checkboxes = form.select('input[name="checks"]')
        self.assertTrue(checkboxes)
        self.assertTrue(all(checkbox.get("checked") is None for checkbox in checkboxes))
        self.assertTrue(all(checkbox.get("autocomplete") == "off" for checkbox in checkboxes))

    def test_language_consistency_check_items_are_checked_by_default(self):
        self._configure_provider()

        response = self.client.get("/language-consistency")

        self.assertEqual(response.status_code, 200)
        soup = BeautifulSoup(response.get_data(as_text=True), "html.parser")
        form = _required_tag(soup.find("form", {"data-require-checks": "true"}))
        self.assertEqual(form.get("data-check-required-message"), "请至少选择一个跨语种检查项。")
        self.assertEqual(form.get("data-prevent-double-submit"), "true")
        self.assertEqual(form.get("data-submitting-label"), "提交中...")
        self.assertIn("请勿重复提交", form.get("data-submitting-message", ""))
        submission_token = _required_tag(form.select_one('input[name="submission_token"]'))
        self.assertRegex(submission_token.get("value", ""), r"^[0-9a-f]{32}$")
        progress = _required_tag(form.select_one("[data-submit-progress]"))
        self.assertTrue(progress.has_attr("hidden"))
        self.assertIsNotNone(form.select_one('input[name="document_a"]'))
        self.assertIsNotNone(form.select_one('input[name="document_b"]'))
        checkboxes = form.select('input[name="checks"]')
        self.assertTrue(checkboxes)
        self.assertTrue(all(checkbox.get("checked") is not None for checkbox in checkboxes))

    def test_saml_user_page_redirects_to_saml_login(self):
        self.app.config["AUTH"] = _saml_auth_config()

        response = self.client.get("/")

        self.assertEqual(response.status_code, 302)
        self.assertIn("/auth/saml/login?next=/", response.headers["Location"])

    def test_saml_login_stores_request_id(self):
        self.app.config["AUTH"] = _saml_auth_config()
        fake_auth = _FakeSamlAuth()

        with patch("app.routes.create_saml_auth", return_value=fake_auth):
            response = self.client.get("/auth/saml/login?next=/consistency")

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers["Location"], "https://sso.example.com/login?SAMLRequest=test")
        self.assertEqual(fake_auth.return_to, "/consistency")
        with self.client.session_transaction() as session:
            self.assertEqual(session["saml_request_id"], "REQ-1")

    def test_saml_acs_saves_session_identity(self):
        self.app.config["AUTH"] = _saml_auth_config()
        fake_auth = _FakeSamlAuth()
        with self.client.session_transaction() as session:
            session["saml_request_id"] = "REQ-1"

        with patch("app.routes.create_saml_auth", return_value=fake_auth):
            response = self.client.post(
                "/auth/saml/acs",
                data={"SAMLResponse": "test", "RelayState": "/consistency"},
            )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers["Location"], "/consistency")
        self.assertEqual(fake_auth.processed_request_id, "REQ-1")
        with self.client.session_transaction() as session:
            self.assertEqual(session[SAML_USER_SESSION_KEY], {"user_id": "100086", "username": "张三"})
            self.assertNotIn("saml_request_id", session)

    def test_create_task_uses_saml_session_identity(self):
        model_id = self._configure_provider("saml:100086")
        self.app.config["AUTH"] = _saml_auth_config()
        with self.client.session_transaction() as session:
            session[SAML_USER_SESSION_KEY] = {"user_id": "100086", "username": "张三"}
        with self.app.app_context():
            item = get_db().execute("SELECT id FROM check_items WHERE code = 'typo'").fetchone()

        response = self.client.post(
            "/",
            data={
                "document": (io.BytesIO("测试文档".encode("utf-8")), "doc.txt"),
                "checks": [str(item["id"])],
                "model_id": model_id,
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            task = get_db().execute("SELECT owner_subject, owner_name_snapshot, owner_source FROM tasks").fetchone()
        self.assertEqual(task["owner_subject"], "saml:100086")
        self.assertEqual(task["owner_name_snapshot"], "张三")
        self.assertEqual(task["owner_source"], "saml")

    def test_saml_metadata_uses_sp_config_only(self):
        self.app.config["AUTH"] = _saml_auth_config()

        response = self.client.get("/auth/saml/metadata")

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn("EntityDescriptor", html)
        self.assertIn("https://doc.example.com/auth/saml/metadata", html)
        self.assertIn("https://doc.example.com/auth/saml/acs", html)

    def test_saml_admin_settings_still_uses_local_admin_login(self):
        self.app.config["AUTH"] = _saml_auth_config()

        response = self.client.get("/admin/settings")

        self.assertEqual(response.status_code, 200)
        self.assertIn("系统设置", response.get_data(as_text=True))

    def test_saml_admin_task_page_redirects_to_saml_login_for_same_user(self):
        self.app.config["AUTH"] = _saml_auth_config()

        response = self.client.get("/admin/tasks")

        self.assertEqual(response.status_code, 302)
        self.assertIn("/auth/saml/login?next=/admin/tasks", response.headers["Location"])

    def test_create_consistency_task_rejects_missing_checks_before_saving_file(self):
        model_id = self._configure_provider()

        response = self.client.post(
            "/consistency",
            data={
                "master_documents": (io.BytesIO("素材参数 10A".encode("utf-8")), "master.txt"),
                "related_documents": (io.BytesIO("资料参数 12A".encode("utf-8")), "related.txt"),
                "model_id": model_id,
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            total = get_db().execute("SELECT COUNT(*) AS total FROM tasks").fetchone()["total"]
        self.assertEqual(total, 0)
        self.assertEqual(list(Path(self.app.config["UPLOAD_FOLDER"]).iterdir()), [])

    def test_create_consistency_task_saves_combined_document_text(self):
        model_id = self._configure_provider()
        with self.app.app_context():
            item = get_db().execute(
                "SELECT id, code, name, prompt FROM check_items WHERE code = 'consistency-cross-document'"
            ).fetchone()

        response = self.client.post(
            "/consistency",
            data={
                "master_documents": (_xlsx_bytes([["项目", "参数"], ["素材参数", "10A"]], title="素材参数表"), "master.xlsx"),
                "related_documents": (io.BytesIO("资料参数 12A".encode("utf-8")), "related.txt"),
                "checks": [str(item["id"])],
                "model_id": model_id,
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            task = get_db().execute(
                "SELECT task_type, original_filename, document_text, checks_snapshot_json FROM tasks"
            ).fetchone()
        self.assertEqual(task["task_type"], "consistency_check")
        self.assertEqual(task["original_filename"], "素材文档：master.xlsx / 资料：related.txt")
        self.assertIn("## 素材文档1：master.xlsx", task["document_text"])
        self.assertIn("# 工作表：素材参数表", task["document_text"])
        self.assertIn("素材参数 | 10A", task["document_text"])
        self.assertIn("## 资料1：related.txt", task["document_text"])
        self.assertEqual(
            json.loads(task["checks_snapshot_json"]),
            [
                {
                    "id": item["id"],
                    "code": item["code"],
                    "name": item["name"],
                    "prompt": item["prompt"],
                }
            ],
        )
        page = self.client.get("/consistency")
        self.assertEqual(page.status_code, 200)
        self.assertIn("素材文档：master.xlsx / 资料：related.txt", page.get_data(as_text=True))

    def test_create_consistency_task_removes_files_when_database_insert_fails(self):
        model_id = self._configure_provider()
        with self.app.app_context():
            item = get_db().execute(
                "SELECT id FROM check_items WHERE code = 'consistency-cross-document'"
            ).fetchone()
        self._reject_task_inserts()

        response = self.client.post(
            "/consistency",
            data={
                "master_documents": (io.BytesIO("素材参数 10A".encode("utf-8")), "master.txt"),
                "related_documents": (io.BytesIO("资料参数 12A".encode("utf-8")), "related.txt"),
                "checks": [str(item["id"])],
                "model_id": model_id,
            },
            content_type="multipart/form-data",
            follow_redirects=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn("创建多文档对照任务失败", response.get_data(as_text=True))
        with self.app.app_context():
            total = get_db().execute("SELECT COUNT(*) AS total FROM tasks").fetchone()["total"]
        self.assertEqual(total, 0)
        self.assertEqual(list(Path(self.app.config["UPLOAD_FOLDER"]).iterdir()), [])

    def test_consistency_task_title_uses_document_metadata_for_legacy_task(self):
        task = {
            "original_filename": "多文档对照检查：素材3个 / 资料1个",
            "document_meta_json": json.dumps(
                {
                    "groups": [
                        {
                            "role": "master",
                            "label": "素材文档",
                            "files": [
                                {"original_filename": "需求说明.docx"},
                                {"original_filename": "参数表.xlsx"},
                                {"original_filename": "会议纪要.pdf"},
                            ],
                        },
                        {
                            "role": "related",
                            "label": "资料",
                            "files": [{"original_filename": "投标文件.docx"}],
                        },
                    ]
                },
                ensure_ascii=False,
            ),
        }

        self.assertEqual(
            _consistency_task_title(task),
            "素材文档：需求说明.docx、参数表.xlsx 等3个 / 资料：投标文件.docx",
        )
        self.assertEqual(
            _consistency_task_title(task, include_all=True),
            "素材文档：需求说明.docx、参数表.xlsx、会议纪要.pdf / 资料：投标文件.docx",
        )

    def test_create_language_consistency_task_rejects_missing_checks_before_saving_file(self):
        model_id = self._configure_provider()

        response = self.client.post(
            "/language-consistency",
            data={
                "document_a": (io.BytesIO("中文参数 10A".encode("utf-8")), "zh.txt"),
                "document_b": (io.BytesIO("English parameter 10A".encode("utf-8")), "en.txt"),
                "model_id": model_id,
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            total = get_db().execute("SELECT COUNT(*) AS total FROM tasks").fetchone()["total"]
        self.assertEqual(total, 0)
        self.assertEqual(list(Path(self.app.config["UPLOAD_FOLDER"]).iterdir()), [])

    def test_create_language_consistency_task_saves_static_precheck(self):
        model_id = self._configure_provider()
        with self.app.app_context():
            item = get_db().execute(
                "SELECT id, code, name, prompt FROM check_items WHERE code = 'language-consistency-cross-lingual'"
            ).fetchone()

        response = self.client.post(
            "/language-consistency",
            data={
                "document_a": (
                    io.BytesIO("1. 安装要求\n设备电流为 10A。\n访问 https://example.com/a。".encode("utf-8")),
                    "zh.txt",
                ),
                "document_b": (
                    io.BytesIO("1. Installation requirements\nThe device current is 12A.\nVisit https://example.com/b.".encode("utf-8")),
                    "en.txt",
                ),
                "checks": [str(item["id"])],
                "model_id": model_id,
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            task = get_db().execute(
                "SELECT task_type, original_filename, file_type, document_text, document_meta_json, checks_snapshot_json FROM tasks"
            ).fetchone()
        self.assertEqual(task["task_type"], LANGUAGE_CONSISTENCY_TASK_TYPE)
        self.assertEqual(task["file_type"], "双文档")
        self.assertIn("跨语种检查：zh.txt / en.txt", task["original_filename"])
        self.assertIn("# 静态预检摘要", task["document_text"])
        self.assertIn("文档A独有硬线索", task["document_text"])
        self.assertIn("文档B独有硬线索", task["document_text"])
        self.assertIn("10a", task["document_text"])
        self.assertIn("12a", task["document_text"])
        self.assertIn("# 文档A：zh.txt", task["document_text"])
        self.assertIn("# 文档B：en.txt", task["document_text"])
        meta = json.loads(task["document_meta_json"])
        self.assertEqual([group["role"] for group in meta["groups"]], ["document_a", "document_b"])
        self.assertIn("文档A独有硬线索", meta["static_precheck"])
        self.assertEqual(
            json.loads(task["checks_snapshot_json"]),
            [
                {
                    "id": item["id"],
                    "code": item["code"],
                    "name": item["name"],
                    "prompt": item["prompt"],
                }
            ],
        )

    def test_create_language_consistency_task_removes_files_when_database_insert_fails(self):
        model_id = self._configure_provider()
        with self.app.app_context():
            item = get_db().execute(
                "SELECT id FROM check_items WHERE code = 'language-consistency-cross-lingual'"
            ).fetchone()
        self._reject_task_inserts()

        response = self.client.post(
            "/language-consistency",
            data={
                "document_a": (io.BytesIO("中文参数 10A".encode("utf-8")), "zh.txt"),
                "document_b": (io.BytesIO("English parameter 10A".encode("utf-8")), "en.txt"),
                "checks": [str(item["id"])],
                "model_id": model_id,
                "submission_token": "b" * 32,
            },
            content_type="multipart/form-data",
            follow_redirects=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn("创建跨语种检查任务失败", response.get_data(as_text=True))
        with self.app.app_context():
            total = get_db().execute("SELECT COUNT(*) AS total FROM tasks").fetchone()["total"]
        self.assertEqual(total, 0)
        self.assertEqual(list(Path(self.app.config["UPLOAD_FOLDER"]).iterdir()), [])

    def test_duplicate_language_consistency_submission_creates_one_task(self):
        model_id = self._configure_provider()
        submission_token = "a" * 32
        with self.app.app_context():
            item = get_db().execute(
                "SELECT id FROM check_items WHERE code = 'language-consistency-cross-lingual'"
            ).fetchone()

        def submit():
            return self.client.post(
                "/language-consistency",
                data={
                    "document_a": (io.BytesIO("中文参数 10A".encode("utf-8")), "zh.txt"),
                    "document_b": (io.BytesIO("English parameter 10A".encode("utf-8")), "en.txt"),
                    "checks": [str(item["id"])],
                    "model_id": model_id,
                    "submission_token": submission_token,
                },
                content_type="multipart/form-data",
            )

        first_response = submit()
        second_response = submit()

        self.assertEqual(first_response.status_code, 302)
        self.assertEqual(second_response.status_code, 302)
        with self.app.app_context():
            tasks = get_db().execute(
                "SELECT submission_token FROM tasks WHERE task_type = ?",
                (LANGUAGE_CONSISTENCY_TASK_TYPE,),
            ).fetchall()
        self.assertEqual(len(tasks), 1)
        self.assertEqual(tasks[0]["submission_token"], submission_token)
        self.assertEqual(len(list(Path(self.app.config["UPLOAD_FOLDER"]).iterdir())), 2)

if __name__ == "__main__":
    unittest.main()
