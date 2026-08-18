import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from app.common_terms import CommonTermRule, build_common_terms_report, load_common_terms


class CommonTermsTest(unittest.TestCase):
    def test_loads_workbook_and_splits_multiple_discouraged_terms(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "common_terms.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            assert sheet is not None
            sheet.append(["常用词", "常见错误/不推荐用法"])
            sheet.append(["OpenAI", "Open AI；open-ai、OPEN AI"])
            sheet.append(["登录", "登陆"])
            workbook.save(path)
            workbook.close()

            rules = load_common_terms(path)

        self.assertEqual(len(rules), 2)
        self.assertEqual(rules[0].standard, "OpenAI")
        self.assertEqual(rules[0].discouraged, ("Open AI", "open-ai", "OPEN AI"))
        self.assertEqual(rules[0].language_scope, "all")
        self.assertEqual(rules[1].discouraged, ("登陆",))

    def test_loads_optional_language_scope_and_keeps_scopes_separate(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "common_terms.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            assert sheet is not None
            sheet.append(["常用词", "常见错误/不推荐用法", "适用语种"])
            sheet.append(["App", "APP；app", "中文"])
            sheet.append(["App", "APP EN", "全部"])
            workbook.save(path)
            workbook.close()

            rules = load_common_terms(path)

        self.assertEqual(len(rules), 2)
        self.assertEqual(rules[0].language_scope, "zh")
        self.assertEqual(rules[1].language_scope, "all")

    def test_rejects_unsupported_language_scope(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "common_terms.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            assert sheet is not None
            sheet.append(["常用词", "常见错误/不推荐用法", "适用语种"])
            sheet.append(["App", "APP；app", "英文"])
            workbook.save(path)
            workbook.close()

            with self.assertRaisesRegex(ValueError, "目前支持留空/全部/all，或中文/zh"):
                load_common_terms(path)

    def test_reports_discouraged_terms_and_incorrect_case(self):
        rules = [
            CommonTermRule("OpenAI", ("Open AI", "open-ai")),
            CommonTermRule("登录", ("登陆",)),
        ]
        document_text = (
            "file: doc.txt\n\n[第3页]\n"
            "OpenAI 是正确写法，openai 和 OPENAI 大小写错误，Open AI 不推荐。\n"
            "openaiService 是更长的标识符，不应按子串命中。请先登陆。"
        )

        report = build_common_terms_report(
            document_text,
            rules,
            source_path=Path("common_terms.xlsx"),
            issue_limit=20,
        )

        self.assertIn("发现 4 类常用词写法问题", report["summary"])
        self.assertEqual(len(report["items"]), 4)
        descriptions = "\n".join(item["description"] for item in report["items"])
        self.assertIn("“openai”与常用词表规定的大小写“OpenAI”不一致", descriptions)
        self.assertIn("“OPENAI”与常用词表规定的大小写“OpenAI”不一致", descriptions)
        self.assertIn("不推荐用法“Open AI”", descriptions)
        self.assertIn("不推荐用法“登陆”", descriptions)
        self.assertNotIn("openaiService", descriptions)
        self.assertTrue(all("文件：doc.txt" in item["location"] for item in report["items"]))
        self.assertTrue(all("页码：第3页" in item["location"] for item in report["items"]))

    def test_accepts_exact_case_and_does_not_match_inside_longer_identifier(self):
        rules = [CommonTermRule("OpenAI", ("Open AI",))]
        document_text = "OpenAI OpenAIService preOpenAI openaiService"

        report = build_common_terms_report(
            document_text,
            rules,
            source_path=Path("common_terms.xlsx"),
            issue_limit=20,
        )

        self.assertEqual(report["items"], [])
        self.assertIn("未发现错误、不推荐用法或大小写不一致问题", report["summary"])

    def test_exact_standard_is_not_rejected_when_repeated_in_discouraged_cell(self):
        rules = [CommonTermRule("OpenAI", ("OpenAI", "Open AI"))]

        report = build_common_terms_report(
            "OpenAI",
            rules,
            source_path=Path("common_terms.xlsx"),
            issue_limit=20,
        )

        self.assertEqual(report["items"], [])

    def test_chinese_only_rule_checks_discouraged_terms_and_case_in_chinese_document(self):
        rules = [CommonTermRule("App", ("APP", "app"), language_scope="zh")]
        document_text = (
            "这是面向客户发布的中文操作指南，介绍移动应用的安装、登录、配置、使用和维护方法。"
            "用户应按照以下步骤完成操作，并在操作完成后检查系统状态和业务结果。"
            "请打开 APP，然后在 app 中选择设置。"
        )

        report = build_common_terms_report(
            document_text,
            rules,
            source_path=Path("common_terms.xlsx"),
            issue_limit=20,
        )

        self.assertEqual(len(report["items"]), 2)
        descriptions = "\n".join(item["description"] for item in report["items"])
        self.assertIn("不推荐用法“APP”", descriptions)
        self.assertIn("不推荐用法“app”", descriptions)
        self.assertIn("文档语种估计：中文为主", report["summary"])
        self.assertNotIn("已跳过", report["summary"])

    def test_model_identifiers_do_not_skip_chinese_only_rules(self):
        rules = [CommonTermRule("App", ("APP",), language_scope="zh")]
        document_text = (
            "这是中文产品资料，用于说明设备安装、配置、操作、验证和维护方法。" * 3
            + "SUN2000-50KTL-M3 S5735-L48T4X-A1 NetCol5000-A UPS5000-E V100R001C00 " * 4
            + "请打开 APP 并检查设备状态。"
        )

        report = build_common_terms_report(
            document_text,
            rules,
            source_path=Path("common_terms.xlsx"),
            issue_limit=20,
        )

        self.assertEqual(len(report["items"]), 1)
        self.assertIn("不推荐用法“APP”", report["items"][0]["description"])
        self.assertIn("文档语种估计：中文为主", report["summary"])
        self.assertNotIn("已跳过", report["summary"])

    def test_chinese_only_rule_is_skipped_for_english_document(self):
        rules = [CommonTermRule("App", ("APP", "app"), language_scope="zh")]
        document_text = (
            "This English customer guide explains how to install, configure, operate, and maintain "
            "the mobile APP. Open the app, select Settings, save the changes, and verify the result."
        )

        report = build_common_terms_report(
            document_text,
            rules,
            source_path=Path("common_terms.xlsx"),
            issue_limit=20,
        )

        self.assertEqual(report["items"], [])
        self.assertIn("文档语种估计：拉丁语系为主", report["summary"])
        self.assertIn("已跳过 1 条不适用于当前文档语种", report["summary"])

    def test_chinese_only_scope_also_controls_automatic_case_check(self):
        rules = [CommonTermRule("App", language_scope="zh")]
        chinese_text = (
            "这是中文客户操作指南，用于说明移动应用的安装、配置、使用、验证和维护方法。"
            "用户完成准备工作后，请打开 APP 并检查运行状态。"
        )
        english_text = (
            "This English customer guide explains installation, configuration, operation, verification, "
            "and maintenance. Open the APP and check the operating status."
        )

        chinese_report = build_common_terms_report(
            chinese_text,
            rules,
            source_path=Path("common_terms.xlsx"),
            issue_limit=20,
        )
        english_report = build_common_terms_report(
            english_text,
            rules,
            source_path=Path("common_terms.xlsx"),
            issue_limit=20,
        )

        self.assertEqual(len(chinese_report["items"]), 1)
        self.assertIn("大小写“App”不一致", chinese_report["items"][0]["description"])
        self.assertEqual(english_report["items"], [])
        self.assertIn("已跳过 1 条不适用于当前文档语种", english_report["summary"])

    def test_chinese_only_rule_is_skipped_for_mixed_or_short_document(self):
        rules = [CommonTermRule("App", ("APP", "app"), language_scope="zh")]
        mixed_text = (
            "这是中文说明文字，用于介绍安装配置操作、使用限制、结果验证和维护注意事项。" * 2
            + " This English section explains installation configuration operation verification maintenance " * 3
            + "APP app"
        )

        mixed_report = build_common_terms_report(
            mixed_text,
            rules,
            source_path=Path("common_terms.xlsx"),
            issue_limit=20,
        )
        short_report = build_common_terms_report(
            "打开 APP 或 app。",
            rules,
            source_path=Path("common_terms.xlsx"),
            issue_limit=20,
        )

        self.assertEqual(mixed_report["items"], [])
        self.assertIn("文档语种估计：中英混合", mixed_report["summary"])
        self.assertEqual(short_report["items"], [])
        self.assertIn("语种特征较少", short_report["summary"])

    def test_chinese_only_rule_is_skipped_for_japanese_document(self):
        rules = [CommonTermRule("App", ("APP", "app"), language_scope="zh")]
        document_text = (
            "この文書では、製品のインストール、設定、操作、確認、および保守手順について説明します。"
            "APP を開き、app の設定画面で必要な項目を選択して、処理結果を確認してください。"
        )

        report = build_common_terms_report(
            document_text,
            rules,
            source_path=Path("common_terms.xlsx"),
            issue_limit=20,
        )

        self.assertEqual(report["items"], [])
        self.assertIn("文档语种估计：其他语种为主", report["summary"])
        self.assertIn("已跳过 1 条不适用于当前文档语种", report["summary"])

    def test_all_language_rule_still_checks_english_document(self):
        rules = [CommonTermRule("OpenAI", ("Open AI",), language_scope="all")]
        document_text = (
            "This English document describes the Open AI service configuration, request parameters, "
            "response fields, error handling, and operational verification for customer deployments."
        )

        report = build_common_terms_report(
            document_text,
            rules,
            source_path=Path("common_terms.xlsx"),
            issue_limit=20,
        )

        self.assertEqual(len(report["items"]), 1)
        self.assertIn("不推荐用法“Open AI”", report["items"][0]["description"])


if __name__ == "__main__":
    unittest.main()
