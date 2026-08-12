import csv
import json
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
import xlrd

from .term_locations import DocumentLocationIndex, TERM_LOCATIONS_PER_ISSUE, excerpt_for
from .text_language import TEXT_LANGUAGE_CHINESE, estimate_text_language, text_language_label


COMMON_TERMS_CHECK_CODE = "common-terms"
COMMON_TERMS_FILENAMES = (
    "common_terms.xlsx",
    "common_terms.xlsm",
    "common_terms.xls",
    "common_terms.csv",
    "common_words.xlsx",
    "common_words.xlsm",
    "common_words.xls",
    "common_words.csv",
    "常用词检查表.xlsx",
    "常用词检查表.xlsm",
    "常用词检查表.xls",
    "常用词检查表.csv",
    "常用词表.xlsx",
    "常用词表.xlsm",
    "常用词表.xls",
    "常用词表.csv",
)

_STANDARD_HEADER_NAMES = {
    "常用词",
    "正确写法",
    "规范写法",
    "标准写法",
    "推荐写法",
}
_DISCOURAGED_HEADER_NAMES = {
    "常见错误/不推荐用法",
    "常见错误／不推荐用法",
    "常见错误和不推荐用法",
    "常见错误及不推荐用法",
    "常见错误",
    "错误用法",
    "不推荐用法",
}
_LANGUAGE_HEADER_NAMES = {
    "适用语种",
    "适用语言",
    "文档语种",
    "规则适用语种",
}
_EMPTY_DISCOURAGED_VALUES = {"", "-", "--", "—", "/", "无", "不涉及", "na", "n/a"}
_DISCOURAGED_SEPARATOR_RE = re.compile(r"[\r\n,，;；、]+")
_LANGUAGE_SCOPE_ALL = "all"
_LANGUAGE_SCOPE_CHINESE = TEXT_LANGUAGE_CHINESE
_LANGUAGE_SCOPE_ALL_VALUES = {
    "",
    "*",
    "all",
    "any",
    "全部",
    "所有",
    "不限",
    "通用",
    "全部/all",
    "all/全部",
}
_LANGUAGE_SCOPE_CHINESE_VALUES = {
    "zh",
    "zh-cn",
    "zh-hans",
    "zh-hant",
    "中文",
    "中文为主",
    "中文文档",
    "仅中文",
    "简体中文",
    "繁体中文",
    "中文/zh",
    "zh/中文",
}


@dataclass(frozen=True)
class CommonTermRule:
    standard: str
    discouraged: tuple[str, ...] = ()
    sheet: str = ""
    row_number: int = 0
    language_scope: str = _LANGUAGE_SCOPE_ALL


def find_common_terms_file(
    *,
    root_dir: Path,
    instance_dir: Path,
    configured_path: str | Path | None = None,
) -> Path | None:
    for path in common_terms_file_candidates(
        root_dir=root_dir,
        instance_dir=instance_dir,
        configured_path=configured_path,
    ):
        if path.is_file():
            return path
    return None


def common_terms_file_candidates(
    *,
    root_dir: Path,
    instance_dir: Path,
    configured_path: str | Path | None = None,
) -> list[Path]:
    candidates = []
    if configured_path:
        candidates.append(_path_from_config(configured_path, root_dir))

    for folder in (instance_dir, root_dir / "data"):
        for filename in COMMON_TERMS_FILENAMES:
            candidates.append(folder / filename)

    result = []
    seen = set()
    for path in candidates:
        key = str(path)
        if key in seen:
            continue
        seen.add(key)
        result.append(path)
    return result


def load_common_terms(path: Path) -> list[CommonTermRule]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _load_openpyxl_terms(path)
    if suffix == ".xls":
        return _load_xls_terms(path)
    if suffix == ".csv":
        return _load_csv_terms(path)
    raise ValueError("常用词检查表仅支持 xlsx、xlsm、xls、csv 格式。")


def build_common_terms_report(
    document_text: str,
    rules: list[CommonTermRule],
    *,
    source_path: Path,
    issue_limit: int,
) -> dict:
    language = estimate_text_language(document_text)
    active_rules = [rule for rule in rules if _rule_applies_to_language(rule, language)]
    skipped_rule_count = len(rules) - len(active_rules)
    matches = _find_common_term_matches(document_text, active_rules)
    context_summary = _common_terms_context_summary(
        source_path=source_path,
        language=language,
        skipped_rule_count=skipped_rule_count,
    )
    if not matches:
        return {
            "summary": (
                "常用词检查结论：未发现错误、不推荐用法或大小写不一致问题。\n"
                f"{context_summary}"
            ),
            "items": [],
        }

    visible_matches = matches[: max(1, issue_limit)]
    hidden_count = max(0, len(matches) - len(visible_matches))
    occurrence_total = sum(match["count"] for match in matches)
    summary = (
        f"常用词检查结论：发现 {len(matches)} 类常用词写法问题，合计 {occurrence_total} 处命中。\n"
        f"{context_summary}"
    )
    if hidden_count:
        summary += f"\n受报告条目上限限制，仅展示前 {len(visible_matches)} 类，另有 {hidden_count} 类未展开。"

    items = []
    for match in visible_matches:
        if match["kind"] == "case":
            category = "常用词/大小写不一致"
            description = (
                f"文档中的“{match['observed']}”与常用词表规定的大小写“{match['standard']}”不一致，"
                f"共 {match['count']} 处。"
            )
            impact = "大小写不符合统一术语、产品名称或品牌写法要求。"
        else:
            category = "常用词/错误或不推荐用法"
            description = (
                f"文档中出现常见错误或不推荐用法“{match['observed']}”，"
                f"共 {match['count']} 处。"
            )
            impact = "可能造成术语不统一、表达不规范或品牌写法不一致。"
        items.append(
            {
                "category": category,
                "location": "；".join(match["locations"]),
                "excerpt": "；".join(match["excerpts"]),
                "description": description,
                "impact": impact,
                "suggestion": f"请统一修改为词表中的正确写法：{match['standard']}。",
                "type": "issue",
            }
        )
    return {"summary": summary, "items": items}


def build_common_terms_missing_report(*, candidates: list[Path]) -> dict:
    candidate_text = "\n".join(f"- {path}" for path in candidates[:8])
    return {
        "summary": "常用词检查结论：未配置常用词检查表，未执行规则匹配。",
        "items": [
            {
                "category": "常用词检查表配置",
                "location": "本地词表文件",
                "excerpt": "",
                "description": (
                    "未找到可读取的常用词检查表。系统需要读取“常用词”和"
                    "“常见错误/不推荐用法”两列后才能执行检查。"
                ),
                "impact": "本次任务无法判断常用词写法和大小写是否符合词表要求。",
                "suggestion": (
                    "请将检查表放到以下任一路径，推荐使用 instance/common_terms.xlsx：\n"
                    f"{candidate_text}"
                ),
                "type": "suggestion",
            }
        ],
    }


def build_common_terms_invalid_report(*, source_path: Path, error: str) -> dict:
    return {
        "summary": f"常用词检查结论：常用词检查表读取失败。\n词表文件：{source_path}",
        "items": [
            {
                "category": "常用词检查表读取失败",
                "location": str(source_path),
                "excerpt": "",
                "description": str(error or "常用词检查表无法读取。"),
                "impact": "本次任务无法完成常用词规则匹配。",
                "suggestion": (
                    "请确认检查表格式为 xlsx、xlsm、xls 或 csv，并包含"
                    "“常用词”和“常见错误/不推荐用法”两列。"
                ),
                "type": "suggestion",
            }
        ],
    }


def format_common_terms_report(report: dict) -> str:
    return json.dumps(report, ensure_ascii=False, indent=2)


def _path_from_config(path: str | Path, root_dir: Path) -> Path:
    result = Path(path)
    if result.is_absolute():
        return result
    return root_dir / result


def _load_openpyxl_terms(path: Path) -> list[CommonTermRule]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return _dedupe_rules(
            rule
            for sheet in workbook.worksheets
            for rule in _rules_from_rows(sheet.iter_rows(values_only=True), sheet.title)
        )
    finally:
        workbook.close()


def _load_xls_terms(path: Path) -> list[CommonTermRule]:
    workbook = xlrd.open_workbook(str(path), on_demand=True)
    try:
        rules = []
        for sheet in workbook.sheets():
            rows = (
                [sheet.cell_value(row_index, column_index) for column_index in range(sheet.ncols)]
                for row_index in range(sheet.nrows)
            )
            rules.extend(_rules_from_rows(rows, sheet.name))
        return _dedupe_rules(rules)
    finally:
        workbook.release_resources()


def _load_csv_terms(path: Path) -> list[CommonTermRule]:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            with path.open("r", encoding=encoding, newline="") as file:
                return _dedupe_rules(_rules_from_rows(csv.reader(file), "CSV"))
        except UnicodeDecodeError:
            continue
    raise ValueError("无法识别 CSV 编码，请使用 UTF-8 或 GB18030。")


def _rules_from_rows(rows, sheet_name: str):
    header = None
    for row_number, row in enumerate(rows, start=1):
        values = list(row or [])
        if header is None:
            header = _header_mapping(values)
            continue
        standard = _row_cell(values, header["standard"])
        if not standard:
            continue
        discouraged = _split_discouraged(_row_cell(values, header["discouraged"]))
        language_scope = _normalize_language_scope(
            _row_cell(values, header.get("language")),
            sheet_name=sheet_name,
            row_number=row_number,
        )
        yield CommonTermRule(
            standard=standard,
            discouraged=tuple(discouraged),
            language_scope=language_scope,
            sheet=sheet_name,
            row_number=row_number,
        )


def _header_mapping(row: list) -> dict | None:
    standard_index = None
    discouraged_index = None
    language_index = None
    for index, value in enumerate(row):
        kind = _header_kind(value)
        if kind == "standard" and standard_index is None:
            standard_index = index
        if kind == "discouraged" and discouraged_index is None:
            discouraged_index = index
        if kind == "language" and language_index is None:
            language_index = index
    if standard_index is None or discouraged_index is None:
        return None
    return {
        "standard": standard_index,
        "discouraged": discouraged_index,
        "language": language_index,
    }


def _header_kind(value) -> str:
    text = _compact_header(value)
    if not text:
        return ""
    if text in _DISCOURAGED_HEADER_NAMES or "常见错误" in text or "不推荐用法" in text:
        return "discouraged"
    if text in _STANDARD_HEADER_NAMES:
        return "standard"
    if text in _LANGUAGE_HEADER_NAMES:
        return "language"
    return ""


def _compact_header(value) -> str:
    return re.sub(r"[\s:：*＊（）()【】\[\]<>《》]+", "", _cell_text(value))


def _row_cell(row: list, index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    return _cell_text(row[index])


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _split_discouraged(value: str) -> list[str]:
    values = []
    seen = set()
    for part in _DISCOURAGED_SEPARATOR_RE.split(str(value or "")):
        term = part.strip()
        if term.lower() in _EMPTY_DISCOURAGED_VALUES or term in seen:
            continue
        seen.add(term)
        values.append(term)
    return values


def _normalize_language_scope(value: str, *, sheet_name: str, row_number: int) -> str:
    normalized = re.sub(r"[_\s]+", "-", str(value or "").strip().lower())
    if normalized in _LANGUAGE_SCOPE_ALL_VALUES:
        return _LANGUAGE_SCOPE_ALL
    if normalized in _LANGUAGE_SCOPE_CHINESE_VALUES:
        return _LANGUAGE_SCOPE_CHINESE
    raise ValueError(
        f"工作表“{sheet_name}”第 {row_number} 行的“适用语种”值“{value}”无效；"
        "目前支持留空/全部/all，或中文/zh。"
    )


def _dedupe_rules(rules) -> list[CommonTermRule]:
    result = []
    positions = {}
    for rule in rules:
        if not rule.standard:
            continue
        key = (rule.standard, rule.language_scope)
        if key not in positions:
            positions[key] = len(result)
            result.append(rule)
            continue
        index = positions[key]
        existing = result[index]
        discouraged = tuple(dict.fromkeys(existing.discouraged + rule.discouraged))
        result[index] = CommonTermRule(
            standard=existing.standard,
            discouraged=discouraged,
            language_scope=existing.language_scope,
            sheet=existing.sheet,
            row_number=existing.row_number,
        )
    return result


def _rule_applies_to_language(rule: CommonTermRule, language: str) -> bool:
    if rule.language_scope == _LANGUAGE_SCOPE_ALL:
        return True
    return rule.language_scope == language


def _common_terms_context_summary(*, source_path: Path, language: str, skipped_rule_count: int) -> str:
    lines = [
        f"文档语种估计：{text_language_label(language)}。",
        f"词表文件：{source_path}",
    ]
    if skipped_rule_count:
        lines.append(f"已跳过 {skipped_rule_count} 条不适用于当前文档语种的常用词规则。")
    return "\n".join(lines)


def _find_common_term_matches(document_text: str, rules: list[CommonTermRule]) -> list[dict]:
    text = str(document_text or "")
    candidates = []
    for rule in rules:
        for discouraged in rule.discouraged:
            if discouraged == rule.standard:
                continue
            for match in _term_pattern(discouraged).finditer(text):
                candidates.append(_candidate(match, rule.standard, "discouraged"))
        if _has_case(rule.standard):
            for match in _term_pattern(rule.standard, ignore_case=True).finditer(text):
                if match.group(0) != rule.standard:
                    candidates.append(_candidate(match, rule.standard, "case"))

    selected = []
    last_end = -1
    for candidate in sorted(
        candidates,
        key=lambda item: (
            item["start"],
            -(item["end"] - item["start"]),
            0 if item["kind"] == "discouraged" else 1,
            item["standard"],
        ),
    ):
        if candidate["start"] < last_end:
            continue
        selected.append(candidate)
        last_end = candidate["end"]

    document_index = DocumentLocationIndex(text)
    grouped = {}
    for candidate in selected:
        key = (candidate["observed"], candidate["standard"], candidate["kind"])
        match = grouped.setdefault(
            key,
            {
                "observed": candidate["observed"],
                "standard": candidate["standard"],
                "kind": candidate["kind"],
                "count": 0,
                "first_position": candidate["start"],
                "locations": [],
                "excerpts": [],
            },
        )
        match["count"] += 1
        if len(match["locations"]) < TERM_LOCATIONS_PER_ISSUE:
            match["locations"].append(
                document_index.location_for(candidate["start"], candidate["observed"])
            )
            match["excerpts"].append(
                excerpt_for(text, candidate["start"], len(candidate["observed"]))
            )
    return sorted(grouped.values(), key=lambda item: (item["first_position"], item["observed"]))


def _candidate(match: re.Match, standard: str, kind: str) -> dict:
    return {
        "start": match.start(),
        "end": match.end(),
        "observed": match.group(0),
        "standard": standard,
        "kind": kind,
    }


def _has_case(term: str) -> bool:
    return any(character.lower() != character.upper() for character in term)


def _term_pattern(term: str, *, ignore_case: bool = False) -> re.Pattern:
    prefix = r"(?<![A-Za-z0-9_])" if _is_ascii_word_character(term[0]) else ""
    suffix = r"(?![A-Za-z0-9_])" if _is_ascii_word_character(term[-1]) else ""
    flags = re.IGNORECASE if ignore_case else 0
    return re.compile(prefix + re.escape(term) + suffix, flags)


def _is_ascii_word_character(character: str) -> bool:
    return character.isascii() and (character.isalnum() or character == "_")
