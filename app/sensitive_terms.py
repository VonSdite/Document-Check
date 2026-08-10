import csv
import json
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
import xlrd

from .term_locations import DocumentLocationIndex, TERM_LOCATIONS_PER_ISSUE, excerpt_for


SENSITIVE_TERMS_CHECK_CODE = "sensitive-terms"
SENSITIVE_TERMS_FILENAMES = (
    "sensitive_terms.xlsx",
    "sensitive_terms.xlsm",
    "sensitive_terms.xls",
    "sensitive_terms.csv",
    "sensitive_words.xlsx",
    "sensitive_words.xlsm",
    "sensitive_words.xls",
    "sensitive_words.csv",
)
_INVALID_HEADER_NAMES = {
    "不规范用语",
    "不规范词",
    "不规范表达",
    "敏感词",
    "敏感用语",
    "禁用词",
    "原词",
    "原用语",
}
_STANDARD_HEADER_NAMES = {
    "规范用语",
    "规范词",
    "规范表达",
    "标准用语",
    "推荐用语",
    "建议用语",
    "替换为",
}
@dataclass(frozen=True)
class SensitiveTermRule:
    invalid: str
    standard: str
    sheet: str = ""
    row_number: int = 0


def find_sensitive_terms_file(
    *,
    root_dir: Path,
    instance_dir: Path,
    configured_path: str | Path | None = None,
) -> Path | None:
    for path in sensitive_terms_file_candidates(
        root_dir=root_dir,
        instance_dir=instance_dir,
        configured_path=configured_path,
    ):
        if path.is_file():
            return path
    return None


def sensitive_terms_file_candidates(
    *,
    root_dir: Path,
    instance_dir: Path,
    configured_path: str | Path | None = None,
) -> list[Path]:
    candidates = []
    if configured_path:
        candidates.append(_path_from_config(configured_path, root_dir))

    for folder in (instance_dir, root_dir / "data"):
        for filename in SENSITIVE_TERMS_FILENAMES:
            candidates.append(folder / filename)

    result = []
    seen = set()
    for path in candidates:
        resolved_key = str(path)
        if resolved_key in seen:
            continue
        seen.add(resolved_key)
        result.append(path)
    return result


def load_sensitive_terms(path: Path) -> list[SensitiveTermRule]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _load_openpyxl_terms(path)
    if suffix == ".xls":
        return _load_xls_terms(path)
    if suffix == ".csv":
        return _load_csv_terms(path)
    raise ValueError("敏感词表仅支持 xlsx、xlsm、xls、csv 格式。")


def build_sensitive_terms_report(
    document_text: str,
    rules: list[SensitiveTermRule],
    *,
    source_path: Path,
    issue_limit: int,
) -> dict:
    matches = _find_sensitive_term_matches(document_text, rules)
    if not matches:
        return {
            "summary": (
                "敏感词检查结论：未发现词表中的不规范用语。\n"
                f"词表文件：{source_path}"
            ),
            "items": [],
        }

    visible_matches = matches[: max(1, issue_limit)]
    hidden_count = max(0, len(matches) - len(visible_matches))
    occurrence_total = sum(match["count"] for match in matches)
    summary = (
        f"敏感词检查结论：发现 {len(matches)} 类不规范用语，合计 {occurrence_total} 处命中。\n"
        f"词表文件：{source_path}"
    )
    if hidden_count:
        summary += f"\n受报告条目上限限制，仅展示前 {len(visible_matches)} 类，另有 {hidden_count} 类未展开。"

    items = []
    for match in visible_matches:
        location_text = "；".join(match["locations"])
        items.append(
            {
                "category": "敏感词/不规范用语",
                "location": location_text,
                "excerpt": "；".join(match["excerpts"]),
                "description": (
                    f"文档中出现词表中的不规范用语“{match['invalid']}”，"
                    f"共 {match['count']} 处。"
                ),
                "impact": "可能不符合公司统一表述、品牌口径或对外发布用语要求。",
                "suggestion": f"建议统一替换为：{match['standard'] or '请按公司词表补充规范用语'}。",
                "type": "issue",
            }
        )
    return {"summary": summary, "items": items}


def build_sensitive_terms_missing_report(
    *,
    candidates: list[Path],
) -> dict:
    candidate_text = "\n".join(f"- {path}" for path in candidates[:6])
    return {
        "summary": "敏感词检查结论：未配置敏感词表，未执行敏感词匹配。",
        "items": [
            {
                "category": "敏感词表配置",
                "location": "本地词表文件",
                "excerpt": "",
                "description": (
                    "未找到可读取的敏感词表。系统需要从本地 Excel/CSV 词表中读取"
                    "“不规范用语”和“规范用语”两列后才能执行检查。"
                ),
                "impact": "本次任务无法判断文档是否包含公司内部维护的不规范用语。",
                "suggestion": (
                    "请将词表放到以下任一路径，推荐使用 instance/sensitive_terms.xlsx：\n"
                    f"{candidate_text}"
                ),
                "type": "suggestion",
            }
        ],
    }


def build_sensitive_terms_invalid_report(*, source_path: Path, error: str) -> dict:
    return {
        "summary": f"敏感词检查结论：敏感词表读取失败。\n词表文件：{source_path}",
        "items": [
            {
                "category": "敏感词表读取失败",
                "location": str(source_path),
                "excerpt": "",
                "description": str(error or "敏感词表无法读取。"),
                "impact": "本次任务无法完成敏感词匹配。",
                "suggestion": "请确认词表格式为 xlsx、xlsm、xls 或 csv，并包含“不规范用语”和“规范用语”两列。",
                "type": "suggestion",
            }
        ],
    }


def format_sensitive_terms_report(report: dict) -> str:
    return json.dumps(report, ensure_ascii=False, indent=2)


def _path_from_config(path: str | Path, root_dir: Path) -> Path:
    result = Path(path)
    if result.is_absolute():
        return result
    return root_dir / result


def _load_openpyxl_terms(path: Path) -> list[SensitiveTermRule]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return _dedupe_rules(
            rule
            for sheet in workbook.worksheets
            for rule in _rules_from_rows(sheet.iter_rows(values_only=True), sheet.title)
        )
    finally:
        workbook.close()


def _load_xls_terms(path: Path) -> list[SensitiveTermRule]:
    workbook = xlrd.open_workbook(str(path), on_demand=True)
    try:
        rules = []
        for sheet in workbook.sheets():
            rows = (
                [sheet.cell_value(row_index, column_index) for column_index in range(sheet.ncols)]
                for row_index in range(sheet.nrows)
            )
            rules.extend(_rules_from_rows(rows, sheet.name))
        return _dedupe_rules(rules)
    finally:
        workbook.release_resources()


def _load_csv_terms(path: Path) -> list[SensitiveTermRule]:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            with path.open("r", encoding=encoding, newline="") as file:
                return _dedupe_rules(_rules_from_rows(csv.reader(file), "CSV"))
        except UnicodeDecodeError:
            continue
    raise ValueError("无法识别 CSV 编码，请使用 UTF-8 或 GB18030。")


def _rules_from_rows(rows, sheet_name: str):
    header = None
    for row_number, row in enumerate(rows, start=1):
        values = list(row or [])
        if header is None:
            header = _header_mapping(values)
            continue
        invalid = _row_cell(values, header["invalid"])
        standard = _row_cell(values, header["standard"])
        if invalid:
            yield SensitiveTermRule(
                invalid=invalid,
                standard=standard,
                sheet=sheet_name,
                row_number=row_number,
            )


def _header_mapping(row: list) -> dict | None:
    invalid_index = None
    standard_index = None
    for index, value in enumerate(row):
        kind = _header_kind(value)
        if kind == "invalid" and invalid_index is None:
            invalid_index = index
        if kind == "standard" and standard_index is None:
            standard_index = index
    if invalid_index is None or standard_index is None:
        return None
    return {"invalid": invalid_index, "standard": standard_index}


def _header_kind(value) -> str:
    text = _compact_header(value)
    if not text:
        return ""
    if text in _INVALID_HEADER_NAMES or text.startswith("不规范"):
        return "invalid"
    if text in _STANDARD_HEADER_NAMES:
        return "standard"
    if "规范用语" in text and "不规范" not in text:
        return "standard"
    if "标准用语" in text or "推荐用语" in text or "建议用语" in text:
        return "standard"
    return ""


def _compact_header(value) -> str:
    return re.sub(r"[\s:：*＊（）()【】\[\]<>《》]+", "", _cell_text(value))


def _row_cell(row: list, index: int) -> str:
    if index >= len(row):
        return ""
    return _cell_text(row[index])


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _dedupe_rules(rules) -> list[SensitiveTermRule]:
    result = []
    seen = set()
    for rule in rules:
        key = (rule.invalid, rule.standard)
        if not rule.invalid or key in seen:
            continue
        seen.add(key)
        result.append(rule)
    return result


def _find_sensitive_term_matches(document_text: str, rules: list[SensitiveTermRule]) -> list[dict]:
    text = str(document_text or "")
    document_index = DocumentLocationIndex(text)
    matches = []
    for rule in rules:
        occurrences = list(_term_occurrences(text, rule.invalid))
        if not occurrences:
            continue
        samples = occurrences[:TERM_LOCATIONS_PER_ISSUE]
        matches.append(
            {
                "invalid": rule.invalid,
                "standard": rule.standard,
                "count": len(occurrences),
                "first_position": occurrences[0],
                "locations": [
                    document_index.location_for(position, rule.invalid)
                    for position in samples
                ],
                "excerpts": [
                    excerpt_for(text, position, len(rule.invalid))
                    for position in samples
                ],
            }
        )
    return sorted(matches, key=lambda item: (item["first_position"], item["invalid"]))


def _term_occurrences(text: str, term: str):
    if not term:
        return
    for match in re.finditer(re.escape(term), text):
        yield match.start()
