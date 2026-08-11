import hmac
import hashlib
import io
import json
import os
import re
import sqlite3
import uuid
from difflib import SequenceMatcher
import zipfile
from datetime import date, datetime, timedelta
from functools import wraps
from ipaddress import ip_address
from pathlib import Path
from urllib.parse import urlsplit

from flask import (
    abort,
    current_app,
    flash,
    redirect,
    render_template,
    request,
    Response,
    send_file,
    session,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from werkzeug.exceptions import RequestEntityTooLarge

from .auth import SAML_USER_SESSION_KEY, AuthenticationRequired, UserIdentity, current_identity, subject_label
from .config import save_network_config
from .db import (
    default_check_item_codes,
    delete_task_record,
    get_bool_setting,
    get_db,
    get_ip_username,
    get_setting,
    now_text,
    owner_subject_from_ip,
    reset_default_check_item_prompt,
    set_ip_username,
    set_setting,
)
from .documents import DocumentReadError, allowed_file, extension_of, extract_text, format_document_text
from .file_cleanup import (
    describe_failures,
    remove_directory_tree,
    remove_empty_directory as cleanup_remove_empty_directory,
    remove_file,
)
from .images import (
    DEFAULT_PDF_PAGE_IMAGE_MAX_PAGES,
    candidate_pdf_pages_for_image_check,
    default_image_folder,
    extract_images,
    format_image_document_text,
    image_items_from_meta,
    image_path_from_item,
    render_pdf_page_images,
)
from .limits import DEFAULT_ISSUE_OUTPUT_LIMIT, MAX_ISSUE_OUTPUT_LIMIT, normalize_issue_output_limit
from .llm import LLMError, test_model_connection
from .model_discovery import ModelDiscoveryError, fetch_models
from .network import outbound_network_config
from .saml import SamlConfigError, create_saml_auth, saml_sp_metadata
from .task_types import (
    CONSISTENCY_MAX_DATA_FILES,
    CONSISTENCY_MAX_MATERIAL_FILES,
    CONSISTENCY_TASK_TYPE,
    DOCUMENT_TASK_TYPE,
    IMAGE_TASK_TYPE,
    LANGUAGE_CONSISTENCY_TASK_TYPE,
    VIDEO_TASK_TYPE,
    document_groups_from_meta,
    task_type_label,
)
from .tasks import cleanup_task_file_cache, task_file_cache_snapshot
from .videos import allowed_video_file, extract_video_frames, format_video_document_text, video_extension_of


STATUS_LABELS = {
    "queued": "排队中",
    "running": "检查中",
    "completed": "已完成",
    "failed": "失败",
    "canceled": "已取消",
}
DELETABLE_TASK_STATUSES = {"completed", "failed", "canceled"}
TASKS_PER_PAGE = 20
CHECK_ITEM_CONCURRENCY_DEFAULT = 1
TASK_FILE_RETENTION_DAYS_DEFAULT = 0
ISSUE_OUTPUT_LIMIT_DEFAULT = DEFAULT_ISSUE_OUTPUT_LIMIT
UPLOAD_PATH_SAFE_CHARS = 240
UPLOAD_FILENAME_SAFE_CHARS = 180
PROVIDER_TIMEOUT_DEFAULT = 3600
PROVIDER_TIMEOUT_MIN = 30
PROVIDER_TIMEOUT_MAX = 7200
MODEL_TEST_TIMEOUT_MAX = 60
PROVIDER_INPUT_LIMIT_DEFAULT = 500000
PROVIDER_INPUT_LIMIT_MIN = 5000
PROVIDER_INPUT_LIMIT_MAX = 1000000
CONSOLE_USER_ENDPOINTS = {
    "admin_tasks",
    "admin_new_task",
    "admin_consistency",
    "admin_language_consistency",
    "admin_images",
    "admin_videos",
    "admin_models",
}
INVALID_FILENAME_CHARS = re.compile(r'[\x00-\x1f\x7f/\\<>:"|?*]+')
SUBMISSION_TOKEN_RE = re.compile(r"[0-9a-f]{32}")
REPORT_ITEM_TYPES = {
    "issue": "问题",
    "suggestion": "建议",
    "non_issue": "非问题",
}
REPORT_ITEM_TYPE_ORDER = ("issue", "suggestion", "non_issue")
REPORT_COUNT_KEYS = REPORT_ITEM_TYPE_ORDER + (
    "accepted_issue",
    "rejected_issue",
    "pending_issue_acceptance",
    "suppressed",
)
REPORT_ACCEPTANCE_STATUSES = {
    "pending": "未确认",
    "accepted": "接纳",
    "rejected": "不接纳",
}
REPORT_REJECTION_REASONS = {
    "model_hallucination": "模型幻觉",
    "false_positive": "模型误报",
    "evidence_insufficient": "证据不足",
    "not_applicable": "不适用",
    "other": "其他",
}
REPORT_SUPPRESSION_REJECTION_REASONS = {"model_hallucination", "false_positive", "not_applicable"}
REPORT_SUPPRESSION_DESCRIPTION_SIMILARITY_THRESHOLD = 0.56
REPORT_SUPPRESSION_DESCRIPTION_REPLACEMENTS = (
    ("不统一", "不一致"),
    ("不相同", "不一致"),
    ("存在差异", "不一致"),
    ("矛盾", "冲突"),
    ("有误", "错误"),
    ("不正确", "错误"),
    ("没有提供", "缺失"),
    ("未提供", "缺失"),
    ("没有说明", "缺失"),
    ("未说明", "缺失"),
    ("缺少", "缺失"),
    ("遗留", "保留"),
    ("残留", "保留"),
    ("资料", "文档"),
)
REPORT_ITEM_FIELDS = (
    ("severity_label", "严重程度"),
    ("confidence_label", "证据可信度"),
    ("category", "问题类型"),
    ("location", "位置"),
    ("excerpt", "原文/证据"),
    ("description", "问题描述"),
    ("impact", "影响"),
    ("suggestion", "修改建议"),
)
REPORT_SUPPRESSION_FIELDS = ("category", "location", "excerpt", "description", "impact", "suggestion")
REPORT_SEVERITY_LABELS = {
    "critical": "致命",
    "high": "高",
    "medium": "中",
    "low": "低",
}
REPORT_CONFIDENCE_LABELS = {
    "high": "高",
    "medium": "中",
    "low": "低",
}
REPORT_SEVERITY_ORDER = {value: index for index, value in enumerate(REPORT_SEVERITY_LABELS)}
REPORT_CONFIDENCE_ORDER = {value: index for index, value in enumerate(REPORT_CONFIDENCE_LABELS)}
MEDIA_REPORT_ITEM_FIELDS = (("media_summary", "AI检查结论"),)
MEDIA_REPORT_ITEM_DETAIL_FIELDS = (
    ("category", "问题类型"),
    ("location", "位置/画面"),
    ("excerpt", "依据/证据"),
    ("impact", "影响"),
    ("suggestion", "建议"),
)
REPORT_EXPORT_MIMETYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
REPORT_EXPORT_HEADER_FILL = PatternFill("solid", fgColor="EAF0F8")
REPORT_EXPORT_HEADER_FONT = Font(bold=True)
REPORT_TOTAL_EXPORT_ROWS = (
    ("问题", "issue"),
    ("建议", "suggestion"),
    ("非问题", "non_issue"),
    ("接纳问题", "accepted_issue"),
    ("不接纳问题", "rejected_issue"),
    ("待确认问题", "pending_issue_acceptance"),
    ("已忽略误报", "suppressed"),
    ("问题检出率", "issue_detection_rate"),
    ("问题接纳率", "issue_acceptance_rate"),
    ("合计", "total"),
)
REPORT_ITEM_TYPE_LABEL = "条目判定"
REPORT_ITEM_START_RE = re.compile(
    r"^(?:(?:问题|建议|风险|疑点|不一致|偏差|错误|缺失)\s*\d*[:：]|"
    r"(?:\d{1,3}[.、)]|\(\d{1,3}\)|（\d{1,3}）)\s*(?:[*_`~]{1,3}\s*)?\S)"
)
REPORT_ITEM_PREFIX_RE = re.compile(
    r"^\s*(?:(?:[-+*]\s+)|(?:#{1,6}\s+)|(?:[*_`~]{1,3}\s*))*"
)
REPORT_JSON_ITEM_KEYS = ("items", "issues", "report_items", "findings", "problems")
REPORT_JSON_SUMMARY_KEYS = ("summary", "overall", "conclusion", "总体结论", "总结")
REPORT_STATUS_KEYS = ("status", "状态", "classification", "item_type", "结论类型", "问题状态", "type")
REPORT_FIELD_ALIASES = {
    "severity": ("severity", "priority", "risk_level", "严重程度", "风险等级", "优先级"),
    "confidence": ("confidence", "certainty", "evidence_confidence", "可信度", "置信度", "证据可信度"),
    "category": ("category", "issue_type", "problem_type", "type", "问题类型", "类型", "检查类型"),
    "location": ("location", "position", "where", "位置", "位置线索", "页码", "章节", "图片位置"),
    "excerpt": (
        "excerpt",
        "quote",
        "original",
        "evidence",
        "document_a_evidence",
        "document_b_evidence",
        "原文摘录",
        "原文",
        "证据",
        "文档A证据",
        "文档B证据",
        "文档线索",
        "图片可见证据",
        "可见线索",
    ),
    "description": ("description", "issue", "problem", "finding", "疑似问题", "问题描述", "偏差说明", "差异说明", "冲突或缺失说明", "问题判断"),
    "impact": ("impact", "risk", "影响", "影响说明", "客户影响", "可能影响"),
    "suggestion": ("suggestion", "recommendation", "fix", "修改建议", "建议修改", "建议处理方式", "需核对的依据"),
}
REPORT_NO_ACTION_IMPACT_MARKERS = (
    "无实质影响",
    "无实际影响",
    "没有实质影响",
    "没有实际影响",
    "无明显影响",
    "不造成实质影响",
    "影响不大",
    "影响较小",
    "影响很小",
    "无影响",
    "不影响理解",
    "不影响使用",
    "nosubstantiveimpact",
    "nomaterialimpact",
    "nosignificantimpact",
    "norealimpact",
    "noactualimpact",
    "doesnotaffect",
)
REPORT_NO_ACTION_SUGGESTION_MARKERS = (
    "无需修改",
    "无须修改",
    "不需修改",
    "不需要修改",
    "无需处理",
    "无须处理",
    "不需处理",
    "不需要处理",
    "无需调整",
    "无须调整",
    "不需调整",
    "无需修正",
    "保持不变",
    "nomodificationrequired",
    "nomodificationneeded",
    "noneedtomodify",
    "noneedtochange",
    "nochangeneeded",
    "noactionrequired",
    "noactionneeded",
)
REPORT_LEGACY_LABEL_FIELDS = {
    "问题类型": "category",
    "类型": "category",
    "对象类型": "category",
    "位置": "location",
    "位置线索": "location",
    "图片名称或位置": "location",
    "图片位置": "location",
    "文档线索": "location",
    "原文": "excerpt",
    "原文摘录": "excerpt",
    "文档A证据": "excerpt",
    "文档B证据": "excerpt",
    "资料表述": "excerpt",
    "冲突表述": "excerpt",
    "图片可见内容": "excerpt",
    "图片可见证据": "excerpt",
    "可见内容线索": "excerpt",
    "可见线索": "excerpt",
    "识别到的文字": "excerpt",
    "问题": "description",
    "问题描述": "description",
    "疑似问题": "description",
    "偏差说明": "description",
    "差异说明": "description",
    "冲突或缺失说明": "description",
    "问题判断": "description",
    "不匹配原因": "description",
    "理由": "description",
    "影响": "impact",
    "影响说明": "impact",
    "客户影响": "impact",
    "可能影响": "impact",
    "建议": "suggestion",
    "修改建议": "suggestion",
    "建议修改": "suggestion",
    "建议处理方式": "suggestion",
    "建议补充的标题形式": "suggestion",
}
LANGUAGE_STATIC_TOKEN_RE = re.compile(
    r"https?://[^\s<>\]\)\"']+"
    r"|[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}"
    r"|\b\d{1,3}(?:\.\d{1,3}){3}\b"
    r"|\bv?\d+(?:\.\d+){1,4}\b"
    r"|\b\d{4}[-/年]\d{1,2}(?:[-/月]\d{1,2}日?)?\b"
    r"|\b\d+(?:[.,]\d+)*(?:\s?(?:%|ms|s|m|mm|cm|km|kg|g|KB|MB|GB|TB|V|A|W|Hz|kHz|MHz|GHz|°C|℃))?\b",
    re.IGNORECASE,
)
LANGUAGE_HEADING_RE = re.compile(
    r"^\s*(?:#{1,6}\s+|第[一二三四五六七八九十百千万\d]+[章节篇部]\s*|"
    r"(?:\d+|[A-Z])(?:[.\-、)]\d*){0,5}[.\-、)]?\s+|[一二三四五六七八九十]+[、.]\s*)\S"
)


def register_routes(app):
    app.add_template_global(STATUS_LABELS, "STATUS_LABELS")
    app.add_template_global(REPORT_ITEM_FIELDS, "REPORT_ITEM_FIELDS")
    app.add_template_global(REPORT_ITEM_TYPE_LABEL, "REPORT_ITEM_TYPE_LABEL")
    app.add_template_global(REPORT_ACCEPTANCE_STATUSES, "REPORT_ACCEPTANCE_STATUSES")
    app.add_template_global(REPORT_REJECTION_REASONS, "REPORT_REJECTION_REASONS")
    app.add_template_global(lambda: app.config["ADMIN_URL"], "admin_url")
    app.add_template_global(subject_label, "subject_label")
    app.add_template_global(_owner_display, "owner_display")
    app.add_template_global(_owner_meta, "owner_meta")
    app.add_template_global(_consistency_task_title, "consistency_task_title")
    app.add_template_global(_current_relative_url, "current_relative_url")
    app.add_template_global(_task_source_files_available, "task_source_files_available")

    @app.context_processor
    def inject_globals():
        identity = current_identity()
        auth_config = current_app.config.get("AUTH", {})
        return {
            "platform_mode": _platform_enabled(),
            "auth_mode": auth_config.get("mode", "ip"),
            "status_labels": STATUS_LABELS,
            "nav_identity": _identity_label(identity),
            "task_type_label": task_type_label,
            "max_upload_mb": _max_upload_mb(),
        }

    @app.errorhandler(RequestEntityTooLarge)
    def request_entity_too_large(error):
        del error
        limit = _max_upload_mb()
        flash(
            f"上传文件过大，当前上传上限为 {limit}MB。请压缩视频，或在本地 config.yaml 中调整 server.max_upload_mb 后重启服务。",
            "error",
        )
        return redirect(_request_entity_too_large_redirect()), 303

    @app.before_request
    def require_saml_user_session():
        if not _platform_enabled() or not _saml_mode_enabled() or not _needs_saml_user_session(request.endpoint):
            return None
        if _has_saml_user_session():
            return None
        return redirect(url_for("saml_login", next=_current_relative_url()))

    @app.get("/auth/saml/login")
    def saml_login():
        if not _saml_mode_enabled():
            abort(404)
        try:
            auth = create_saml_auth()
            redirect_url = auth.login(return_to=_safe_next_path(request.args.get("next")))
        except SamlConfigError as error:
            abort(503, description=str(error))
        except Exception:
            current_app.logger.exception("生成 SAML 登录请求失败")
            abort(503, description="SAML 登录配置无效，请联系管理员。")
        session["saml_request_id"] = auth.get_last_request_id()
        return redirect(redirect_url)

    @app.post("/auth/saml/acs")
    def saml_acs():
        if not _saml_mode_enabled():
            abort(404)
        try:
            auth = create_saml_auth()
            request_id = session.pop("saml_request_id", None)
            auth.process_response(request_id=request_id)
        except SamlConfigError as error:
            abort(503, description=str(error))
        except Exception:
            current_app.logger.exception("处理 SAML 回调失败")
            abort(401, description="SAML 登录失败，请重新从公司统一入口访问。")

        if auth.get_errors() or not auth.is_authenticated():
            current_app.logger.warning("SAML 回调校验失败：%s", ", ".join(auth.get_errors()))
            abort(401, description="SAML 登录失败，请重新从公司统一入口访问。")

        user_id, username = _saml_user_from_response(auth)
        if not user_id:
            abort(401, description="SAML 响应缺少用户 ID，请联系管理员检查 SSO 属性映射。")
        session[SAML_USER_SESSION_KEY] = {"user_id": user_id, "username": username or user_id}
        return redirect(_safe_next_path(request.form.get("RelayState")))

    @app.get("/auth/saml/metadata")
    def saml_metadata():
        if not _saml_mode_enabled():
            abort(404)
        try:
            metadata = saml_sp_metadata()
        except SamlConfigError as error:
            abort(503, description=str(error))
        except Exception:
            current_app.logger.exception("生成 SAML metadata 失败")
            abort(503, description="SAML SP metadata 配置无效，请联系管理员。")
        return Response(metadata, mimetype="application/samlmetadata+xml")

    @app.post("/auth/saml/logout")
    def saml_logout():
        if not _saml_mode_enabled():
            abort(404)
        session.pop(SAML_USER_SESSION_KEY, None)
        session.pop("saml_request_id", None)
        return redirect(url_for("user_tasks"))

    @app.route("/", methods=["GET", "POST"])
    def user_tasks():
        if not _platform_enabled():
            if request.method == "POST":
                return create_task_for_identity(current_identity(), admin_created=True)
            return _render_admin_tasks_page()

        identity = _current_user_identity()
        if request.method == "POST":
            return create_task_for_identity(identity, admin_created=False)
        page = _page_arg()
        total = get_db().execute(
            "SELECT COUNT(*) AS total FROM tasks WHERE COALESCE(owner_subject, 'ip:' || ip) = ? AND task_type = ?",
            (identity.subject, DOCUMENT_TASK_TYPE),
        ).fetchone()["total"]
        page = _bounded_page(page, total, TASKS_PER_PAGE)
        rows = get_db().execute(
            """
            SELECT t.*,
                   COALESCE(NULLIF(t.owner_name_snapshot, ''), NULLIF(t.username_snapshot, ''), '') AS current_owner_name,
                   COALESCE(NULLIF(t.owner_name_snapshot, ''), NULLIF(t.username_snapshot, ''), '') AS current_username,
                   COALESCE(t.owner_subject, 'ip:' || t.ip) AS effective_owner_subject
            FROM tasks t
            WHERE COALESCE(t.owner_subject, 'ip:' || t.ip) = ? AND t.task_type = ?
            ORDER BY created_at DESC, id DESC
            LIMIT ? OFFSET ?
            """,
            (identity.subject, DOCUMENT_TASK_TYPE, TASKS_PER_PAGE, (page - 1) * TASKS_PER_PAGE),
        ).fetchall()
        stats = _task_stats_for_where("COALESCE(owner_subject, 'ip:' || ip) = ? AND task_type = ?", (identity.subject, DOCUMENT_TASK_TYPE))
        return render_template(
            "user_tasks.html",
            ip=identity.ip,
            identity=identity,
            tasks=rows,
            stats=stats,
            pagination=_pagination(page, total, TASKS_PER_PAGE),
            check_items=get_enabled_check_items(),
            models=get_enabled_models(identity.subject),
            active_nav=DOCUMENT_TASK_TYPE,
        )

    @app.route("/tasks/new", methods=["GET", "POST"])
    def user_new_task():
        if not _platform_enabled():
            if request.method == "POST":
                return create_task_for_identity(current_identity(), admin_created=True)
            return redirect(url_for("user_tasks"))

        identity = _current_user_identity()
        if request.method == "POST":
            return create_task_for_identity(identity, admin_created=False)
        return redirect(url_for("user_tasks"))

    @app.route("/consistency", methods=["GET", "POST"])
    def user_consistency():
        if not _platform_enabled():
            if request.method == "POST":
                return create_consistency_task_for_identity(current_identity(), admin_created=True)
            return _render_admin_consistency_page()

        identity = _current_user_identity()
        if request.method == "POST":
            return create_consistency_task_for_identity(identity, admin_created=False)

        page = _page_arg()
        total = get_db().execute(
            "SELECT COUNT(*) AS total FROM tasks WHERE COALESCE(owner_subject, 'ip:' || ip) = ? AND task_type = ?",
            (identity.subject, CONSISTENCY_TASK_TYPE),
        ).fetchone()["total"]
        page = _bounded_page(page, total, TASKS_PER_PAGE)
        rows = get_db().execute(
            """
            SELECT t.*,
                   COALESCE(NULLIF(t.owner_name_snapshot, ''), NULLIF(t.username_snapshot, ''), '') AS current_owner_name,
                   COALESCE(NULLIF(t.owner_name_snapshot, ''), NULLIF(t.username_snapshot, ''), '') AS current_username,
                   COALESCE(t.owner_subject, 'ip:' || t.ip) AS effective_owner_subject
            FROM tasks t
            WHERE COALESCE(t.owner_subject, 'ip:' || t.ip) = ? AND t.task_type = ?
            ORDER BY created_at DESC, id DESC
            LIMIT ? OFFSET ?
            """,
            (identity.subject, CONSISTENCY_TASK_TYPE, TASKS_PER_PAGE, (page - 1) * TASKS_PER_PAGE),
        ).fetchall()
        stats = _task_stats_for_where("COALESCE(owner_subject, 'ip:' || ip) = ? AND task_type = ?", (identity.subject, CONSISTENCY_TASK_TYPE))
        return render_template(
            "user_consistency.html",
            ip=identity.ip,
            identity=identity,
            tasks=rows,
            stats=stats,
            pagination=_pagination(page, total, TASKS_PER_PAGE),
            check_items=get_enabled_check_items(CONSISTENCY_TASK_TYPE),
            models=get_enabled_models(identity.subject),
            active_nav=CONSISTENCY_TASK_TYPE,
        )

    @app.route("/language-consistency", methods=["GET", "POST"])
    def user_language_consistency():
        if not _platform_enabled():
            if request.method == "POST":
                return create_language_consistency_task_for_identity(current_identity(), admin_created=True)
            return _render_admin_language_consistency_page()

        identity = _current_user_identity()
        if request.method == "POST":
            return create_language_consistency_task_for_identity(identity, admin_created=False)

        page = _page_arg()
        total = get_db().execute(
            "SELECT COUNT(*) AS total FROM tasks WHERE COALESCE(owner_subject, 'ip:' || ip) = ? AND task_type = ?",
            (identity.subject, LANGUAGE_CONSISTENCY_TASK_TYPE),
        ).fetchone()["total"]
        page = _bounded_page(page, total, TASKS_PER_PAGE)
        rows = get_db().execute(
            """
            SELECT t.*,
                   COALESCE(NULLIF(t.owner_name_snapshot, ''), NULLIF(t.username_snapshot, ''), '') AS current_owner_name,
                   COALESCE(NULLIF(t.owner_name_snapshot, ''), NULLIF(t.username_snapshot, ''), '') AS current_username,
                   COALESCE(t.owner_subject, 'ip:' || t.ip) AS effective_owner_subject
            FROM tasks t
            WHERE COALESCE(t.owner_subject, 'ip:' || t.ip) = ? AND t.task_type = ?
            ORDER BY created_at DESC, id DESC
            LIMIT ? OFFSET ?
            """,
            (identity.subject, LANGUAGE_CONSISTENCY_TASK_TYPE, TASKS_PER_PAGE, (page - 1) * TASKS_PER_PAGE),
        ).fetchall()
        stats = _task_stats_for_where(
            "COALESCE(owner_subject, 'ip:' || ip) = ? AND task_type = ?",
            (identity.subject, LANGUAGE_CONSISTENCY_TASK_TYPE),
        )
        return render_template(
            "user_language_consistency.html",
            ip=identity.ip,
            identity=identity,
            tasks=rows,
            stats=stats,
            pagination=_pagination(page, total, TASKS_PER_PAGE),
            check_items=get_enabled_check_items(LANGUAGE_CONSISTENCY_TASK_TYPE),
            models=get_enabled_models(identity.subject),
            submission_token=uuid.uuid4().hex,
            active_nav=LANGUAGE_CONSISTENCY_TASK_TYPE,
        )

    @app.route("/images", methods=["GET", "POST"])
    def user_images():
        if not _platform_enabled():
            if request.method == "POST":
                return create_image_task_for_identity(current_identity(), admin_created=True)
            return _render_admin_images_page()

        identity = _current_user_identity()
        if request.method == "POST":
            return create_image_task_for_identity(identity, admin_created=False)

        page = _page_arg()
        total = get_db().execute(
            "SELECT COUNT(*) AS total FROM tasks WHERE COALESCE(owner_subject, 'ip:' || ip) = ? AND task_type = ?",
            (identity.subject, IMAGE_TASK_TYPE),
        ).fetchone()["total"]
        page = _bounded_page(page, total, TASKS_PER_PAGE)
        rows = get_db().execute(
            """
            SELECT t.*,
                   COALESCE(NULLIF(t.owner_name_snapshot, ''), NULLIF(t.username_snapshot, ''), '') AS current_owner_name,
                   COALESCE(NULLIF(t.owner_name_snapshot, ''), NULLIF(t.username_snapshot, ''), '') AS current_username,
                   COALESCE(t.owner_subject, 'ip:' || t.ip) AS effective_owner_subject
            FROM tasks t
            WHERE COALESCE(t.owner_subject, 'ip:' || t.ip) = ? AND t.task_type = ?
            ORDER BY created_at DESC, id DESC
            LIMIT ? OFFSET ?
            """,
            (identity.subject, IMAGE_TASK_TYPE, TASKS_PER_PAGE, (page - 1) * TASKS_PER_PAGE),
        ).fetchall()
        stats = _task_stats_for_where("COALESCE(owner_subject, 'ip:' || ip) = ? AND task_type = ?", (identity.subject, IMAGE_TASK_TYPE))
        return render_template(
            "user_images.html",
            ip=identity.ip,
            identity=identity,
            tasks=rows,
            stats=stats,
            pagination=_pagination(page, total, TASKS_PER_PAGE),
            check_items=get_enabled_check_items(IMAGE_TASK_TYPE),
            models=get_enabled_models(identity.subject),
            active_nav=IMAGE_TASK_TYPE,
        )

    @app.route("/videos", methods=["GET", "POST"])
    def user_videos():
        if not _platform_enabled():
            if request.method == "POST":
                return create_video_task_for_identity(current_identity(), admin_created=True)
            return _render_admin_videos_page()

        identity = _current_user_identity()
        if request.method == "POST":
            return create_video_task_for_identity(identity, admin_created=False)

        page = _page_arg()
        total = get_db().execute(
            "SELECT COUNT(*) AS total FROM tasks WHERE COALESCE(owner_subject, 'ip:' || ip) = ? AND task_type = ?",
            (identity.subject, VIDEO_TASK_TYPE),
        ).fetchone()["total"]
        page = _bounded_page(page, total, TASKS_PER_PAGE)
        rows = get_db().execute(
            """
            SELECT t.*,
                   COALESCE(NULLIF(t.owner_name_snapshot, ''), NULLIF(t.username_snapshot, ''), '') AS current_owner_name,
                   COALESCE(NULLIF(t.owner_name_snapshot, ''), NULLIF(t.username_snapshot, ''), '') AS current_username,
                   COALESCE(t.owner_subject, 'ip:' || t.ip) AS effective_owner_subject
            FROM tasks t
            WHERE COALESCE(t.owner_subject, 'ip:' || t.ip) = ? AND t.task_type = ?
            ORDER BY created_at DESC, id DESC
            LIMIT ? OFFSET ?
            """,
            (identity.subject, VIDEO_TASK_TYPE, TASKS_PER_PAGE, (page - 1) * TASKS_PER_PAGE),
        ).fetchall()
        stats = _task_stats_for_where("COALESCE(owner_subject, 'ip:' || ip) = ? AND task_type = ?", (identity.subject, VIDEO_TASK_TYPE))
        return render_template(
            "user_videos.html",
            ip=identity.ip,
            identity=identity,
            tasks=rows,
            stats=stats,
            pagination=_pagination(page, total, TASKS_PER_PAGE),
            check_items=get_enabled_check_items(VIDEO_TASK_TYPE),
            models=get_enabled_models(identity.subject),
            active_nav=VIDEO_TASK_TYPE,
        )

    @app.get("/tasks/<int:task_id>")
    def user_task_detail(task_id):
        task = _get_user_task_or_local_admin(task_id)
        results = _task_results(task)
        back_endpoint = _task_list_endpoint(not _platform_enabled(), task["task_type"])
        return render_template(
            "task_detail.html",
            mode="admin" if not _platform_enabled() else "user",
            task=task,
            results=results,
            report_totals=_report_item_totals(results),
            report_item_types=REPORT_ITEM_TYPES,
            report_item_fields=_report_item_fields_for_task(task["task_type"]),
            media_report=_is_media_report_task_type(task["task_type"]),
            report_classification_url=url_for("admin_update_report_item_type" if not _platform_enabled() else "user_update_report_item_type", task_id=task_id),
            document_groups=_task_document_groups(task),
            active_nav=task["task_type"] or DOCUMENT_TASK_TYPE,
            back_url=_safe_next_path(request.args.get("next"), url_for(back_endpoint)),
        )

    @app.post("/tasks/<int:task_id>/report-items")
    def user_update_report_item_type(task_id):
        task = _get_user_task_or_local_admin(task_id)
        return _update_report_item_type(task)

    @app.get("/tasks/<int:task_id>/export")
    def user_export_task(task_id):
        task = _get_user_task_or_local_admin(task_id)
        return _export_task_report(task)

    @app.get("/tasks/<int:task_id>/export.xlsx")
    def user_export_task_excel(task_id):
        task = _get_user_task_or_local_admin(task_id)
        return _export_task_report_excel(task)

    @app.get("/tasks/<int:task_id>/document")
    def user_download_task_document(task_id):
        task = _get_user_task_or_local_admin(task_id)
        return _download_task_document(task, "user_task_detail")

    @app.post("/tasks/<int:task_id>/cancel")
    def user_cancel_task(task_id):
        task = _get_user_task_or_local_admin(task_id)
        _cancel_task(task)
        flash("已提交取消请求。", "success")
        return redirect(_task_action_redirect("user_tasks"))

    @app.post("/tasks/<int:task_id>/delete")
    def user_delete_task(task_id):
        task = _get_user_task_or_local_admin(task_id)
        if _delete_task(task):
            flash("任务已删除。", "success")
        return redirect(url_for(_task_list_endpoint(False, task["task_type"])))

    @app.post("/tasks/bulk-delete")
    def user_bulk_delete_tasks():
        return _bulk_delete_tasks(_get_user_task_or_local_admin, admin_created=False)

    @app.route("/models", methods=["GET", "POST"])
    def user_models():
        return _model_management_response(_model_page_identity(), "user_models")

    @app.post("/models/fetch")
    def user_fetch_models():
        _model_page_identity()
        data = request.get_json(silent=True) or {}
        if not isinstance(data, dict):
            return {"error": "请求数据格式不正确。"}, 400
        provider_data = _provider_connection_data(data, "模型拉取")
        if isinstance(provider_data, str):
            return {"error": provider_data}, 400
        network = outbound_network_config()
        try:
            models = fetch_models(
                api_base=provider_data["api_base"],
                api_key=provider_data["api_key"],
                proxy_mode=network["proxy_mode"],
                proxy=network["proxy"],
                ssl_verify=network["ssl_verify"],
                request_timeout=provider_data["request_timeout"],
            )
        except ModelDiscoveryError as exc:
            return {"error": str(exc)}, 400
        return {"fetched_models": models, "fetched_count": len(models)}

    @app.post("/models/test")
    def user_test_model():
        _model_page_identity()
        data = request.get_json(silent=True) or {}
        if not isinstance(data, dict):
            return {"ok": False, "error": "请求数据格式不正确。"}, 400
        provider_data = _provider_connection_data(data, "模型测试")
        if isinstance(provider_data, str):
            return {"ok": False, "error": provider_data}, 400
        model_name = str(data.get("model_name") or "").strip()
        if not model_name:
            return {"ok": False, "error": "请先填写模型 ID。"}, 400
        network = outbound_network_config()
        try:
            message = test_model_connection(
                api_base=provider_data["api_base"],
                api_key=provider_data["api_key"],
                proxy_mode=network["proxy_mode"],
                proxy=network["proxy"],
                ssl_verify=network["ssl_verify"],
                request_timeout=min(provider_data["request_timeout"], MODEL_TEST_TIMEOUT_MAX),
                model_name=model_name,
                force_disable_thinking=_form_bool(data.get("force_disable_thinking")),
            )
        except LLMError as exc:
            return {"ok": False, "error": str(exc)}, 400
        return {"ok": True, "message": message}

    admin_prefix = app.config["ADMIN_URL"]

    @app.route(f"{admin_prefix}/login", methods=["GET", "POST"])
    def admin_login():
        if not _platform_enabled():
            return redirect(url_for("user_tasks"))
        if request.method == "POST":
            username = request.form.get("username", "")
            password = request.form.get("password", "")
            ok = hmac.compare_digest(username, current_app.config["ADMIN_USERNAME"]) and hmac.compare_digest(
                password, current_app.config["ADMIN_PASSWORD"]
            )
            if ok:
                session["admin_logged_in"] = True
                flash("管理员已登录。", "success")
                return redirect(url_for("admin_dashboard"))
            flash("账号或密码不正确。", "error")
        return render_template("admin_login.html")

    @app.post(f"{admin_prefix}/logout")
    def admin_logout():
        session.pop("admin_logged_in", None)
        if not _platform_enabled():
            return redirect(url_for("user_tasks"))
        flash("管理员已退出。", "success")
        return redirect(url_for("admin_login"))

    @app.get(admin_prefix)
    @admin_required
    def admin_dashboard():
        if not _platform_enabled():
            return redirect(url_for("user_tasks"))
        selected_range = _admin_overview_range()
        overview = _admin_overview_data(selected_range["start_at"], selected_range["end_at"])
        return render_template(
            "admin_overview.html",
            selected_range=selected_range,
            totals=overview["totals"],
            daily_rows=overview["daily_rows"],
            user_rows=overview["user_rows"],
            active_nav="overview",
        )

    @app.route(f"{admin_prefix}/tasks", methods=["GET", "POST"])
    @admin_required
    def admin_tasks():
        if request.method == "POST":
            return create_task_for_identity(_console_user_identity(), admin_created=True)
        return _render_admin_tasks_page()

    @app.route(f"{admin_prefix}/tasks/new", methods=["GET", "POST"])
    @admin_required
    def admin_new_task():
        if request.method == "POST":
            return create_task_for_identity(_console_user_identity(), admin_created=True)
        return redirect(url_for("admin_tasks"))

    @app.get(f"{admin_prefix}/task-statuses")
    @admin_required
    def admin_task_statuses():
        task_type = str(request.args.get("task_type") or DOCUMENT_TASK_TYPE)
        allowed_task_types = {
            DOCUMENT_TASK_TYPE,
            CONSISTENCY_TASK_TYPE,
            LANGUAGE_CONSISTENCY_TASK_TYPE,
            IMAGE_TASK_TYPE,
            VIDEO_TASK_TYPE,
        }
        if task_type not in allowed_task_types:
            return {"error": "任务类型无效。"}, 400
        task_ids = []
        for value in str(request.args.get("ids") or "").split(","):
            value = value.strip()
            if value.isdigit():
                task_ids.append(int(value))
        task_ids = list(dict.fromkeys(task_ids))[:100]
        mode_clause, mode_params = _mode_subject_filter("t")
        rows = []
        if task_ids:
            placeholders = ",".join("?" for _ in task_ids)
            rows = get_db().execute(
                f"""
                SELECT t.id, t.status, t.progress
                FROM tasks t
                WHERE t.task_type = ? AND {mode_clause} AND t.id IN ({placeholders})
                """,
                (task_type, *mode_params, *task_ids),
            ).fetchall()
        counts = get_db().execute(
            f"""
            SELECT
                COUNT(*) AS tasks,
                COALESCE(SUM(CASE WHEN t.status = 'queued' THEN 1 ELSE 0 END), 0) AS queued,
                COALESCE(SUM(CASE WHEN t.status = 'running' THEN 1 ELSE 0 END), 0) AS running,
                COALESCE(SUM(CASE WHEN t.status = 'completed' THEN 1 ELSE 0 END), 0) AS completed
            FROM tasks t
            WHERE t.task_type = ? AND {mode_clause}
            """,
            (task_type, *mode_params),
        ).fetchone()
        return {
            "active": bool(counts["queued"] or counts["running"]),
            "counts": {key: int(counts[key] or 0) for key in ("tasks", "queued", "running", "completed")},
            "tasks": [
                {
                    "id": row["id"],
                    "status": row["status"],
                    "status_label": STATUS_LABELS.get(row["status"], row["status"]),
                    "progress": int(row["progress"] or 0),
                }
                for row in rows
            ],
        }

    @app.route(f"{admin_prefix}/consistency", methods=["GET", "POST"])
    @admin_required
    def admin_consistency():
        if request.method == "POST":
            return create_consistency_task_for_identity(_console_user_identity(), admin_created=True)
        return _render_admin_consistency_page()

    @app.route(f"{admin_prefix}/language-consistency", methods=["GET", "POST"])
    @admin_required
    def admin_language_consistency():
        if request.method == "POST":
            return create_language_consistency_task_for_identity(_console_user_identity(), admin_created=True)
        return _render_admin_language_consistency_page()

    @app.route(f"{admin_prefix}/images", methods=["GET", "POST"])
    @admin_required
    def admin_images():
        if request.method == "POST":
            return create_image_task_for_identity(_console_user_identity(), admin_created=True)
        return _render_admin_images_page()

    @app.route(f"{admin_prefix}/videos", methods=["GET", "POST"])
    @admin_required
    def admin_videos():
        if request.method == "POST":
            return create_video_task_for_identity(_console_user_identity(), admin_created=True)
        return _render_admin_videos_page()

    @app.route(f"{admin_prefix}/models", methods=["GET", "POST"])
    @admin_required
    def admin_models():
        if not _platform_enabled():
            return redirect(url_for("user_models"))
        return _model_management_response(_console_user_identity(), "admin_models")

    @app.get(f"{admin_prefix}/tasks/<int:task_id>")
    @admin_required
    def admin_task_detail(task_id):
        task = _get_task_or_404(task_id)
        results = _task_results(task)
        back_endpoint = _task_list_endpoint(True, task["task_type"])
        return render_template(
            "task_detail.html",
            mode="admin",
            task=task,
            results=results,
            report_totals=_report_item_totals(results),
            report_item_types=REPORT_ITEM_TYPES,
            report_item_fields=_report_item_fields_for_task(task["task_type"]),
            media_report=_is_media_report_task_type(task["task_type"]),
            report_classification_url=url_for("admin_update_report_item_type", task_id=task_id),
            document_groups=_task_document_groups(task),
            active_nav=task["task_type"] or DOCUMENT_TASK_TYPE,
            back_url=_safe_next_path(request.args.get("next"), url_for(back_endpoint)),
        )

    @app.post(f"{admin_prefix}/tasks/<int:task_id>/report-items")
    @admin_required
    def admin_update_report_item_type(task_id):
        task = _get_task_or_404(task_id)
        return _update_report_item_type(task)

    @app.get(f"{admin_prefix}/tasks/<int:task_id>/export")
    @admin_required
    def admin_export_task(task_id):
        task = _get_task_or_404(task_id)
        return _export_task_report(task)

    @app.get(f"{admin_prefix}/tasks/<int:task_id>/export.xlsx")
    @admin_required
    def admin_export_task_excel(task_id):
        task = _get_task_or_404(task_id)
        return _export_task_report_excel(task)

    @app.get(f"{admin_prefix}/tasks/<int:task_id>/document")
    @admin_required
    def admin_download_task_document(task_id):
        task = _get_task_or_404(task_id)
        return _download_task_document(task, "admin_task_detail")

    @app.post(f"{admin_prefix}/tasks/<int:task_id>/cancel")
    @admin_required
    def admin_cancel_task(task_id):
        task = _get_task_or_404(task_id)
        _cancel_task(task)
        flash("已提交取消请求。", "success")
        return redirect(_task_action_redirect("admin_tasks"))

    @app.post(f"{admin_prefix}/tasks/<int:task_id>/delete")
    @admin_required
    def admin_delete_task(task_id):
        task = _get_task_or_404(task_id)
        if _delete_task(task):
            flash("任务已删除。", "success")
        return redirect(url_for(_task_list_endpoint(True, task["task_type"])))

    @app.post(f"{admin_prefix}/tasks/bulk-delete")
    @admin_required
    def admin_bulk_delete_tasks():
        return _bulk_delete_tasks(_get_task_or_404, admin_created=True)

    @app.route(f"{admin_prefix}/prompts", methods=["GET", "POST"])
    @admin_required
    def admin_prompts():
        return redirect(url_for("admin_settings"))

    @app.get(f"{admin_prefix}/settings/task-cache")
    @admin_required
    def admin_task_file_cache():
        snapshot = task_file_cache_snapshot(current_app)
        items = []
        for item in snapshot["items"]:
            row = dict(item)
            if row["task_type"] in {CONSISTENCY_TASK_TYPE, LANGUAGE_CONSISTENCY_TASK_TYPE}:
                row["title"] = _consistency_task_title(row)
            else:
                row["title"] = str(row["original_filename"] or f"任务 #{row['id']}")
            row["task_type_label"] = task_type_label(row["task_type"])
            row["report_url"] = url_for("admin_task_detail", task_id=row["id"])
            row.pop("document_meta_json", None)
            items.append(row)
        snapshot["items"] = items
        return snapshot

    @app.post(f"{admin_prefix}/settings/task-cache/cleanup")
    @admin_required
    def admin_cleanup_task_file_cache():
        data = request.get_json(silent=True)
        if not isinstance(data, dict) or not isinstance(data.get("task_ids"), list):
            return {"ok": False, "error": "请选择需要清理的任务。"}, 400
        task_ids = []
        for value in data["task_ids"]:
            try:
                task_id = int(value)
            except (TypeError, ValueError):
                continue
            if task_id > 0 and task_id not in task_ids:
                task_ids.append(task_id)
        if not task_ids:
            return {"ok": False, "error": "请选择需要清理的任务。"}, 400
        return {"ok": True, **cleanup_task_file_cache(current_app, task_ids)}

    @app.route(f"{admin_prefix}/settings", methods=["GET", "POST"])
    @admin_required
    def admin_settings():
        db = get_db()
        if request.method == "POST":
            action = request.form.get("action", "concurrency")
            if action == "concurrency":
                try:
                    global_concurrency = max(1, int(request.form.get("global_concurrency", "3")))
                    user_concurrency = max(1, int(request.form.get("user_concurrency", "1")))
                    check_item_concurrency = max(
                        1,
                        int(request.form.get("check_item_concurrency", str(CHECK_ITEM_CONCURRENCY_DEFAULT))),
                    )
                    image_page_check_max_pages = max(
                        1,
                        int(request.form.get("image_page_check_max_pages", str(DEFAULT_PDF_PAGE_IMAGE_MAX_PAGES))),
                    )
                    issue_output_limit = normalize_issue_output_limit(
                        int(request.form.get("issue_output_limit", str(ISSUE_OUTPUT_LIMIT_DEFAULT)))
                    )
                    task_file_retention_days = max(
                        0,
                        int(request.form.get("task_file_retention_days", str(TASK_FILE_RETENTION_DAYS_DEFAULT))),
                    )
                except ValueError:
                    flash("任务设置必须是整数，任务文件保留天数可为 0，其余必须为正整数。", "error")
                    return redirect(url_for("admin_settings"))
                set_setting("global_concurrency", global_concurrency)
                set_setting("user_concurrency", user_concurrency)
                set_setting("check_item_concurrency", check_item_concurrency)
                set_setting("image_page_check_max_pages", image_page_check_max_pages)
                set_setting("issue_output_limit", issue_output_limit)
                set_setting("task_file_retention_days", task_file_retention_days)
                flash("任务设置已保存。", "success")
                return redirect(url_for("admin_settings"))

            if action == "diagnostics":
                llm_stream_trace_enabled = request.form.get("llm_stream_trace_enabled") == "on"
                set_setting("llm_stream_trace_enabled", llm_stream_trace_enabled)
                if _wants_json_response():
                    return {"llm_stream_trace_enabled": llm_stream_trace_enabled}
                flash("定位日志设置已保存。", "success")
                return redirect(url_for("admin_settings"))

            if action == "network":
                proxy_mode = request.form.get("proxy_mode", "direct")
                proxy = request.form.get("proxy", "")
                if proxy_mode == "custom" and not str(proxy or "").strip():
                    flash("自定义代理模式需要填写代理地址。", "error")
                    return redirect(url_for("admin_settings"))
                network = save_network_config(
                    current_app.config["ROOT_DIR"],
                    {
                        "proxy_mode": proxy_mode,
                        "proxy": proxy,
                        "ssl_verify": request.form.get("ssl_verify") == "on",
                    },
                )
                current_app.config["NETWORK"] = network
                flash("系统出站网络配置已保存。", "success")
                return redirect(url_for("admin_settings"))

            if action == "ip_username":
                if not _ip_username_management_enabled():
                    abort(404)
                ip = request.form.get("ip", "").strip()
                username = request.form.get("username", "").strip()
                if not _valid_ip(ip):
                    if _wants_json_response():
                        return {"ok": False, "error": "请输入有效的 IP 地址。"}, 400
                    flash("请输入有效的 IP 地址。", "error")
                    return redirect(url_for("admin_settings", tab="ip_users"))
                set_ip_username(ip, username)
                if _wants_json_response():
                    return {"ok": True, "ip": ip, "username": username}
                flash("IP 用户名已保存。" if username else "IP 用户名已清除。", "success")
                return redirect(url_for("admin_settings", tab="ip_users"))

            if action == "report_suppression_rule":
                rule_id = request.form.get("rule_id", "")
                operation = request.form.get("operation", "")
                if not rule_id.isdigit():
                    flash("误报忽略规则不存在。", "error")
                    return redirect(url_for("admin_settings"))
                if operation == "enable":
                    db.execute(
                        "UPDATE report_suppression_rules SET enabled = 1, updated_at = ? WHERE id = ?",
                        (now_text(), int(rule_id)),
                    )
                    db.commit()
                    flash("误报忽略规则已启用。", "success")
                    return redirect(url_for("admin_settings"))
                if operation == "disable":
                    db.execute(
                        "UPDATE report_suppression_rules SET enabled = 0, updated_at = ? WHERE id = ?",
                        (now_text(), int(rule_id)),
                    )
                    db.commit()
                    flash("误报忽略规则已停用。", "success")
                    return redirect(url_for("admin_settings"))
                if operation == "delete":
                    db.execute("DELETE FROM report_suppression_rules WHERE id = ?", (int(rule_id),))
                    db.commit()
                    flash("误报忽略规则已删除。", "success")
                    return redirect(url_for("admin_settings"))
                flash("未知误报忽略规则操作。", "error")
                return redirect(url_for("admin_settings"))

            if action == "create_check_item":
                task_type = _check_item_task_type(request.form.get("task_type"))
                name = request.form.get("name", "").strip()
                description = request.form.get("description", "").strip()
                prompt = request.form.get("prompt", "").strip()
                enabled = 1 if request.form.get("enabled") == "on" else 0
                if not name or not prompt:
                    flash("检查项名称和提示词不能为空。", "error")
                    return redirect(url_for("admin_settings"))
                now = now_text()
                db.execute(
                    """
                    INSERT INTO check_items(task_type, code, name, description, prompt, enabled, sort_order, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        task_type,
                        f"{_check_item_code_prefix(task_type)}-{uuid.uuid4().hex}",
                        name,
                        description,
                        prompt,
                        enabled,
                        _next_check_item_sort_order(db, task_type),
                        now,
                        now,
                    ),
                )
                db.commit()
                flash("扩展检查项已创建。", "success")
                return redirect(url_for("admin_settings"))

            if action == "reorder_check_items":
                task_type = _check_item_task_type(request.form.get("task_type"))
                item_ids = [int(value) for value in request.form.getlist("item_ids") if value.isdigit()]
                if not item_ids:
                    if request.headers.get("X-Requested-With") == "fetch":
                        return Response("检查项顺序不能为空。", status=400)
                    flash("检查项顺序不能为空。", "error")
                    return redirect(url_for("admin_settings"))
                _reorder_check_items(db, item_ids, task_type)
                db.commit()
                if request.headers.get("X-Requested-With") == "fetch":
                    return Response(status=204)
                flash("检查项顺序已保存。", "success")
                return redirect(url_for("admin_settings"))

            if action == "delete_check_item":
                item_id = request.form.get("item_id")
                if not item_id or not item_id.isdigit():
                    flash("检查项不存在，无法删除。", "error")
                    return redirect(url_for("admin_settings"))
                item = db.execute("SELECT code FROM check_items WHERE id = ?", (item_id,)).fetchone()
                if item is None:
                    flash("检查项不存在，无法删除。", "error")
                    return redirect(url_for("admin_settings"))
                if item["code"] in default_check_item_codes():
                    flash("内置检查项不能删除。", "error")
                    return redirect(url_for("admin_settings"))
                db.execute("DELETE FROM check_items WHERE id = ?", (item_id,))
                db.commit()
                flash("扩展检查项已删除。", "success")
                return redirect(url_for("admin_settings"))

            if action == "prompt" and request.form.get("reset_prompt") == "1":
                item_id = request.form.get("item_id")
                if not item_id or not item_id.isdigit():
                    flash("检查项不存在，无法重置。", "error")
                    return redirect(url_for("admin_settings"))
                if not reset_default_check_item_prompt(int(item_id)):
                    flash("该检查项没有默认提示词可重置。", "error")
                    return redirect(url_for("admin_settings"))
                flash("检查项提示词已重置为默认内容。", "success")
                return redirect(url_for("admin_settings"))

            if action != "prompt":
                flash("未知设置操作。", "error")
                return redirect(url_for("admin_settings"))

            item_id = request.form.get("item_id")
            name = request.form.get("name", "").strip()
            description = request.form.get("description", "").strip()
            prompt = request.form.get("prompt", "").strip()
            enabled = 1 if request.form.get("enabled") == "on" else 0
            if not item_id or not item_id.isdigit() or not name or not prompt:
                flash("检查项名称和提示词不能为空。", "error")
                return redirect(url_for("admin_settings"))
            if db.execute("SELECT 1 FROM check_items WHERE id = ?", (item_id,)).fetchone() is None:
                flash("检查项不存在，无法保存。", "error")
                return redirect(url_for("admin_settings"))
            db.execute(
                """
                UPDATE check_items
                SET name = ?, description = ?, prompt = ?, enabled = ?, updated_at = ?
                WHERE id = ?
                """,
                (name, description, prompt, enabled, now_text(), item_id),
            )
            db.commit()
            flash("检查项提示词已保存。", "success")
            return redirect(url_for("admin_settings"))

        document_check_items = _check_items_for_task_type(db, DOCUMENT_TASK_TYPE)
        consistency_check_items = _check_items_for_task_type(db, CONSISTENCY_TASK_TYPE)
        language_consistency_check_items = _check_items_for_task_type(db, LANGUAGE_CONSISTENCY_TASK_TYPE)
        image_check_items = _check_items_for_task_type(db, IMAGE_TASK_TYPE)
        video_check_items = _check_items_for_task_type(db, VIDEO_TASK_TYPE)
        settings_tab = _settings_tab()
        return render_template(
            "admin_settings.html",
            check_item_groups=[
                {
                    "task_type": DOCUMENT_TASK_TYPE,
                    "tab_title": "单文档检查",
                    "title": "单文档检查提示词",
                    "description": "内置检查项不可删除；扩展检查项可新增、停用或删除。",
                    "new_title": "新增单文档检查项",
                    "name_placeholder": "例如：术语一致性检查",
                    "description_placeholder": "用于向用户说明该检查项的范围",
                    "prompt_placeholder": "描述该检查项的审查角色、关注范围和输出要求",
                    "items": document_check_items,
                    "default_check_codes": default_check_item_codes(DOCUMENT_TASK_TYPE),
                },
                {
                    "task_type": CONSISTENCY_TASK_TYPE,
                    "tab_title": "多文档对照",
                    "title": "多文档对照提示词",
                    "description": "内置检查项不可删除；扩展检查项可新增、停用或删除，提交多文档对照任务时可多选。",
                    "new_title": "新增多文档对照项",
                    "name_placeholder": "例如：关键参数一致性检查",
                    "description_placeholder": "用于说明该多文档对照项的比对范围",
                    "prompt_placeholder": "描述素材与资料的比对规则、关注范围和输出要求",
                    "items": consistency_check_items,
                    "default_check_codes": default_check_item_codes(CONSISTENCY_TASK_TYPE),
                },
                {
                    "task_type": LANGUAGE_CONSISTENCY_TASK_TYPE,
                    "tab_title": "跨语种检查",
                    "title": "跨语种检查提示词",
                    "description": "内置检查项不可删除；扩展检查项可新增、停用或删除，提交跨语种检查任务时可多选。",
                    "new_title": "新增跨语种检查项",
                    "name_placeholder": "例如：翻译缺失与事实差异检查",
                    "description_placeholder": "用于说明该跨语种检查项的范围",
                    "prompt_placeholder": "描述两份不同语种文档的比对规则、关注范围和中文输出要求",
                    "items": language_consistency_check_items,
                    "default_check_codes": default_check_item_codes(LANGUAGE_CONSISTENCY_TASK_TYPE),
                },
                {
                    "task_type": IMAGE_TASK_TYPE,
                    "tab_title": "图片检查",
                    "title": "图片检查提示词",
                    "description": "内置检查项不可删除；扩展检查项可新增、停用或删除，提交图片检查任务时可多选。",
                    "new_title": "新增图片检查项",
                    "name_placeholder": "例如：端子标识完整性检查",
                    "description_placeholder": "用于说明该图片检查项的范围",
                    "prompt_placeholder": "描述图片审查角色、关注范围、判断规则和输出要求",
                    "items": image_check_items,
                    "default_check_codes": default_check_item_codes(IMAGE_TASK_TYPE),
                },
                {
                    "task_type": VIDEO_TASK_TYPE,
                    "tab_title": "视频检查",
                    "title": "视频检查提示词",
                    "description": "内置检查项不可删除；扩展检查项可新增、停用或删除，提交视频检查任务时可多选。",
                    "new_title": "新增视频检查项",
                    "name_placeholder": "例如：安装力矩与工具使用检查",
                    "description_placeholder": "用于说明该视频检查项的范围",
                    "prompt_placeholder": "描述视频质检角色、关注范围、判断规则和输出要求",
                    "items": video_check_items,
                    "default_check_codes": default_check_item_codes(VIDEO_TASK_TYPE),
                },
            ],
            global_concurrency=get_setting("global_concurrency", 3),
            user_concurrency=get_setting("user_concurrency", 1),
            check_item_concurrency=get_setting("check_item_concurrency", CHECK_ITEM_CONCURRENCY_DEFAULT),
            image_page_check_max_pages=get_setting("image_page_check_max_pages", DEFAULT_PDF_PAGE_IMAGE_MAX_PAGES),
            issue_output_limit=get_setting("issue_output_limit", ISSUE_OUTPUT_LIMIT_DEFAULT),
            max_issue_output_limit=MAX_ISSUE_OUTPUT_LIMIT,
            task_file_retention_days=get_setting("task_file_retention_days", TASK_FILE_RETENTION_DAYS_DEFAULT),
            network=current_app.config["NETWORK"],
            llm_stream_trace_enabled=get_bool_setting("llm_stream_trace_enabled", False),
            settings_tab=settings_tab,
            ip_username_management_enabled=_ip_username_management_enabled(),
            ip_username_rows=_ip_username_rows() if _ip_username_management_enabled() else [],
            report_suppression_rules=_report_suppression_rule_rows(),
        )


def _identity_label(identity: UserIdentity) -> str:
    if identity.display_name:
        return f"{identity.subject}-{identity.display_name}"
    return identity.label


def _wants_json_response() -> bool:
    return request.headers.get("X-Requested-With") == "fetch" or request.accept_mimetypes.best == "application/json"


def _report_suppression_rule_rows() -> list[dict]:
    rows = get_db().execute(
        """
        SELECT r.*, c.name AS check_name
        FROM report_suppression_rules r
        LEFT JOIN check_items c ON c.code = r.check_code
        ORDER BY r.enabled ASC, r.updated_at DESC, r.id DESC
        """
    ).fetchall()
    result = []
    for row in rows:
        snapshot = _parse_report_suppression_item_json(row["item_json"])
        result.append(
            {
                "id": row["id"],
                "enabled": bool(row["enabled"]),
                "task_type": row["task_type"],
                "task_type_label": task_type_label(row["task_type"]),
                "check_code": row["check_code"],
                "check_name": row["check_name"] or row["check_code"],
                "reason": row["reason"] or "",
                "hit_count": row["hit_count"],
                "last_hit_at": row["last_hit_at"] or "",
                "source_task_id": row["source_task_id"],
                "created_at": row["created_at"],
                "updated_at": row["updated_at"],
                "item": snapshot,
            }
        )
    return result


def _parse_report_suppression_item_json(raw: str | None) -> dict:
    if not raw:
        return {}
    try:
        value = json.loads(raw)
    except json.JSONDecodeError:
        return {}
    return value if isinstance(value, dict) else {}


def _form_bool(value) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "on"}


def _platform_enabled() -> bool:
    return bool(current_app.config.get("PLATFORM", True))


def _max_upload_mb() -> int:
    try:
        return max(1, int(current_app.config.get("MAX_UPLOAD_MB") or 1))
    except (TypeError, ValueError):
        return 1


def _request_entity_too_large_redirect() -> str:
    upload_endpoints = {
        "user_tasks",
        "user_new_task",
        "user_consistency",
        "user_language_consistency",
        "user_images",
        "user_videos",
        "admin_tasks",
        "admin_consistency",
        "admin_language_consistency",
        "admin_images",
        "admin_videos",
    }
    if request.endpoint in upload_endpoints:
        return url_for(request.endpoint)
    referrer = _same_origin_referrer_path()
    if referrer:
        return referrer
    return url_for("user_tasks")


def _same_origin_referrer_path() -> str:
    referrer = str(request.referrer or "").strip()
    if not referrer:
        return ""
    parsed = urlsplit(referrer)
    if parsed.netloc and parsed.netloc != request.host:
        return ""
    path = parsed.path or "/"
    if parsed.query:
        path = f"{path}?{parsed.query}"
    return _safe_next_path(path)


def _auth_mode() -> str:
    auth_config = current_app.config.get("AUTH", {})
    if not isinstance(auth_config, dict):
        return "ip"
    return str(auth_config.get("mode") or "ip").strip().lower()


def _mode_subject_prefix() -> str:
    mode = _auth_mode()
    if mode == "trusted_header":
        return "trusted_header:"
    if mode == "saml":
        return "saml:"
    return "ip:"


def _owner_subject_expr(table_alias: str = "t") -> str:
    prefix = f"{table_alias}." if table_alias else ""
    return f"COALESCE({prefix}owner_subject, 'ip:' || {prefix}ip)"


def _mode_subject_filter(table_alias: str = "t") -> tuple[str, tuple[str]]:
    return f"instr({_owner_subject_expr(table_alias)}, ?) = 1", (_mode_subject_prefix(),)


def _ip_username_management_enabled() -> bool:
    return _auth_mode() == "ip"


def _settings_tab() -> str:
    tab = request.args.get("tab", "general").strip()
    if tab == "ip_users" and _ip_username_management_enabled():
        return tab
    return "general"


def _valid_ip(value: str) -> bool:
    try:
        ip_address(str(value or "").strip())
    except ValueError:
        return False
    return True


def _ip_username_rows():
    return get_db().execute(
        """
        WITH known_ips AS (
            SELECT ip
            FROM tasks
            WHERE ip IS NOT NULL
              AND ip != ''
              AND instr(COALESCE(owner_subject, 'ip:' || ip), 'ip:') = 1
            UNION
            SELECT ip FROM ip_usernames
        )
        SELECT
            k.ip,
            COALESCE(u.username, '') AS username,
            COUNT(t.id) AS task_count,
            MAX(t.created_at) AS last_task_at
        FROM known_ips k
        LEFT JOIN ip_usernames u ON u.ip = k.ip
        LEFT JOIN tasks t
            ON t.ip = k.ip
           AND instr(COALESCE(t.owner_subject, 'ip:' || t.ip), 'ip:') = 1
        GROUP BY k.ip, u.username
        ORDER BY COALESCE(MAX(t.created_at), '') DESC, k.ip ASC
        """
    ).fetchall()


def _saml_mode_enabled() -> bool:
    return _auth_mode() == "saml"


def _is_user_endpoint(endpoint: str | None) -> bool:
    return bool(endpoint and endpoint.startswith("user_"))


def _needs_saml_user_session(endpoint: str | None) -> bool:
    if _is_user_endpoint(endpoint):
        return True
    return bool(endpoint in CONSOLE_USER_ENDPOINTS and session.get("admin_logged_in"))


def _has_saml_user_session() -> bool:
    saml_user = session.get(SAML_USER_SESSION_KEY)
    return isinstance(saml_user, dict) and bool(str(saml_user.get("user_id") or "").strip())


def _current_relative_url() -> str:
    path = request.full_path if request.query_string else request.path
    script_root = request.script_root.rstrip("/")
    return f"{script_root}{path}".rstrip("?") or url_for("user_tasks")


def _safe_next_path(value, fallback: str | None = None) -> str:
    fallback = fallback or url_for("user_tasks")
    value = str(value or "").strip()
    if not value:
        return fallback
    parsed = urlsplit(value)
    if parsed.scheme or parsed.netloc or not value.startswith("/") or value.startswith("//"):
        return fallback
    script_root = request.script_root.rstrip("/")
    if script_root and value != script_root and not value.startswith(f"{script_root}/"):
        return fallback
    return value


def _saml_user_from_response(auth) -> tuple[str, str]:
    saml_config = current_app.config.get("AUTH", {}).get("saml", {})
    user_id_attribute = str(saml_config.get("user_id_attribute") or "").strip()
    username_attribute = str(saml_config.get("username_attribute") or "").strip()
    attributes = auth.get_attributes() or {}
    friendly_attributes = getattr(auth, "get_friendlyname_attributes", lambda: {})() or {}

    if user_id_attribute:
        user_id = _saml_attribute_value(attributes, user_id_attribute) or _saml_attribute_value(
            friendly_attributes, user_id_attribute
        )
    else:
        user_id = str(auth.get_nameid() or "").strip()
    username = ""
    if username_attribute:
        username = _saml_attribute_value(attributes, username_attribute) or _saml_attribute_value(
            friendly_attributes, username_attribute
        )
    return user_id, username or user_id


def _saml_attribute_value(attributes: dict, name: str) -> str:
    value = attributes.get(name) if isinstance(attributes, dict) else None
    if isinstance(value, (list, tuple)):
        value = value[0] if value else ""
    return str(value or "").strip()


def _current_user_identity() -> UserIdentity:
    try:
        return current_identity(require_sso=True)
    except AuthenticationRequired:
        abort(401, description="未收到 SSO 用户信息，请通过公司统一入口访问。")


def _console_user_identity() -> UserIdentity:
    if _platform_enabled():
        return _current_user_identity()
    return current_identity()


def _owner_display(task) -> str:
    ip = str(_row_value(task, "ip") or "").strip()
    subject = (
        _row_value(task, "effective_owner_subject")
        or _row_value(task, "owner_subject")
        or owner_subject_from_ip(ip)
    )
    subject = str(subject)
    if subject.startswith("ip:"):
        current_ip_username = _row_value(task, "current_ip_username")
        if current_ip_username:
            return str(current_ip_username)
        if not _row_value(task, "ip_username_lookup_complete", False):
            current_ip_username = get_ip_username(ip or subject[3:])
            if current_ip_username:
                return current_ip_username
    current_owner_name = _row_value(task, "current_owner_name")
    if current_owner_name:
        return str(current_owner_name)
    owner_name_snapshot = _row_value(task, "owner_name_snapshot")
    if owner_name_snapshot:
        return str(owner_name_snapshot)
    username_snapshot = _row_value(task, "username_snapshot")
    if username_snapshot:
        return str(username_snapshot)
    return subject_label(subject)


def _owner_meta(task) -> str:
    ip = str(_row_value(task, "ip") or "").strip()
    subject = (
        _row_value(task, "effective_owner_subject")
        or _row_value(task, "owner_subject")
        or owner_subject_from_ip(ip)
    )
    subject = str(subject)
    if subject.startswith("ip:"):
        subject_ip = subject[3:].strip()
        display = _owner_display(task)
        if display and display not in {subject_ip, ip}:
            return f"IP {ip or subject_ip}"
        return ""
    if subject and ip:
        return f"{subject} · IP {ip}"
    if subject:
        return subject
    if ip:
        return f"IP {ip}"
    return ""


def _row_value(row, key: str, default=None):
    if row is None:
        return default
    if hasattr(row, "keys") and key in row.keys():
        return row[key]
    if isinstance(row, dict):
        return row.get(key, default)
    return default


def _consistency_task_title(task, include_all: bool = False) -> str:
    groups = document_groups_from_meta(_row_value(task, "document_meta_json"))
    title = _consistency_title_from_groups(groups, include_all=include_all)
    if title:
        return title
    return str(_row_value(task, "original_filename", "多文档对照检查") or "多文档对照检查")


def _consistency_title_from_groups(groups: list[dict], *, include_all: bool = False) -> str:
    parts = []
    for group in groups:
        names = [
            Path(str(file_info.get("original_filename") or "")).name.strip()
            for file_info in group.get("files", [])
        ]
        names = [name for name in names if name]
        if not names:
            continue
        label = str(group.get("label") or "文档").strip() or "文档"
        if include_all or len(names) <= 2:
            summary = "、".join(names)
        else:
            summary = f"{'、'.join(names[:2])} 等{len(names)}个"
        parts.append(f"{label}：{summary}")
    return " / ".join(parts)


def _render_admin_tasks_page():
    return _render_admin_task_list(
        task_type=DOCUMENT_TASK_TYPE,
        template_name="admin_tasks.html",
        totals_task_type=DOCUMENT_TASK_TYPE,
        check_items=get_enabled_check_items(),
    )


def _render_admin_consistency_page():
    return _render_admin_task_list(
        task_type=CONSISTENCY_TASK_TYPE,
        template_name="admin_consistency.html",
        totals_task_type=CONSISTENCY_TASK_TYPE,
        check_items=get_enabled_check_items(CONSISTENCY_TASK_TYPE),
    )


def _render_admin_language_consistency_page():
    return _render_admin_task_list(
        task_type=LANGUAGE_CONSISTENCY_TASK_TYPE,
        template_name="admin_language_consistency.html",
        totals_task_type=LANGUAGE_CONSISTENCY_TASK_TYPE,
        check_items=get_enabled_check_items(LANGUAGE_CONSISTENCY_TASK_TYPE),
    )


def _render_admin_images_page():
    return _render_admin_task_list(
        task_type=IMAGE_TASK_TYPE,
        template_name="admin_images.html",
        totals_task_type=IMAGE_TASK_TYPE,
        check_items=get_enabled_check_items(IMAGE_TASK_TYPE),
    )


def _render_admin_videos_page():
    return _render_admin_task_list(
        task_type=VIDEO_TASK_TYPE,
        template_name="admin_videos.html",
        totals_task_type=VIDEO_TASK_TYPE,
        check_items=get_enabled_check_items(VIDEO_TASK_TYPE),
    )


def _render_admin_task_list(*, task_type: str, template_name: str, totals_task_type: str, check_items):
    identity = _console_user_identity()
    status = request.args.get("status", "")
    owner = request.args.get("owner", request.args.get("ip", "")).strip()
    page = _page_arg()
    params = []
    clauses = []
    join_ip_usernames = _auth_mode() == "ip"
    ip_username_join = "LEFT JOIN ip_usernames iu ON iu.ip = t.ip" if join_ip_usernames else ""
    owner_name_expr = (
        "COALESCE(NULLIF(iu.username, ''), NULLIF(t.owner_name_snapshot, ''), NULLIF(t.username_snapshot, ''), '')"
        if join_ip_usernames
        else "COALESCE(NULLIF(t.owner_name_snapshot, ''), NULLIF(t.username_snapshot, ''), '')"
    )
    current_ip_username_expr = "COALESCE(iu.username, '')" if join_ip_usernames else "''"
    mode_clause, mode_params = _mode_subject_filter("t")
    clauses.append(mode_clause)
    params.extend(mode_params)
    if status:
        clauses.append("t.status = ?")
        params.append(status)
    if owner:
        owner_name_filter = "OR COALESCE(iu.username, '') LIKE ?" if join_ip_usernames else ""
        clauses.append(
            f"""
            (
                COALESCE(t.owner_subject, 'ip:' || t.ip) LIKE ?
                OR t.ip LIKE ?
                OR COALESCE(t.owner_name_snapshot, t.username_snapshot, '') LIKE ?
                {owner_name_filter}
            )
            """
        )
        owner_like = f"%{owner}%"
        params.extend([owner_like, owner_like, owner_like])
        if join_ip_usernames:
            params.append(owner_like)
    clauses.append("t.task_type = ?")
    params.append(task_type)
    where = f"WHERE {' AND '.join(clauses)}"
    total = get_db().execute(
        f"""
        SELECT COUNT(*) AS total
        FROM tasks t
        {ip_username_join}
        {where}
        """,
        tuple(params),
    ).fetchone()["total"]
    page = _bounded_page(page, total, TASKS_PER_PAGE)
    rows = get_db().execute(
        f"""
        SELECT
               t.id, t.task_type, t.ip, t.username_snapshot,
               t.owner_subject, t.owner_name_snapshot, t.owner_source,
               t.original_filename, t.stored_filename, t.file_type, t.file_size,
               t.provider_name, t.model_name, t.status, t.progress,
               t.created_at, t.document_meta_json,
               {current_ip_username_expr} AS current_ip_username,
               {1 if join_ip_usernames else 0} AS ip_username_lookup_complete,
               {owner_name_expr} AS current_owner_name,
               {owner_name_expr} AS current_username,
               COALESCE(t.owner_subject, 'ip:' || t.ip) AS effective_owner_subject
        FROM tasks t
        {ip_username_join}
        {where}
        ORDER BY t.created_at DESC, t.id DESC
        LIMIT ? OFFSET ?
        """,
        tuple(params + [TASKS_PER_PAGE, (page - 1) * TASKS_PER_PAGE]),
    ).fetchall()
    return render_template(
        template_name,
        tasks=rows,
        status=status,
        owner=owner,
        ip=owner,
        pagination=_pagination(page, total, TASKS_PER_PAGE),
        totals=_admin_totals(totals_task_type),
        global_concurrency=get_setting("global_concurrency", 3),
        user_concurrency=get_setting("user_concurrency", 1),
        check_items=check_items,
        models=get_enabled_models(identity.subject),
        submission_token=uuid.uuid4().hex,
        refresh_url=url_for("admin_task_statuses", task_type=task_type),
        active_nav=task_type,
    )


def _check_item_task_type(value: str | None) -> str:
    if value == CONSISTENCY_TASK_TYPE:
        return CONSISTENCY_TASK_TYPE
    if value == LANGUAGE_CONSISTENCY_TASK_TYPE:
        return LANGUAGE_CONSISTENCY_TASK_TYPE
    if value == IMAGE_TASK_TYPE:
        return IMAGE_TASK_TYPE
    if value == VIDEO_TASK_TYPE:
        return VIDEO_TASK_TYPE
    return DOCUMENT_TASK_TYPE


def _check_item_code_prefix(task_type: str) -> str:
    if task_type == CONSISTENCY_TASK_TYPE:
        return "custom-consistency"
    if task_type == LANGUAGE_CONSISTENCY_TASK_TYPE:
        return "custom-language-consistency"
    if task_type == IMAGE_TASK_TYPE:
        return "custom-image"
    if task_type == VIDEO_TASK_TYPE:
        return "custom-video"
    return "custom"


def _check_items_for_task_type(db, task_type: str):
    return db.execute(
        """
        SELECT *
        FROM check_items
        WHERE task_type = ?
        ORDER BY sort_order ASC, id ASC
        """,
        (task_type,),
    ).fetchall()


def get_enabled_check_items(task_type: str = DOCUMENT_TASK_TYPE):
    return get_db().execute(
        """
        SELECT *
        FROM check_items
        WHERE task_type = ? AND enabled = 1
        ORDER BY sort_order ASC, id ASC
        """,
        (task_type,),
    ).fetchall()


def _enabled_check_item_snapshots(db, check_ids: list[int], task_type: str) -> list[dict]:
    unique_ids = []
    seen = set()
    for check_id in check_ids:
        if check_id not in seen:
            unique_ids.append(check_id)
            seen.add(check_id)
    if not unique_ids:
        return []

    placeholders = ",".join("?" for _ in unique_ids)
    rows = db.execute(
        f"""
        SELECT id, code, name, prompt
        FROM check_items
        WHERE id IN ({placeholders}) AND task_type = ? AND enabled = 1
        ORDER BY sort_order ASC, id ASC
        """,
        tuple(unique_ids + [task_type]),
    ).fetchall()
    return [
        {
            "id": row["id"],
            "code": row["code"],
            "name": row["name"],
            "prompt": row["prompt"],
        }
        for row in rows
    ]


def _next_check_item_sort_order(db, task_type: str = DOCUMENT_TASK_TYPE) -> int:
    row = db.execute(
        "SELECT MIN(sort_order) AS value FROM check_items WHERE task_type = ?",
        (task_type,),
    ).fetchone()
    if row is None or row["value"] is None:
        return 10
    return int(row["value"]) - 10


def _reorder_check_items(db, item_ids: list[int], task_type: str = DOCUMENT_TASK_TYPE) -> list[int]:
    rows = db.execute(
        """
        SELECT id
        FROM check_items
        WHERE task_type = ?
        ORDER BY sort_order ASC, id ASC
        """,
        (task_type,),
    ).fetchall()
    existing_ids = [int(row["id"]) for row in rows]
    existing_set = set(existing_ids)
    ordered_ids = []
    seen_ids = set()
    for item_id in item_ids:
        if item_id in existing_set and item_id not in seen_ids:
            ordered_ids.append(item_id)
            seen_ids.add(item_id)
    ordered_ids.extend(item_id for item_id in existing_ids if item_id not in seen_ids)

    updated_at = now_text()
    for index, item_id in enumerate(ordered_ids, start=1):
        db.execute(
            "UPDATE check_items SET sort_order = ?, updated_at = ? WHERE id = ?",
            (index * 10, updated_at, item_id),
        )
    return ordered_ids


def _model_page_identity() -> UserIdentity:
    if _platform_enabled():
        return _current_user_identity()
    return current_identity()


def _model_management_response(identity: UserIdentity, redirect_endpoint: str):
    if request.method == "POST":
        action = request.form.get("action", "save")
        provider_id = request.form.get("provider_id")
        if action == "delete" and provider_id:
            _delete_user_model_provider(identity.subject, provider_id)
            flash("模型提供商已删除。", "success")
            return redirect(url_for(redirect_endpoint))

        provider_data = _provider_form_data()
        if isinstance(provider_data, str):
            flash(provider_data, "error")
            return redirect(url_for(redirect_endpoint))

        if provider_id and not _user_provider_exists(identity.subject, provider_id):
            flash("模型提供商不存在。", "error")
            return redirect(url_for(redirect_endpoint))

        _save_user_model_provider(identity.subject, provider_id, provider_data)
        flash("模型提供商已保存。", "success")
        return redirect(url_for(redirect_endpoint))

    providers = _load_user_model_providers(identity.subject)
    models_by_provider = {
        provider["id"]: sorted(
            _provider_model_options(provider),
            key=lambda model: (model["model_name"], model["force_disable_thinking"]),
        )
        for provider in providers
    }
    return render_template(
        "user_models.html",
        providers=providers,
        models_by_provider=models_by_provider,
        active_nav="models",
    )


def _provider_form_data() -> dict | str:
    return _normalize_provider_input(
        {
            "name": request.form.get("name", ""),
            "api_base": request.form.get("api_base", ""),
            "api_key": request.form.get("api_key", ""),
            "request_timeout": request.form.get("request_timeout", str(PROVIDER_TIMEOUT_DEFAULT)),
            "max_input_chars": request.form.get("max_input_chars", str(PROVIDER_INPUT_LIMIT_DEFAULT)),
            "is_active": request.form.get("is_active") == "on",
            "models": _parse_model_configs(
                request.form.get("model_configs", ""),
                request.form.get("models", ""),
            ),
        },
        require_models=True,
    )


def _provider_connection_data(data: dict, name: str) -> dict | str:
    return _normalize_provider_input(
        {
            "name": name,
            "api_base": data.get("api_base", ""),
            "api_key": data.get("api_key", ""),
            "request_timeout": data.get("request_timeout", str(PROVIDER_TIMEOUT_DEFAULT)),
            "max_input_chars": str(PROVIDER_INPUT_LIMIT_DEFAULT),
            "is_active": True,
            "models": [{"model_name": "placeholder", "force_disable_thinking": False}],
        },
        require_models=False,
    )


def _normalize_provider_input(value: dict, *, require_models: bool) -> dict | str:
    name = str(value.get("name") or "").strip()
    api_base = str(value.get("api_base") or "").strip().rstrip("/")
    api_key = str(value.get("api_key") or "").strip()
    if not name or not api_base:
        return "提供商名称和 API 地址不能为空。"
    if not _is_chat_completions_endpoint(api_base):
        return "API 地址必须填写完整的 /chat/completions 请求地址。"
    try:
        request_timeout = int(value.get("request_timeout") or PROVIDER_TIMEOUT_DEFAULT)
    except (TypeError, ValueError):
        return "超时时间必须是整数秒。"
    try:
        max_input_chars = int(value.get("max_input_chars") or PROVIDER_INPUT_LIMIT_DEFAULT)
    except (TypeError, ValueError):
        return "文本上限必须是整数。"
    if request_timeout < PROVIDER_TIMEOUT_MIN or request_timeout > PROVIDER_TIMEOUT_MAX:
        return f"超时时间需在 {PROVIDER_TIMEOUT_MIN}-{PROVIDER_TIMEOUT_MAX} 秒之间。"
    if max_input_chars < PROVIDER_INPUT_LIMIT_MIN or max_input_chars > PROVIDER_INPUT_LIMIT_MAX:
        return f"文本上限需在 {PROVIDER_INPUT_LIMIT_MIN}-{PROVIDER_INPUT_LIMIT_MAX} 字之间。"
    model_configs = value.get("models") or []
    if require_models and not model_configs:
        return "至少需要填写一个模型 ID。"
    return {
        "name": name,
        "api_base": api_base,
        "api_key": api_key,
        "request_timeout": request_timeout,
        "max_input_chars": max_input_chars,
        "is_active": bool(value.get("is_active")),
        "models": model_configs,
    }


def _load_user_model_providers(owner_subject: str) -> list[dict]:
    rows = get_db().execute(
        """
        SELECT *
        FROM user_model_providers
        WHERE owner_subject = ?
        ORDER BY updated_at DESC, id DESC
        """,
        (owner_subject,),
    ).fetchall()
    return [_provider_from_row(row, _load_user_model_configs(row["id"])) for row in rows]


def _load_user_model_configs(provider_id: int) -> list[dict]:
    rows = get_db().execute(
        """
        SELECT model_name, force_disable_thinking
        FROM user_model_configs
        WHERE provider_id = ?
        ORDER BY sort_order ASC, id ASC
        """,
        (provider_id,),
    ).fetchall()
    return [
        {
            "model_name": row["model_name"],
            "force_disable_thinking": bool(row["force_disable_thinking"]),
        }
        for row in rows
    ]


def _provider_from_row(row, models: list[dict]) -> dict:
    return {
        "id": row["id"],
        "owner_subject": row["owner_subject"],
        "name": row["name"],
        "api_base": row["api_base"],
        "api_key": row["api_key"] or "",
        "request_timeout": row["request_timeout"],
        "max_input_chars": row["max_input_chars"],
        "is_active": bool(row["is_active"]),
        "models": models,
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def _user_provider_exists(owner_subject: str, provider_id) -> bool:
    return (
        get_db()
        .execute(
            "SELECT 1 FROM user_model_providers WHERE id = ? AND owner_subject = ?",
            (provider_id, owner_subject),
        )
        .fetchone()
        is not None
    )


def _save_user_model_provider(owner_subject: str, provider_id, provider_data: dict):
    db = get_db()
    now = now_text()
    if provider_id:
        db.execute(
            """
            UPDATE user_model_providers
            SET name = ?, api_base = ?, api_key = ?,
                request_timeout = ?, max_input_chars = ?, is_active = ?, updated_at = ?
            WHERE id = ? AND owner_subject = ?
            """,
            (
                provider_data["name"],
                provider_data["api_base"],
                provider_data["api_key"],
                provider_data["request_timeout"],
                provider_data["max_input_chars"],
                1 if provider_data["is_active"] else 0,
                now,
                provider_id,
                owner_subject,
            ),
        )
        saved_provider_id = int(provider_id)
    else:
        cursor = db.execute(
            """
            INSERT INTO user_model_providers(
                owner_subject, name, api_base, api_key,
                request_timeout, max_input_chars, is_active, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                owner_subject,
                provider_data["name"],
                provider_data["api_base"],
                provider_data["api_key"],
                provider_data["request_timeout"],
                provider_data["max_input_chars"],
                1 if provider_data["is_active"] else 0,
                now,
                now,
            ),
        )
        saved_provider_id = cursor.lastrowid
        if saved_provider_id is None:
            raise RuntimeError("模型提供商保存失败，请稍后重试。")
    _replace_user_model_configs(saved_provider_id, provider_data["models"], now)
    db.commit()


def _replace_user_model_configs(provider_id: int, model_configs: list[dict], updated_at: str):
    db = get_db()
    db.execute("DELETE FROM user_model_configs WHERE provider_id = ?", (provider_id,))
    for index, model_config in enumerate(model_configs, start=1):
        db.execute(
            """
            INSERT INTO user_model_configs(
                provider_id, model_name, force_disable_thinking, sort_order, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                provider_id,
                model_config["model_name"],
                1 if model_config["force_disable_thinking"] else 0,
                index * 10,
                updated_at,
                updated_at,
            ),
        )


def _delete_user_model_provider(owner_subject: str, provider_id):
    get_db().execute(
        "DELETE FROM user_model_providers WHERE id = ? AND owner_subject = ?",
        (provider_id, owner_subject),
    )
    get_db().commit()


def _parse_model_configs(model_configs_json: str, models_text: str = "") -> list[dict]:
    configs = []
    try:
        value = json.loads(model_configs_json) if model_configs_json else []
    except json.JSONDecodeError:
        value = []

    if isinstance(value, list):
        for item in value:
            if isinstance(item, dict):
                model_name = str(item.get("model_name") or item.get("id") or "").strip()
                force_disable_thinking = _form_bool(item.get("force_disable_thinking", True))
            else:
                model_name = str(item or "").strip()
                force_disable_thinking = True
            configs.append(
                {
                    "model_name": model_name,
                    "force_disable_thinking": force_disable_thinking,
                }
            )

    if not configs:
        configs = [
            {
                "model_name": line.strip(),
                "force_disable_thinking": True,
            }
            for line in str(models_text or "").splitlines()
            if line.strip()
        ]

    result = []
    seen = set()
    for config in configs:
        model_name = str(config.get("model_name") or "").strip()
        force_disable_thinking = bool(config.get("force_disable_thinking"))
        key = (model_name, force_disable_thinking)
        if not model_name or key in seen:
            continue
        seen.add(key)
        result.append(
            {
                "model_name": model_name,
                "force_disable_thinking": force_disable_thinking,
            }
        )
    return result


def _provider_model_options(provider: dict) -> list[dict]:
    return [
        {
            "model_name": _model_config_name(model_config),
            "force_disable_thinking": _model_config_force_disable_thinking(model_config),
            "enabled": True,
        }
        for model_config in provider["models"]
        if _model_config_name(model_config)
    ]


def _model_config_name(model_config) -> str:
    if isinstance(model_config, dict):
        return str(model_config.get("model_name") or model_config.get("id") or "").strip()
    return str(model_config or "").strip()


def _model_config_force_disable_thinking(model_config) -> bool:
    if not isinstance(model_config, dict):
        return True
    return _form_bool(model_config.get("force_disable_thinking", True))


def get_enabled_models(owner_subject: str | None = None):
    if owner_subject is None:
        owner_subject = current_identity().subject
    models = []
    for provider in _load_user_model_providers(owner_subject):
        if not provider["is_active"]:
            continue
        for model_config in provider["models"]:
            models.append(_model_option(provider, model_config))
    return sorted(models, key=lambda model: (model["provider_name"], model["model_name"], model["force_disable_thinking"]))


def _model_option(provider: dict, model_name) -> dict:
    if isinstance(model_name, dict):
        model_config = model_name
        model_name = str(model_config.get("model_name") or model_config.get("id") or "").strip()
        force_disable_thinking = bool(model_config.get("force_disable_thinking"))
    else:
        model_name = str(model_name or "").strip()
        force_disable_thinking = False
    return {
        "id": f"{provider['id']}:{1 if force_disable_thinking else 0}:{model_name}",
        "provider_id": provider["id"],
        "provider_name": provider["name"],
        "model_name": model_name,
        "force_disable_thinking": force_disable_thinking,
        "api_base": provider["api_base"],
        "api_key": provider["api_key"],
        "request_timeout": provider["request_timeout"],
        "max_input_chars": provider["max_input_chars"],
    }


def _is_chat_completions_endpoint(value: str) -> bool:
    endpoint = str(value or "").strip().rstrip("/")
    return endpoint.startswith(("http://", "https://")) and endpoint.endswith("/chat/completions")


def _find_enabled_model(model_id: str, owner_subject: str | None = None) -> dict | None:
    if ":" not in model_id:
        return None
    if owner_subject is None:
        owner_subject = current_identity().subject
    force_disable_thinking = None
    parts = model_id.split(":", 2)
    if len(parts) == 3 and parts[1] in {"0", "1"}:
        provider_id, thinking_flag, model_name = parts
        force_disable_thinking = thinking_flag == "1"
    else:
        provider_id, model_name = model_id.split(":", 1)
    for provider in _load_user_model_providers(owner_subject):
        if str(provider["id"]) != str(provider_id) or not provider["is_active"]:
            continue
        for model_config in provider["models"]:
            option = _model_option(provider, model_config)
            if option["model_name"] == model_name and (
                force_disable_thinking is None or option["force_disable_thinking"] == force_disable_thinking
            ):
                return option
        return None
    return None


def _admin_overview_range() -> dict:
    today = date.today()
    default_start = today
    start_date = _date_arg("start_date", default_start)
    end_date = _date_arg("end_date", today)
    if start_date > end_date:
        start_date, end_date = end_date, start_date
    return {
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "start_at": f"{start_date.isoformat()} 00:00:00",
        "end_at": f"{(end_date + timedelta(days=1)).isoformat()} 00:00:00",
        "days": (end_date - start_date).days + 1,
        "today": today.isoformat(),
        "last_7_start": (today - timedelta(days=6)).isoformat(),
        "last_30_start": (today - timedelta(days=29)).isoformat(),
    }


def _date_arg(name: str, default: date) -> date:
    value = request.args.get(name, "").strip()
    if not value:
        return default
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return default


def _admin_overview_data(start_at: str, end_at: str) -> dict:
    db = get_db()
    mode_clause, mode_params = _mode_subject_filter("")
    totals = db.execute(
        f"""
        SELECT
            COUNT(*) AS tasks,
            COUNT(DISTINCT COALESCE(owner_subject, 'ip:' || ip)) AS users,
            COALESCE(SUM(CASE WHEN task_type = ? THEN 1 ELSE 0 END), 0) AS document_tasks,
            COALESCE(SUM(CASE WHEN task_type = ? THEN 1 ELSE 0 END), 0) AS consistency_tasks,
            COALESCE(SUM(CASE WHEN task_type = ? THEN 1 ELSE 0 END), 0) AS language_consistency_tasks,
            COALESCE(SUM(CASE WHEN task_type = ? THEN 1 ELSE 0 END), 0) AS image_tasks,
            COALESCE(SUM(CASE WHEN task_type = ? THEN 1 ELSE 0 END), 0) AS video_tasks,
            COALESCE(SUM(CASE WHEN status = 'queued' THEN 1 ELSE 0 END), 0) AS queued,
            COALESCE(SUM(CASE WHEN status = 'running' THEN 1 ELSE 0 END), 0) AS running,
            COALESCE(SUM(CASE WHEN status = 'completed' THEN 1 ELSE 0 END), 0) AS completed,
            COALESCE(SUM(CASE WHEN status = 'failed' THEN 1 ELSE 0 END), 0) AS failed,
            COALESCE(SUM(CASE WHEN status = 'canceled' THEN 1 ELSE 0 END), 0) AS canceled
        FROM tasks
        WHERE created_at >= ? AND created_at < ? AND {mode_clause}
        """,
        (
            DOCUMENT_TASK_TYPE,
            CONSISTENCY_TASK_TYPE,
            LANGUAGE_CONSISTENCY_TASK_TYPE,
            IMAGE_TASK_TYPE,
            VIDEO_TASK_TYPE,
            start_at,
            end_at,
            *mode_params,
        ),
    ).fetchone()
    totals = dict(totals or {})
    totals["report_items"] = _admin_report_item_totals_for_where(
        f"created_at >= ? AND created_at < ? AND {mode_clause}",
        (start_at, end_at, *mode_params),
    )
    mode_clause, mode_params = _mode_subject_filter("")
    daily_rows = db.execute(
        f"""
        SELECT
            substr(created_at, 1, 10) AS day,
            COUNT(DISTINCT COALESCE(owner_subject, 'ip:' || ip)) AS users,
            COUNT(*) AS tasks,
            COALESCE(SUM(CASE WHEN task_type = ? THEN 1 ELSE 0 END), 0) AS document_tasks,
            COALESCE(SUM(CASE WHEN task_type = ? THEN 1 ELSE 0 END), 0) AS consistency_tasks,
            COALESCE(SUM(CASE WHEN task_type = ? THEN 1 ELSE 0 END), 0) AS language_consistency_tasks,
            COALESCE(SUM(CASE WHEN task_type = ? THEN 1 ELSE 0 END), 0) AS image_tasks,
            COALESCE(SUM(CASE WHEN task_type = ? THEN 1 ELSE 0 END), 0) AS video_tasks
        FROM tasks
        WHERE created_at >= ? AND created_at < ? AND {mode_clause}
        GROUP BY day
        ORDER BY day DESC
        """,
        (
            DOCUMENT_TASK_TYPE,
            CONSISTENCY_TASK_TYPE,
            LANGUAGE_CONSISTENCY_TASK_TYPE,
            IMAGE_TASK_TYPE,
            VIDEO_TASK_TYPE,
            start_at,
            end_at,
            *mode_params,
        ),
    ).fetchall()
    join_ip_usernames = _auth_mode() == "ip"
    ip_username_join = "LEFT JOIN ip_usernames iu ON iu.ip = t.ip" if join_ip_usernames else ""
    username_expr = (
        "COALESCE(NULLIF(MAX(iu.username), ''), NULLIF(MAX(t.owner_name_snapshot), ''), NULLIF(MAX(t.username_snapshot), ''))"
        if join_ip_usernames
        else "COALESCE(NULLIF(MAX(t.owner_name_snapshot), ''), NULLIF(MAX(t.username_snapshot), ''))"
    )
    mode_clause, mode_params = _mode_subject_filter("t")
    user_rows = db.execute(
        f"""
        SELECT
            COALESCE(t.owner_subject, 'ip:' || t.ip) AS subject,
            MIN(t.ip) AS ip,
            {username_expr} AS username,
            COUNT(*) AS tasks,
            COALESCE(SUM(CASE WHEN t.task_type = ? THEN 1 ELSE 0 END), 0) AS document_tasks,
            COALESCE(SUM(CASE WHEN t.task_type = ? THEN 1 ELSE 0 END), 0) AS consistency_tasks,
            COALESCE(SUM(CASE WHEN t.task_type = ? THEN 1 ELSE 0 END), 0) AS language_consistency_tasks,
            COALESCE(SUM(CASE WHEN t.task_type = ? THEN 1 ELSE 0 END), 0) AS image_tasks,
            COALESCE(SUM(CASE WHEN t.task_type = ? THEN 1 ELSE 0 END), 0) AS video_tasks,
            MAX(t.created_at) AS last_task_at
        FROM tasks t
        {ip_username_join}
        WHERE t.created_at >= ? AND t.created_at < ? AND {mode_clause}
        GROUP BY COALESCE(t.owner_subject, 'ip:' || t.ip)
        ORDER BY tasks DESC, last_task_at DESC, COALESCE(t.owner_subject, 'ip:' || t.ip) ASC
        LIMIT 10
        """,
        (
            DOCUMENT_TASK_TYPE,
            CONSISTENCY_TASK_TYPE,
            LANGUAGE_CONSISTENCY_TASK_TYPE,
            IMAGE_TASK_TYPE,
            VIDEO_TASK_TYPE,
            start_at,
            end_at,
            *mode_params,
        ),
    ).fetchall()
    return {
        "totals": totals,
        "daily_rows": daily_rows,
        "user_rows": user_rows,
    }


def _admin_totals(task_type: str = DOCUMENT_TASK_TYPE) -> dict:
    db = get_db()
    mode_clause, mode_params = _mode_subject_filter("")
    row = db.execute(
        f"""
        SELECT
            COUNT(*) AS tasks,
            COALESCE(SUM(CASE WHEN status = 'queued' THEN 1 ELSE 0 END), 0) AS queued,
            COALESCE(SUM(CASE WHEN status = 'running' THEN 1 ELSE 0 END), 0) AS running,
            COALESCE(SUM(CASE WHEN status = 'completed' THEN 1 ELSE 0 END), 0) AS completed,
            COUNT(DISTINCT COALESCE(owner_subject, 'ip:' || ip)) AS users
        FROM tasks
        WHERE task_type = ? AND {mode_clause}
        """,
        (task_type, *mode_params),
    ).fetchone()
    totals = dict(row or {})
    totals["ips"] = totals.get("users", 0)
    totals["report_items"] = _admin_report_item_totals(task_type, mode_clause, mode_params)
    return totals


def _admin_report_item_totals(task_type: str, mode_clause: str, mode_params: tuple[str, ...]) -> dict:
    return _admin_report_item_totals_for_where(
        f"task_type = ? AND {mode_clause}",
        (task_type, *mode_params),
    )


def _admin_report_item_totals_for_where(where_clause: str, params: tuple) -> dict:
    totals = {key: 0 for key in REPORT_ITEM_TYPE_ORDER}
    totals.update(
        {
            "accepted_issue": 0,
            "rejected_issue": 0,
            "pending_issue_acceptance": 0,
            "suppressed": 0,
        }
    )
    db = get_db()
    rows = db.execute(
        f"""
        SELECT
            t.id,
            t.task_type,
            t.updated_at,
            s.source_updated_at,
            s.suppression_version,
            s.issue_count AS issue,
            s.suggestion_count AS suggestion,
            s.non_issue_count AS non_issue,
            s.accepted_issue_count AS accepted_issue,
            s.rejected_issue_count AS rejected_issue,
            s.pending_issue_acceptance_count AS pending_issue_acceptance,
            s.suppressed_count AS suppressed
        FROM tasks t
        LEFT JOIN task_report_stats s ON s.task_id = t.id
        WHERE t.result_json IS NOT NULL
          AND t.result_json != ''
          AND {where_clause}
        """,
        params,
    ).fetchall()
    task_types = {str(row["task_type"] or DOCUMENT_TASK_TYPE) for row in rows}
    suppression_versions = _report_suppression_versions(task_types)
    stale_ids = []
    for row in rows:
        task_type = str(row["task_type"] or DOCUMENT_TASK_TYPE)
        if (
            row["source_updated_at"] == row["updated_at"]
            and row["suppression_version"] == suppression_versions.get(task_type, "0:0:")
        ):
            _add_report_counts(totals, row)
        else:
            stale_ids.append(int(row["id"]))

    if stale_ids:
        rules_by_type = {
            task_type: _enabled_report_suppression_rules(task_type)
            for task_type in task_types
        }
        cache_rows = []
        for chunk_start in range(0, len(stale_ids), 500):
            chunk = stale_ids[chunk_start : chunk_start + 500]
            placeholders = ",".join("?" for _ in chunk)
            stale_rows = db.execute(
                f"SELECT id, task_type, updated_at, result_json FROM tasks WHERE id IN ({placeholders})",
                tuple(chunk),
            ).fetchall()
            for row in stale_rows:
                task_type = str(row["task_type"] or DOCUMENT_TASK_TYPE)
                item_totals = _report_item_totals(
                    _prepare_task_results(
                        _parse_result_json(row["result_json"]),
                        task_type=task_type,
                        task_id=row["id"],
                        suppression_rules=rules_by_type.get(task_type, {}),
                    )
                )
                _add_report_counts(totals, item_totals)
                cache_rows.append(
                    (
                        row["id"],
                        row["updated_at"] or "",
                        suppression_versions.get(task_type, "0:0:"),
                        *[int(item_totals.get(key) or 0) for key in REPORT_COUNT_KEYS],
                        now_text(),
                    )
                )
        db.executemany(
            """
            INSERT INTO task_report_stats(
                task_id, source_updated_at, suppression_version,
                issue_count, suggestion_count, non_issue_count,
                accepted_issue_count, rejected_issue_count,
                pending_issue_acceptance_count, suppressed_count, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(task_id) DO UPDATE SET
                source_updated_at = excluded.source_updated_at,
                suppression_version = excluded.suppression_version,
                issue_count = excluded.issue_count,
                suggestion_count = excluded.suggestion_count,
                non_issue_count = excluded.non_issue_count,
                accepted_issue_count = excluded.accepted_issue_count,
                rejected_issue_count = excluded.rejected_issue_count,
                pending_issue_acceptance_count = excluded.pending_issue_acceptance_count,
                suppressed_count = excluded.suppressed_count,
                updated_at = excluded.updated_at
            """,
            cache_rows,
        )
        db.commit()
    return _finalize_report_counts(totals)


def _report_suppression_versions(task_types: set[str]) -> dict[str, str]:
    versions = {task_type: "0:0:" for task_type in task_types}
    if not task_types:
        return versions
    placeholders = ",".join("?" for _ in task_types)
    rows = get_db().execute(
        f"""
        SELECT task_type, COUNT(*) AS total, COALESCE(SUM(id), 0) AS id_sum, MAX(updated_at) AS latest
        FROM report_suppression_rules
        WHERE enabled = 1 AND task_type IN ({placeholders})
        GROUP BY task_type
        """,
        tuple(sorted(task_types)),
    ).fetchall()
    for row in rows:
        versions[str(row["task_type"])] = f"{int(row['total'] or 0)}:{int(row['id_sum'] or 0)}:{row['latest'] or ''}"
    return versions


def _add_report_counts(target: dict, source) -> None:
    for key in REPORT_COUNT_KEYS:
        target[key] += int(_row_value(source, key, 0) or 0)


def create_task_for_identity(identity: UserIdentity, *, admin_created: bool):
    db = get_db()
    uploads = _selected_uploads("document")
    if not uploads:
        flash("请选择要上传的文档。", "error")
        return _back_to_task_form(admin_created)
    for upload in uploads:
        if not allowed_file(upload.filename):
            flash(
                f"“{upload.filename}”不是支持的文件类型，仅支持 docx、pdf、txt、md、html、xlsx、xlsm、xls 文件。",
                "error",
            )
            return _back_to_task_form(admin_created)

    check_ids = [int(value) for value in request.form.getlist("checks") if value.isdigit()]
    if not check_ids:
        flash("请至少选择一个检查项。", "error")
        return _back_to_task_form(admin_created)
    check_snapshots = _enabled_check_item_snapshots(db, check_ids, DOCUMENT_TASK_TYPE)
    if len(check_snapshots) != len(set(check_ids)):
        flash("请选择当前可用的检查项。", "error")
        return _back_to_task_form(admin_created)

    model_id = request.form.get("model_id", "")
    model = _find_enabled_model(model_id, identity.subject)
    if model is None:
        flash("请选择可用模型。", "error")
        return _back_to_task_form(admin_created)

    saved_paths: list[Path] = []
    try:
        rows = [
            _prepare_document_task_row(upload, identity, model, check_ids, check_snapshots, saved_paths)
            for upload in uploads
        ]
    except DocumentReadError as exc:
        _remove_uploaded_files(saved_paths)
        flash(f"文档读取失败：{exc}", "error")
        return _back_to_task_form(admin_created)
    except RuntimeError as exc:
        _remove_uploaded_files(saved_paths)
        flash(str(exc), "error")
        return _back_to_task_form(admin_created)
    except Exception as exc:
        _remove_uploaded_files(saved_paths)
        current_app.logger.exception("准备单文档检查任务失败")
        flash(_unexpected_upload_preparation_message(exc), "error")
        return _back_to_task_form(admin_created)

    try:
        db.executemany(
            """
            INSERT INTO tasks(
                task_type, ip, username_snapshot, owner_subject, owner_name_snapshot, owner_source,
                original_filename, stored_filename, file_type, file_size,
                document_text, checks_json, checks_snapshot_json, provider_id, provider_name, model_name,
                api_base, api_key, request_timeout, max_input_chars, force_disable_thinking,
                status, progress, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'queued', 0, ?, ?)
            """,
            rows,
        )
        db.commit()
    except Exception:
        db.rollback()
        _remove_uploaded_files(saved_paths)
        current_app.logger.exception("创建单文档检查任务失败")
        flash("创建任务失败，请稍后再试。", "error")
        return _back_to_task_form(admin_created)
    if admin_created:
        return redirect(url_for("admin_tasks"))
    return redirect(url_for("user_tasks"))


def _prepare_document_task_row(
    upload,
    identity: UserIdentity,
    model: dict,
    check_ids: list[int],
    check_snapshots: list[dict],
    saved_paths: list[Path],
):
    file_type = extension_of(upload.filename)
    original_filename = _clean_upload_filename(upload.filename, file_type)
    created_at = now_text()
    stored_filename, destination = _upload_destination(
        original_filename,
        identity.subject,
        created_at,
        file_type,
    )
    file_size = _save_uploaded_file(upload, destination)
    saved_paths.append(destination)
    try:
        document_text = extract_text(destination, file_type).strip()
    except DocumentReadError as exc:
        raise DocumentReadError(f"“{original_filename}”：{exc}") from exc
    if not document_text:
        raise RuntimeError(f"“{original_filename}”未能提取到可检查文本。")
    prepared_document_text = format_document_text(original_filename, document_text)
    if len(prepared_document_text) > model["max_input_chars"]:
        raise RuntimeError(
            f"“{original_filename}”文档文本 {len(prepared_document_text)} 字，"
            f"超过当前模型文本上限 {model['max_input_chars']} 字。"
        )
    owner_name = identity.display_name or None
    return (
        DOCUMENT_TASK_TYPE,
        identity.ip,
        owner_name,
        identity.subject,
        owner_name,
        identity.source,
        original_filename,
        stored_filename,
        file_type,
        file_size,
        prepared_document_text,
        json.dumps(check_ids, ensure_ascii=False),
        json.dumps(check_snapshots, ensure_ascii=False),
        model["provider_id"],
        model["provider_name"],
        model["model_name"],
        model["api_base"],
        model["api_key"],
        model["request_timeout"],
        model["max_input_chars"],
        1 if model["force_disable_thinking"] else 0,
        created_at,
        created_at,
    )


def _unexpected_upload_preparation_message(error: Exception) -> str:
    detail = _compact_user_error(error)
    if detail:
        return f"文档上传或读取失败：{detail}。如持续出现请联系管理员查看日志。"
    return "文档上传或读取失败，请稍后重试；如持续出现请联系管理员查看日志。"


def _compact_user_error(error: Exception, limit: int = 180) -> str:
    text = re.sub(r"\s+", " ", str(error or "").strip())
    if not text:
        return ""
    text = re.sub(r"[A-Za-z]:[\\/][^，。；;]*", "[本地路径]", text)
    if len(text) <= limit:
        return text
    return f"{text[:limit]}..."


def create_image_task_for_identity(identity: UserIdentity, *, admin_created: bool):
    db = get_db()
    upload = request.files.get("document")
    if upload is None or not upload.filename:
        flash("请选择要提取图片的文档。", "error")
        return _back_to_task_form(admin_created, IMAGE_TASK_TYPE)
    file_type = extension_of(upload.filename)
    if file_type != "pdf":
        flash("图片检查仅支持 PDF 文件。", "error")
        return _back_to_task_form(admin_created, IMAGE_TASK_TYPE)

    check_ids = [int(value) for value in request.form.getlist("checks") if value.isdigit()]
    if not check_ids:
        flash("请至少选择一个图片检查项。", "error")
        return _back_to_task_form(admin_created, IMAGE_TASK_TYPE)
    check_snapshots = _enabled_check_item_snapshots(db, check_ids, IMAGE_TASK_TYPE)
    if len(check_snapshots) != len(set(check_ids)):
        flash("请选择当前可用的图片检查项。", "error")
        return _back_to_task_form(admin_created, IMAGE_TASK_TYPE)

    model_id = request.form.get("model_id", "")
    model = _find_enabled_model(model_id, identity.subject)
    if model is None:
        flash("请选择可用模型。", "error")
        return _back_to_task_form(admin_created, IMAGE_TASK_TYPE)

    original_filename = _clean_upload_filename(upload.filename, file_type)
    created_at = now_text()
    stored_filename, destination = _upload_destination(original_filename, identity.subject, created_at, file_type)
    image_dir = _image_output_dir_for_stored(stored_filename)
    try:
        file_size = _save_uploaded_file(upload, destination)
    except Exception:
        current_app.logger.exception("保存图片检查文档失败")
        flash("PDF 上传失败，请稍后再试。", "error")
        return _back_to_task_form(admin_created, IMAGE_TASK_TYPE)

    extracted_text = ""
    text_error = ""
    try:
        extracted_text = extract_text(destination, file_type).strip()
    except DocumentReadError as exc:
        text_error = str(exc)
        current_app.logger.warning(
            "图片检查任务未能提取文档文本 file=%s error=%s",
            original_filename,
            exc,
        )

    image_error = ""
    try:
        images = extract_images(destination, file_type, image_dir, source_filename=original_filename)
    except DocumentReadError as exc:
        images = []
        image_error = str(exc)
        _remove_directory(image_dir)
        current_app.logger.warning(
            "图片检查任务未能提取 PDF 内嵌图片 file=%s error=%s",
            original_filename,
            exc,
        )

    try:
        candidate_pages = candidate_pdf_pages_for_image_check(extracted_text, images)
        page_images, page_selection = render_pdf_page_images(
            destination,
            image_dir,
            source_filename=original_filename,
            max_pages=_image_page_check_max_pages(),
            candidate_pages=candidate_pages,
        )
    except DocumentReadError as exc:
        _remove_uploaded_file(destination)
        _remove_directory(image_dir)
        flash(f"PDF 页面截图生成失败：{exc}", "error")
        return _back_to_task_form(admin_created, IMAGE_TASK_TYPE)
    if not images and not page_images:
        _remove_uploaded_file(destination)
        _remove_directory(image_dir)
        flash("未能从 PDF 中生成可检查页面截图或提取到可检查图片。", "error")
        return _back_to_task_form(admin_created, IMAGE_TASK_TYPE)

    prepared_document_text = format_image_document_text(
        original_filename,
        images,
        document_text=extracted_text,
        text_error=text_error,
        page_images=page_images,
        page_selection=page_selection,
    )

    document_meta = {
        "source_document": {
            "original_filename": original_filename,
            "stored_filename": stored_filename,
            "file_type": file_type,
            "file_size": file_size,
        },
        "image_extraction_error": image_error,
        "page_selection": page_selection,
        "images": [
            {
                **image,
                "relative_path": f"{image_dir.name}/{image['filename']}",
            }
            for image in images
        ],
        "page_images": [
            {
                **image,
                "relative_path": f"{image_dir.name}/{image['filename']}",
            }
            for image in page_images
        ],
    }
    owner_name = identity.display_name or None
    try:
        db.execute(
            """
            INSERT INTO tasks(
                task_type, ip, username_snapshot, owner_subject, owner_name_snapshot, owner_source,
                original_filename, stored_filename, file_type, file_size,
                document_text, document_meta_json, checks_json, checks_snapshot_json,
                provider_id, provider_name, model_name, api_base, api_key, request_timeout,
                max_input_chars, force_disable_thinking,
                status, progress, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'queued', 0, ?, ?)
            """,
            (
                IMAGE_TASK_TYPE,
                identity.ip,
                owner_name,
                identity.subject,
                owner_name,
                identity.source,
                original_filename,
                stored_filename,
                file_type,
                file_size,
                prepared_document_text,
                json.dumps(document_meta, ensure_ascii=False),
                json.dumps(check_ids, ensure_ascii=False),
                json.dumps(check_snapshots, ensure_ascii=False),
                model["provider_id"],
                model["provider_name"],
                model["model_name"],
                model["api_base"],
                model["api_key"],
                model["request_timeout"],
                model["max_input_chars"],
                1 if model["force_disable_thinking"] else 0,
                created_at,
                created_at,
            ),
        )
        db.commit()
    except Exception:
        db.rollback()
        _remove_uploaded_file(destination)
        _remove_directory(image_dir)
        current_app.logger.exception("创建图片检查任务失败")
        flash("创建图片检查任务失败，请稍后再试。", "error")
        return _back_to_task_form(admin_created, IMAGE_TASK_TYPE)
    return redirect(url_for(_task_list_endpoint(admin_created, IMAGE_TASK_TYPE)))


def create_video_task_for_identity(identity: UserIdentity, *, admin_created: bool):
    db = get_db()
    upload = request.files.get("video")
    if upload is None or not upload.filename:
        flash("请选择要质检的视频。", "error")
        return _back_to_task_form(admin_created, VIDEO_TASK_TYPE)
    if not allowed_video_file(upload.filename):
        flash("视频检查仅支持 mp4、mov、mkv、webm、avi、m4v 文件。", "error")
        return _back_to_task_form(admin_created, VIDEO_TASK_TYPE)

    check_ids = [int(value) for value in request.form.getlist("checks") if value.isdigit()]
    if not check_ids:
        flash("请至少选择一个视频检查项。", "error")
        return _back_to_task_form(admin_created, VIDEO_TASK_TYPE)
    check_snapshots = _enabled_check_item_snapshots(db, check_ids, VIDEO_TASK_TYPE)
    if len(check_snapshots) != len(set(check_ids)):
        flash("请选择当前可用的视频检查项。", "error")
        return _back_to_task_form(admin_created, VIDEO_TASK_TYPE)

    model_id = request.form.get("model_id", "")
    model = _find_enabled_model(model_id, identity.subject)
    if model is None:
        flash("请选择可用模型。", "error")
        return _back_to_task_form(admin_created, VIDEO_TASK_TYPE)

    file_type = video_extension_of(upload.filename)
    original_filename = _clean_upload_filename(upload.filename, file_type)
    created_at = now_text()
    stored_filename, destination = _upload_destination(original_filename, identity.subject, created_at, file_type)
    frame_dir = _image_output_dir_for_stored(stored_filename)
    try:
        file_size = _save_uploaded_file(upload, destination)
    except Exception:
        current_app.logger.exception("保存视频检查文件失败")
        flash("视频上传失败，请稍后再试。", "error")
        return _back_to_task_form(admin_created, VIDEO_TASK_TYPE)

    try:
        frames, frame_selection = extract_video_frames(
            destination,
            frame_dir,
            source_filename=original_filename,
        )
    except DocumentReadError as exc:
        _remove_uploaded_file(destination)
        _remove_directory(frame_dir)
        flash(f"视频抽帧失败：{exc}", "error")
        return _back_to_task_form(admin_created, VIDEO_TASK_TYPE)
    if not frames:
        _remove_uploaded_file(destination)
        _remove_directory(frame_dir)
        flash("未能从视频中抽取可检查画面。", "error")
        return _back_to_task_form(admin_created, VIDEO_TASK_TYPE)

    prepared_document_text = format_video_document_text(original_filename, frames, frame_selection)
    if len(prepared_document_text) > model["max_input_chars"]:
        _remove_uploaded_file(destination)
        _remove_directory(frame_dir)
        flash(f"视频帧上下文 {len(prepared_document_text)} 字，超过当前模型文本上限 {model['max_input_chars']} 字。", "error")
        return _back_to_task_form(admin_created, VIDEO_TASK_TYPE)

    document_meta = {
        "source_video": {
            "original_filename": original_filename,
            "stored_filename": stored_filename,
            "file_type": file_type,
            "file_size": file_size,
        },
        "frame_selection": frame_selection,
        "frames": [
            {
                **frame,
                "relative_path": f"{frame_dir.name}/{frame['filename']}",
            }
            for frame in frames
        ],
    }
    owner_name = identity.display_name or None
    try:
        db.execute(
            """
            INSERT INTO tasks(
                task_type, ip, username_snapshot, owner_subject, owner_name_snapshot, owner_source,
                original_filename, stored_filename, file_type, file_size,
                document_text, document_meta_json, checks_json, checks_snapshot_json,
                provider_id, provider_name, model_name, api_base, api_key, request_timeout,
                max_input_chars, force_disable_thinking,
                status, progress, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'queued', 0, ?, ?)
            """,
            (
                VIDEO_TASK_TYPE,
                identity.ip,
                owner_name,
                identity.subject,
                owner_name,
                identity.source,
                original_filename,
                stored_filename,
                file_type,
                file_size,
                prepared_document_text,
                json.dumps(document_meta, ensure_ascii=False),
                json.dumps(check_ids, ensure_ascii=False),
                json.dumps(check_snapshots, ensure_ascii=False),
                model["provider_id"],
                model["provider_name"],
                model["model_name"],
                model["api_base"],
                model["api_key"],
                model["request_timeout"],
                model["max_input_chars"],
                1 if model["force_disable_thinking"] else 0,
                created_at,
                created_at,
            ),
        )
        db.commit()
    except Exception:
        db.rollback()
        _remove_uploaded_file(destination)
        _remove_directory(frame_dir)
        current_app.logger.exception("创建视频检查任务失败")
        flash("创建视频检查任务失败，请稍后再试。", "error")
        return _back_to_task_form(admin_created, VIDEO_TASK_TYPE)
    return redirect(url_for(_task_list_endpoint(admin_created, VIDEO_TASK_TYPE)))


def create_consistency_task_for_identity(identity: UserIdentity, *, admin_created: bool):
    db = get_db()
    master_uploads = _selected_uploads("master_documents")
    related_uploads = _selected_uploads("related_documents")
    if not _validate_consistency_uploads(master_uploads, "素材文档", CONSISTENCY_MAX_MATERIAL_FILES):
        return _back_to_task_form(admin_created, CONSISTENCY_TASK_TYPE)
    if not _validate_consistency_uploads(related_uploads, "资料", CONSISTENCY_MAX_DATA_FILES):
        return _back_to_task_form(admin_created, CONSISTENCY_TASK_TYPE)

    check_ids = [int(value) for value in request.form.getlist("checks") if value.isdigit()]
    if not check_ids:
        flash("请至少选择一个多文档对照项。", "error")
        return _back_to_task_form(admin_created, CONSISTENCY_TASK_TYPE)
    check_snapshots = _enabled_check_item_snapshots(db, check_ids, CONSISTENCY_TASK_TYPE)
    if len(check_snapshots) != len(set(check_ids)):
        flash("请选择当前可用的多文档对照项。", "error")
        return _back_to_task_form(admin_created, CONSISTENCY_TASK_TYPE)

    model_id = request.form.get("model_id", "")
    model = _find_enabled_model(model_id, identity.subject)
    if model is None:
        flash("请选择可用模型。", "error")
        return _back_to_task_form(admin_created, CONSISTENCY_TASK_TYPE)

    created_at = now_text()
    saved_paths = []
    try:
        master_files = _save_consistency_upload_group(master_uploads, identity.subject, created_at, "素材文档", saved_paths)
        related_files = _save_consistency_upload_group(related_uploads, identity.subject, created_at, "资料", saved_paths)
    except DocumentReadError as exc:
        _remove_uploaded_files(saved_paths)
        flash(f"文档读取失败：{exc}", "error")
        return _back_to_task_form(admin_created, CONSISTENCY_TASK_TYPE)
    except Exception:
        _remove_uploaded_files(saved_paths)
        current_app.logger.exception("准备多文档对照任务失败")
        flash("文档上传失败，请稍后再试。", "error")
        return _back_to_task_form(admin_created, CONSISTENCY_TASK_TYPE)

    validation_text = _compose_consistency_validation_text(
        [
            {"label": "素材文档", "files": master_files},
            {"label": "资料", "files": related_files},
        ]
    )
    if len(validation_text) > model["max_input_chars"]:
        _remove_uploaded_files(saved_paths)
        flash(f"文档文本 {len(validation_text)} 字，超过当前模型文本上限 {model['max_input_chars']} 字。", "error")
        return _back_to_task_form(admin_created, CONSISTENCY_TASK_TYPE)

    document_meta = {
        "groups": [
            {
                "role": "master",
                "label": "素材文档",
                "files": [_persisted_file_info(file_info) for file_info in master_files],
            },
            {
                "role": "related",
                "label": "资料",
                "files": [_persisted_file_info(file_info) for file_info in related_files],
            },
        ]
    }
    all_files = master_files + related_files
    first_file = all_files[0]
    file_size = sum(file_info["file_size"] for file_info in all_files)
    original_filename = _consistency_title_from_groups(document_meta["groups"])
    owner_name = identity.display_name or None

    try:
        db.execute(
            """
            INSERT INTO tasks(
                task_type, ip, username_snapshot, owner_subject, owner_name_snapshot, owner_source,
                original_filename, stored_filename, file_type, file_size,
                document_text, document_meta_json, checks_json, checks_snapshot_json,
                provider_id, provider_name, model_name, api_base, api_key, request_timeout,
                max_input_chars, force_disable_thinking,
                status, progress, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'queued', 0, ?, ?)
            """,
            (
                CONSISTENCY_TASK_TYPE,
                identity.ip,
                owner_name,
                identity.subject,
                owner_name,
                identity.source,
                original_filename,
                first_file["stored_filename"],
                "多文档",
                file_size,
                validation_text,
                json.dumps(document_meta, ensure_ascii=False),
                json.dumps(check_ids, ensure_ascii=False),
                json.dumps(check_snapshots, ensure_ascii=False),
                model["provider_id"],
                model["provider_name"],
                model["model_name"],
                model["api_base"],
                model["api_key"],
                model["request_timeout"],
                model["max_input_chars"],
                1 if model["force_disable_thinking"] else 0,
                created_at,
                created_at,
            ),
        )
        db.commit()
    except Exception:
        db.rollback()
        _remove_uploaded_files(saved_paths)
        current_app.logger.exception("创建多文档对照任务失败")
        flash("创建多文档对照任务失败，请稍后再试。", "error")
        return _back_to_task_form(admin_created, CONSISTENCY_TASK_TYPE)
    return redirect(url_for(_task_list_endpoint(admin_created, CONSISTENCY_TASK_TYPE)))


def create_language_consistency_task_for_identity(identity: UserIdentity, *, admin_created: bool):
    db = get_db()
    submission_token = _request_submission_token()
    if _language_consistency_submission_exists(db, identity.subject, submission_token):
        flash("该跨语种检查任务已提交，无需重复提交。", "success")
        return _back_to_task_form(admin_created, LANGUAGE_CONSISTENCY_TASK_TYPE)

    document_a = request.files.get("document_a")
    document_b = request.files.get("document_b")
    if not _validate_language_consistency_upload(document_a, "文档A"):
        return _back_to_task_form(admin_created, LANGUAGE_CONSISTENCY_TASK_TYPE)
    if not _validate_language_consistency_upload(document_b, "文档B"):
        return _back_to_task_form(admin_created, LANGUAGE_CONSISTENCY_TASK_TYPE)

    check_ids = [int(value) for value in request.form.getlist("checks") if value.isdigit()]
    if not check_ids:
        flash("请至少选择一个跨语种检查项。", "error")
        return _back_to_task_form(admin_created, LANGUAGE_CONSISTENCY_TASK_TYPE)
    check_snapshots = _enabled_check_item_snapshots(db, check_ids, LANGUAGE_CONSISTENCY_TASK_TYPE)
    if len(check_snapshots) != len(set(check_ids)):
        flash("请选择当前可用的跨语种检查项。", "error")
        return _back_to_task_form(admin_created, LANGUAGE_CONSISTENCY_TASK_TYPE)

    model_id = request.form.get("model_id", "")
    model = _find_enabled_model(model_id, identity.subject)
    if model is None:
        flash("请选择可用模型。", "error")
        return _back_to_task_form(admin_created, LANGUAGE_CONSISTENCY_TASK_TYPE)

    created_at = now_text()
    saved_paths = []
    try:
        file_a = _save_consistency_upload_group([document_a], identity.subject, created_at, "文档A", saved_paths)[0]
        file_b = _save_consistency_upload_group([document_b], identity.subject, created_at, "文档B", saved_paths)[0]
    except DocumentReadError as exc:
        _remove_uploaded_files(saved_paths)
        flash(f"文档读取失败：{exc}", "error")
        return _back_to_task_form(admin_created, LANGUAGE_CONSISTENCY_TASK_TYPE)
    except Exception:
        _remove_uploaded_files(saved_paths)
        current_app.logger.exception("准备跨语种检查任务失败")
        flash("文档上传失败，请稍后再试。", "error")
        return _back_to_task_form(admin_created, LANGUAGE_CONSISTENCY_TASK_TYPE)

    validation_text = _compose_language_consistency_validation_text(file_a, file_b)
    if len(validation_text) > model["max_input_chars"]:
        _remove_uploaded_files(saved_paths)
        flash(f"文档文本 {len(validation_text)} 字，超过当前模型文本上限 {model['max_input_chars']} 字。", "error")
        return _back_to_task_form(admin_created, LANGUAGE_CONSISTENCY_TASK_TYPE)

    document_meta = {
        "groups": [
            {
                "role": "document_a",
                "label": "文档A",
                "files": [_persisted_file_info(file_a)],
            },
            {
                "role": "document_b",
                "label": "文档B",
                "files": [_persisted_file_info(file_b)],
            },
        ],
        "static_precheck": _language_consistency_static_summary(file_a, file_b),
    }
    file_size = file_a["file_size"] + file_b["file_size"]
    original_filename = f"跨语种检查：{file_a['original_filename']} / {file_b['original_filename']}"
    owner_name = identity.display_name or None

    try:
        db.execute(
            """
            INSERT INTO tasks(
                task_type, ip, username_snapshot, owner_subject, owner_name_snapshot, owner_source, submission_token,
                original_filename, stored_filename, file_type, file_size,
                document_text, document_meta_json, checks_json, checks_snapshot_json,
                provider_id, provider_name, model_name, api_base, api_key, request_timeout,
                max_input_chars, force_disable_thinking,
                status, progress, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'queued', 0, ?, ?)
            """,
            (
                LANGUAGE_CONSISTENCY_TASK_TYPE,
                identity.ip,
                owner_name,
                identity.subject,
                owner_name,
                identity.source,
                submission_token,
                original_filename,
                file_a["stored_filename"],
                "双文档",
                file_size,
                validation_text,
                json.dumps(document_meta, ensure_ascii=False),
                json.dumps(check_ids, ensure_ascii=False),
                json.dumps(check_snapshots, ensure_ascii=False),
                model["provider_id"],
                model["provider_name"],
                model["model_name"],
                model["api_base"],
                model["api_key"],
                model["request_timeout"],
                model["max_input_chars"],
                1 if model["force_disable_thinking"] else 0,
                created_at,
                created_at,
            ),
        )
        db.commit()
    except sqlite3.IntegrityError:
        db.rollback()
        duplicate = _language_consistency_submission_exists(db, identity.subject, submission_token)
        _remove_uploaded_files(saved_paths)
        if duplicate:
            flash("该跨语种检查任务已提交，无需重复提交。", "success")
            return _back_to_task_form(admin_created, LANGUAGE_CONSISTENCY_TASK_TYPE)
        current_app.logger.exception("创建跨语种检查任务失败")
        flash("创建跨语种检查任务失败，请稍后再试。", "error")
        return _back_to_task_form(admin_created, LANGUAGE_CONSISTENCY_TASK_TYPE)
    except Exception:
        db.rollback()
        _remove_uploaded_files(saved_paths)
        current_app.logger.exception("创建跨语种检查任务失败")
        flash("创建跨语种检查任务失败，请稍后再试。", "error")
        return _back_to_task_form(admin_created, LANGUAGE_CONSISTENCY_TASK_TYPE)
    return redirect(url_for(_task_list_endpoint(admin_created, LANGUAGE_CONSISTENCY_TASK_TYPE)))


def _request_submission_token() -> str:
    value = str(request.form.get("submission_token") or "").strip().lower()
    return value if SUBMISSION_TOKEN_RE.fullmatch(value) else uuid.uuid4().hex


def _language_consistency_submission_exists(db, owner_subject: str, submission_token: str) -> bool:
    return (
        db.execute(
            """
            SELECT 1
            FROM tasks
            WHERE task_type = ? AND owner_subject = ? AND submission_token = ?
            LIMIT 1
            """,
            (LANGUAGE_CONSISTENCY_TASK_TYPE, owner_subject, submission_token),
        ).fetchone()
        is not None
    )


def _back_to_task_form(admin_created: bool, task_type: str = DOCUMENT_TASK_TYPE):
    return redirect(url_for(_task_list_endpoint(admin_created, task_type)))


def _task_list_endpoint(admin_created: bool, task_type: str | None = DOCUMENT_TASK_TYPE) -> str:
    if task_type == CONSISTENCY_TASK_TYPE:
        return "admin_consistency" if admin_created else "user_consistency"
    if task_type == LANGUAGE_CONSISTENCY_TASK_TYPE:
        return "admin_language_consistency" if admin_created else "user_language_consistency"
    if task_type == IMAGE_TASK_TYPE:
        return "admin_images" if admin_created else "user_images"
    if task_type == VIDEO_TASK_TYPE:
        return "admin_videos" if admin_created else "user_videos"
    return "admin_tasks" if admin_created else "user_tasks"


def _selected_uploads(field_name: str):
    return [upload for upload in request.files.getlist(field_name) if upload and upload.filename]


def _validate_consistency_uploads(uploads: list, label: str, max_files: int) -> bool:
    if not uploads:
        flash(f"请至少选择 1 个{label}。", "error")
        return False
    if len(uploads) > max_files:
        flash(f"{label}最多上传 {max_files} 个。", "error")
        return False
    for upload in uploads:
        if not allowed_file(upload.filename):
            flash(f"{label}仅支持 docx、pdf、txt、md、html、xlsx、xlsm、xls 文件。", "error")
            return False
    return True


def _validate_language_consistency_upload(upload, label: str) -> bool:
    if upload is None or not upload.filename:
        flash(f"请选择{label}。", "error")
        return False
    if not allowed_file(upload.filename):
        flash(f"{label}仅支持 docx、pdf、txt、md、html、xlsx、xlsm、xls 文件。", "error")
        return False
    return True


def _save_consistency_upload_group(uploads: list, ip: str, created_at: str, label: str, saved_paths: list[Path]) -> list[dict]:
    files = []
    for upload in uploads:
        file_type = extension_of(upload.filename)
        original_filename = _clean_upload_filename(upload.filename, file_type)
        stored_filename, destination = _upload_destination(original_filename, ip, created_at, file_type)
        file_size = _save_uploaded_file(upload, destination)
        saved_paths.append(destination)
        try:
            text = extract_text(destination, file_type).strip()
        except DocumentReadError as exc:
            raise DocumentReadError(f"{label}“{original_filename}”：{exc}") from exc
        if not text:
            raise DocumentReadError(f"{label}“{original_filename}”未能提取到可检查文本")
        files.append(
            {
                "original_filename": original_filename,
                "stored_filename": stored_filename,
                "file_type": file_type,
                "file_size": file_size,
                "text": text,
            }
        )
    return files


def _compose_consistency_validation_text(groups: list[dict]) -> str:
    sections = []
    for group in groups:
        group_parts = [f"# {group['label']}"]
        for index, file_info in enumerate(group["files"], start=1):
            group_parts.append(f"## {group['label']}{index}：{file_info['original_filename']}\n{file_info['text']}")
        sections.append("\n\n".join(group_parts))
    return "\n\n".join(sections).strip()


def _compose_language_consistency_validation_text(file_a: dict, file_b: dict) -> str:
    return "\n\n".join(
        [
            "# 静态预检摘要\n"
            + _language_consistency_static_summary(file_a, file_b)
            + "\n\n说明：静态预检仅提供优先核对线索，最终差异判断需结合两份文档正文。",
            f"# 文档A：{file_a['original_filename']}\n{file_a['text']}",
            f"# 文档B：{file_b['original_filename']}\n{file_b['text']}",
        ]
    ).strip()


def _language_consistency_static_summary(file_a: dict, file_b: dict) -> str:
    profile_a = _document_static_profile(file_a)
    profile_b = _document_static_profile(file_b)
    only_a = _limited_sorted(profile_a["tokens"] - profile_b["tokens"], 40)
    only_b = _limited_sorted(profile_b["tokens"] - profile_a["tokens"], 40)
    ratio = _safe_ratio(profile_b["nonspace_chars"], profile_a["nonspace_chars"])
    lines = [
        (
            f"- 文档A：{file_a['original_filename']}；格式：{file_a['file_type']}；"
            f"语种估计：{profile_a['language']}；非空白字符：{profile_a['nonspace_chars']}；"
            f"段落：{profile_a['paragraphs']}；标题线索：{len(profile_a['headings'])}"
        ),
        (
            f"- 文档B：{file_b['original_filename']}；格式：{file_b['file_type']}；"
            f"语种估计：{profile_b['language']}；非空白字符：{profile_b['nonspace_chars']}；"
            f"段落：{profile_b['paragraphs']}；标题线索：{len(profile_b['headings'])}"
        ),
        f"- 长度比例：文档B / 文档A = {ratio}",
        f"- 文档A独有硬线索：{_format_preview_list(only_a)}",
        f"- 文档B独有硬线索：{_format_preview_list(only_b)}",
        f"- 文档A标题线索：{_format_preview_list(profile_a['headings'])}",
        f"- 文档B标题线索：{_format_preview_list(profile_b['headings'])}",
    ]
    return "\n".join(lines)


def _document_static_profile(file_info: dict) -> dict:
    text = str(file_info.get("text") or "")
    nonspace_text = re.sub(r"\s+", "", text)
    paragraphs = [part for part in re.split(r"\n\s*\n+", text.strip()) if part.strip()]
    tokens = {
        _normalize_static_token(match.group(0))
        for match in LANGUAGE_STATIC_TOKEN_RE.finditer(text)
    }
    tokens = {token for token in tokens if token}
    cjk_chars = len(re.findall(r"[\u4e00-\u9fff]", text))
    latin_chars = len(re.findall(r"[A-Za-z]", text))
    return {
        "language": _estimate_text_language(cjk_chars, latin_chars),
        "nonspace_chars": len(nonspace_text),
        "paragraphs": len(paragraphs),
        "tokens": tokens,
        "headings": _extract_static_headings(text, 12),
    }


def _extract_static_headings(text: str, limit: int) -> list[str]:
    headings = []
    seen = set()
    for line in text.splitlines():
        value = re.sub(r"\s+", " ", line).strip()
        if not value or len(value) > 120 or not LANGUAGE_HEADING_RE.match(value):
            continue
        if value in seen:
            continue
        seen.add(value)
        headings.append(value)
        if len(headings) >= limit:
            break
    return headings


def _estimate_text_language(cjk_chars: int, latin_chars: int) -> str:
    if cjk_chars >= 40 and latin_chars >= 80 and min(cjk_chars, latin_chars) / max(cjk_chars, latin_chars) >= 0.2:
        return "中英混合"
    if cjk_chars >= max(20, latin_chars):
        return "中文为主"
    if latin_chars >= max(40, cjk_chars):
        return "拉丁语系为主"
    if cjk_chars or latin_chars:
        return "语种特征较少，需人工确认"
    return "未识别"


def _normalize_static_token(value: str) -> str:
    return value.strip(" \t\r\n,.;:，。；：、()（）[]【】<>《》\"'“”‘’").lower()


def _limited_sorted(values: set[str] | list[str], limit: int) -> list[str]:
    return sorted(values, key=lambda value: (len(value), value))[:limit]


def _format_preview_list(values: list[str]) -> str:
    return "、".join(values) if values else "未发现"


def _safe_ratio(numerator: int, denominator: int) -> str:
    if denominator <= 0:
        return "无法计算"
    return f"{numerator / denominator:.2f}"


def _persisted_file_info(file_info: dict) -> dict:
    return {
        "original_filename": file_info["original_filename"],
        "stored_filename": file_info["stored_filename"],
        "file_type": file_info["file_type"],
        "file_size": file_info["file_size"],
    }


def admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not _platform_enabled():
            return view(*args, **kwargs)
        if not session.get("admin_logged_in"):
            return redirect(url_for("admin_login"))
        return view(*args, **kwargs)

    return wrapped


def _get_task_or_404(task_id: int):
    join_ip_usernames = _auth_mode() == "ip"
    ip_username_join = "LEFT JOIN ip_usernames iu ON iu.ip = t.ip" if join_ip_usernames else ""
    owner_name_expr = (
        "COALESCE(NULLIF(iu.username, ''), NULLIF(t.owner_name_snapshot, ''), NULLIF(t.username_snapshot, ''), '')"
        if join_ip_usernames
        else "COALESCE(NULLIF(t.owner_name_snapshot, ''), NULLIF(t.username_snapshot, ''), '')"
    )
    clauses = ["t.id = ?"]
    params: list[object] = [task_id]
    if _platform_enabled():
        mode_clause, mode_params = _mode_subject_filter("t")
        clauses.append(mode_clause)
        params.extend(mode_params)
    task = get_db().execute(
        f"""
        SELECT t.*,
               {owner_name_expr} AS current_owner_name,
               {owner_name_expr} AS current_username,
               COALESCE(t.owner_subject, 'ip:' || t.ip) AS effective_owner_subject
        FROM tasks t
        {ip_username_join}
        WHERE {' AND '.join(clauses)}
        """,
        tuple(params),
    ).fetchone()
    if task is None:
        abort(404)
    return task


def _get_user_task(task_id: int):
    identity = _current_user_identity()
    task = get_db().execute(
        """
        SELECT t.*,
               COALESCE(NULLIF(t.owner_name_snapshot, ''), NULLIF(t.username_snapshot, ''), '') AS current_owner_name,
               COALESCE(NULLIF(t.owner_name_snapshot, ''), NULLIF(t.username_snapshot, ''), '') AS current_username,
               COALESCE(t.owner_subject, 'ip:' || t.ip) AS effective_owner_subject
        FROM tasks t
        WHERE t.id = ? AND COALESCE(t.owner_subject, 'ip:' || t.ip) = ?
        """,
        (task_id, identity.subject),
    ).fetchone()
    if task is None:
        abort(404)
    return task


def _get_user_task_or_local_admin(task_id: int):
    if not _platform_enabled():
        return _get_task_or_404(task_id)
    return _get_user_task(task_id)


def _cancel_task(task):
    if task["status"] in {"completed", "failed", "canceled"}:
        return
    db = get_db()
    db.execute(
        """
        UPDATE tasks
        SET cancel_requested = 1,
            status = 'canceled',
            progress = 0,
            api_key = NULL,
            claim_token = NULL,
            lease_expires_at = NULL,
            updated_at = ?,
            finished_at = ?
        WHERE id = ? AND status NOT IN ('completed', 'failed', 'canceled')
        """,
        (now_text(), now_text(), task["id"]),
    )
    db.commit()


def _delete_task(task):
    if task["status"] not in DELETABLE_TASK_STATUSES:
        flash("排队中或运行中的任务不能直接删除，请先取消后再删除。", "error")
        return False
    db = get_db()
    paths = _task_upload_paths(task)
    image_dirs = {path.parent for path in paths if _image_folder() in path.parents}
    failures = _remove_uploaded_files(paths)
    if failures:
        current_app.logger.warning(
            "删除任务文件失败 task_id=%s failures=%s",
            task["id"],
            "; ".join(f"{path}: {error}" for path, error in failures),
        )
        flash(
            f"任务文件正被其他程序使用，暂时无法删除：{describe_failures(failures)}。"
            "请关闭正在下载、预览或扫描该文件的程序后稍后重试。",
            "error",
        )
        return False
    for image_dir in image_dirs:
        _remove_empty_directory(image_dir)
    delete_task_record(db, task["id"])
    db.commit()
    return True


def _bulk_delete_tasks(task_loader, *, admin_created: bool):
    raw_task_ids = request.form.getlist("task_ids")
    if len(raw_task_ids) > TASKS_PER_PAGE:
        flash(f"每次最多批量删除 {TASKS_PER_PAGE} 个任务。", "error")
        return redirect(_task_action_redirect("admin_tasks" if admin_created else "user_tasks"))

    task_ids = []
    for raw_task_id in raw_task_ids:
        try:
            task_id = int(raw_task_id)
        except (TypeError, ValueError):
            continue
        if task_id > 0 and task_id not in task_ids:
            task_ids.append(task_id)

    if not task_ids:
        flash("请先选择需要删除的任务。", "error")
        return redirect(_task_action_redirect("admin_tasks" if admin_created else "user_tasks"))

    tasks = [task_loader(task_id) for task_id in task_ids]
    fallback_endpoint = _task_list_endpoint(admin_created, tasks[0]["task_type"])
    redirect_url = _task_action_redirect(fallback_endpoint)
    deletable_tasks = [task for task in tasks if task["status"] in DELETABLE_TASK_STATUSES]
    skipped_count = len(tasks) - len(deletable_tasks)
    deleted_count = sum(1 for task in deletable_tasks if _delete_task(task))

    if deleted_count:
        flash(f"已批量删除 {deleted_count} 个任务。", "success")
    if skipped_count:
        flash(f"已跳过 {skipped_count} 个排队中或运行中的任务。", "error")
    return redirect(redirect_url)


def _task_action_redirect(default_endpoint: str):
    return _safe_next_path(request.form.get("next"), url_for(default_endpoint))


def _download_task_document(task, fallback_endpoint: str):
    if task["task_type"] in {CONSISTENCY_TASK_TYPE, LANGUAGE_CONSISTENCY_TASK_TYPE}:
        return _download_task_documents_zip(task, fallback_endpoint)

    upload_path = _task_upload_path(task)
    if not upload_path.is_file():
        flash("原文件已清理或缺失，无法下载。", "error")
        return redirect(request.referrer or url_for(fallback_endpoint, task_id=task["id"]))
    return send_file(
        upload_path,
        as_attachment=True,
        download_name=task["original_filename"],
    )


def _remove_uploaded_file(path: Path):
    ok, error = remove_file(path)
    if not ok:
        current_app.logger.warning("删除文件失败 path=%s error=%s", path, error)
    return ok, error


def _save_uploaded_file(upload, destination: Path) -> int:
    try:
        upload.save(destination)
        return os.path.getsize(destination)
    except Exception:
        _remove_uploaded_file(destination)
        raise


def _remove_uploaded_files(paths: list[Path]):
    failures = []
    for path in paths:
        ok, error = _remove_uploaded_file(path)
        if not ok:
            failures.append((path, error))
    return failures


def _remove_directory(path: Path):
    ok, error = remove_directory_tree(path)
    if not ok:
        current_app.logger.warning("删除目录失败 path=%s error=%s", path, error)
    return ok


def _remove_empty_directory(path: Path):
    return cleanup_remove_empty_directory(path)


def _task_upload_path(task) -> Path:
    return Path(current_app.config["UPLOAD_FOLDER"]) / Path(task["stored_filename"]).name


def _task_upload_paths(task) -> list[Path]:
    paths = _task_source_file_paths(task)
    for image in _task_image_items(task):
        image_path = image_path_from_item(_image_folder(), image)
        if image_path is not None:
            paths.append(image_path)
    return paths


def _task_source_file_paths(task) -> list[Path]:
    groups = _task_document_groups(task)
    upload_folder = Path(current_app.config["UPLOAD_FOLDER"])
    if not groups:
        return [_task_upload_path(task)]

    paths = []
    for group in groups:
        for file_info in group["files"]:
            stored_filename = Path(str(file_info.get("stored_filename") or "")).name
            if stored_filename:
                paths.append(upload_folder / stored_filename)
    return paths


def _task_source_files_available(task) -> bool:
    groups = _task_document_groups(task)
    if task["task_type"] in {CONSISTENCY_TASK_TYPE, LANGUAGE_CONSISTENCY_TASK_TYPE} and not groups:
        return False
    if groups and any(
        not Path(str(file_info.get("stored_filename") or "")).name
        for group in groups
        for file_info in group["files"]
    ):
        return False
    paths = _task_source_file_paths(task)
    return bool(paths) and all(path.is_file() for path in paths)


def _task_document_groups(task) -> list[dict]:
    return document_groups_from_meta(task["document_meta_json"])


def _task_image_items(task) -> list[dict]:
    raw = task["document_meta_json"]
    return image_items_from_meta(raw) + image_items_from_meta(raw, "page_images") + image_items_from_meta(raw, "frames")


def _image_folder() -> Path:
    configured = current_app.config.get("IMAGE_FOLDER")
    if configured:
        return Path(configured)
    return default_image_folder(current_app.config["UPLOAD_FOLDER"])


def _image_output_dir_for_stored(stored_filename: str) -> Path:
    folder = _image_folder()
    stem = Path(stored_filename).stem
    return folder / _safe_filename_part(stem, "task-images")


def _image_page_check_max_pages() -> int:
    return max(1, _int_setting("image_page_check_max_pages", DEFAULT_PDF_PAGE_IMAGE_MAX_PAGES))


def _int_setting(key: str, default: int) -> int:
    value = get_setting(key, default)
    if value is None:
        return default
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _download_task_documents_zip(task, fallback_endpoint: str):
    groups = _task_document_groups(task)
    if not groups:
        flash("文档信息缺失，无法下载。", "error")
        return redirect(request.referrer or url_for(fallback_endpoint, task_id=task["id"]))
    if not _task_source_files_available(task):
        flash("部分或全部原文件已清理或缺失，无法完整下载。", "error")
        return redirect(request.referrer or url_for(fallback_endpoint, task_id=task["id"]))

    upload_folder = Path(current_app.config["UPLOAD_FOLDER"])
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        used_names = set()
        for group in groups:
            for file_info in group["files"]:
                stored_filename = Path(str(file_info.get("stored_filename") or "")).name
                if not stored_filename:
                    continue
                upload_path = upload_folder / stored_filename
                archive_name = _unique_archive_name(
                    used_names,
                    f"{group['label']}/{Path(str(file_info.get('original_filename') or stored_filename)).name}",
                )
                archive.write(upload_path, archive_name)
    buffer.seek(0)
    return send_file(
        buffer,
        mimetype="application/zip",
        as_attachment=True,
        download_name=f"{task['task_type'] or 'document-check'}-{task['id']}-documents.zip",
    )


def _unique_archive_name(used_names: set[str], archive_name: str) -> str:
    archive_name = archive_name.strip("/\\") or "document"
    if archive_name not in used_names:
        used_names.add(archive_name)
        return archive_name
    path = Path(archive_name)
    parent = str(path.parent)
    stem = path.stem or "document"
    suffix = path.suffix
    for index in range(2, 1000):
        candidate_name = f"{stem}-{index}{suffix}"
        candidate = f"{parent}/{candidate_name}" if parent and parent != "." else candidate_name
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
    raise RuntimeError("压缩包内文件名过多，无法生成唯一名称")


def _clean_upload_filename(filename: str, file_type: str) -> str:
    name = Path(filename.replace("\\", "/")).name.strip()
    name = _safe_filename_part(name, f"document.{file_type}")
    if "." not in name:
        name = f"{name}.{file_type}"
    return name


def _upload_destination(original_filename: str, ip: str, created_at: str, file_type: str) -> tuple[str, Path]:
    upload_dir = Path(current_app.config["UPLOAD_FOLDER"])
    upload_dir.mkdir(parents=True, exist_ok=True)
    timestamp = re.sub(r"\D", "", created_at) or now_text().replace("-", "").replace(":", "").replace(" ", "")
    raw_stem = _limit_utf8_bytes(_safe_filename_part(Path(original_filename).stem, "document"), 140)
    raw_ip_part = _limit_utf8_bytes(_safe_filename_part(ip, "0.0.0.0"), 80)
    token = uuid.uuid4().hex[:12]
    stored_filename = _stored_upload_filename(raw_stem, raw_ip_part, timestamp, token, file_type, upload_dir)
    destination = upload_dir / stored_filename
    if not destination.exists():
        return stored_filename, destination

    for index in range(2, 1000):
        candidate = _stored_upload_filename(raw_stem, raw_ip_part, timestamp, token, file_type, upload_dir, index=index)
        destination = upload_dir / candidate
        if not destination.exists():
            return candidate, destination
    raise RuntimeError("无法保存上传文档，请稍后再试")


def _stored_upload_filename(
    stem: str,
    ip_part: str,
    timestamp: str,
    token: str,
    file_type: str,
    upload_dir: Path,
    *,
    index: int | None = None,
) -> str:
    budget = _upload_filename_char_budget(upload_dir)
    index_suffix = f"-{index}" if index else ""
    fixed_suffix = f"__{timestamp}_{token}{index_suffix}.{file_type}"
    if budget < len(fixed_suffix) + 2:
        raise RuntimeError("上传目录路径过长，无法生成可保存文件名，请将项目目录移动到更短路径后重试。")
    max_ip_chars = max(1, budget - len(fixed_suffix) - 1)
    ip_part = _limit_chars(ip_part, max_ip_chars)
    suffix = f"_{ip_part}_{timestamp}_{token}{index_suffix}.{file_type}"
    max_stem_chars = max(1, budget - len(suffix))
    return f"{_limit_chars(stem, max_stem_chars)}{suffix}"


def _upload_filename_char_budget(upload_dir: Path) -> int:
    try:
        upload_dir_text = str(upload_dir.resolve())
    except OSError:
        upload_dir_text = str(upload_dir.absolute())
    path_budget = UPLOAD_PATH_SAFE_CHARS - len(upload_dir_text) - 1
    return min(UPLOAD_FILENAME_SAFE_CHARS, path_budget)


def _safe_filename_part(value: str, fallback: str) -> str:
    value = INVALID_FILENAME_CHARS.sub("_", value).strip(" ._")
    value = re.sub(r"_+", "_", value)
    return value or fallback


def _limit_chars(value: str, max_chars: int) -> str:
    value = str(value or "").strip(" ._")
    if max_chars <= 0:
        return "d"
    value = value[:max_chars].strip(" ._")
    if value:
        return value
    return "document"[:max_chars] or "d"


def _limit_utf8_bytes(value: str, max_bytes: int) -> str:
    while len(value.encode("utf-8")) > max_bytes:
        value = value[:-1]
    return value or "document"


def _is_media_report_task_type(task_type: str | None) -> bool:
    return (task_type or DOCUMENT_TASK_TYPE) in {IMAGE_TASK_TYPE, VIDEO_TASK_TYPE}


def _report_item_fields_for_task(task_type: str | None) -> tuple[tuple[str, str], ...]:
    if _is_media_report_task_type(task_type):
        return MEDIA_REPORT_ITEM_FIELDS
    return REPORT_ITEM_FIELDS


def _media_report_item_text(item: dict) -> str:
    description = str(item.get("description") or "").strip()
    parts = [description] if description else []
    for field, label in MEDIA_REPORT_ITEM_DETAIL_FIELDS:
        value = str(item.get(field) or "").strip()
        if value:
            parts.append(f"{label}：{value}")
    return "\n".join(parts).strip()


def _task_results(task):
    return _prepare_task_results(
        _raw_task_results(task),
        task_type=task["task_type"] or DOCUMENT_TASK_TYPE,
        task_id=task["id"],
        record_suppression_hits=True,
    )


def _raw_task_results(task) -> list[dict]:
    return _parse_result_json(task["result_json"])


def _parse_result_json(result_json) -> list[dict]:
    if not result_json:
        return []
    try:
        data = json.loads(result_json)
    except json.JSONDecodeError:
        return []
    if not isinstance(data, list):
        return []
    return [item for item in data if isinstance(item, dict)]


def _prepare_task_results(
    results: list[dict],
    *,
    task_type: str | None = None,
    task_id: int | None = None,
    record_suppression_hits: bool = False,
    suppression_rules: dict[str, list[dict]] | None = None,
) -> list[dict]:
    prepared = []
    if suppression_rules is None:
        suppression_rules = _enabled_report_suppression_rules(task_type) if task_type else {}
    for result in results:
        item = dict(result)
        result_code = str(item.get("code") or "")
        structured_report = _result_structured_report(item)
        report_items = _result_report_items(item, structured_report)
        classifications = item.get("item_classifications")
        if not isinstance(classifications, dict):
            classifications = {}
        acceptances = item.get("item_acceptances")
        if not isinstance(acceptances, dict):
            acceptances = {}
        for report_item in report_items:
            saved_type = classifications.get(report_item["id"])
            report_item["type"] = _normalize_report_item_type(saved_type) or report_item["type"]
            report_item["type_label"] = REPORT_ITEM_TYPES[report_item["type"]]
            acceptance = _normalize_report_acceptance(acceptances.get(report_item["id"]))
            report_item.update(acceptance)
            report_item["media_summary"] = _media_report_item_text(report_item)
        original_report_item_count = len(report_items)
        report_items = _deduplicate_report_items(report_items)
        report_items.sort(key=_report_item_priority_key)
        duplicate_count = original_report_item_count - len(report_items)
        report_items, suppressed_items = _apply_report_suppression(
            task_type=task_type,
            task_id=task_id,
            result_code=result_code,
            report_items=report_items,
            suppression_rules=suppression_rules,
            record_hits=record_suppression_hits,
        )
        report_items, report_limit = _limit_ranked_report_items(
            report_items,
            issue_output_limit=item.get("issue_output_limit"),
            original_count=original_report_item_count,
            duplicate_count=duplicate_count,
        )
        for display_index, report_item in enumerate(report_items, start=1):
            report_item["index"] = display_index
        item["result_summary"] = _result_report_summary(item, structured_report)
        item["report_items"] = report_items
        item["report_limit"] = report_limit
        item["suppressed_report_items"] = suppressed_items
        item["report_counts"] = _count_report_items(report_items, suppressed_count=len(suppressed_items))
        prepared.append(item)
    return prepared


def _enabled_report_suppression_rules(task_type: str | None) -> dict[str, list[dict]]:
    if not task_type:
        return {}
    rows = get_db().execute(
        """
        SELECT id, check_code, reason, item_json
        FROM report_suppression_rules
        WHERE task_type = ? AND enabled = 1
        """,
        (task_type,),
    ).fetchall()
    rules = {}
    for row in rows:
        snapshot = _parse_report_suppression_item_json(row["item_json"])
        description = _report_suppression_description(snapshot)
        if not description:
            continue
        rules.setdefault(row["check_code"], []).append(
            {
                "id": row["id"],
                "reason": row["reason"] or "",
                "description": description,
            }
        )
    return rules


def _apply_report_suppression(
    *,
    task_type: str | None,
    task_id: int | None,
    result_code: str,
    report_items: list[dict],
    suppression_rules: dict[str, list[dict]],
    record_hits: bool,
) -> tuple[list[dict], list[dict]]:
    if not task_type or not suppression_rules:
        return report_items, []

    visible_items = []
    suppressed_items = []
    db = get_db() if record_hits and task_id is not None else None
    for item in report_items:
        rule, similarity = _matching_report_suppression_rule(
            item,
            suppression_rules.get(result_code, []),
        )
        if rule is None:
            visible_items.append(item)
            continue
        suppressed = dict(item)
        suppressed["suppression_rule_id"] = rule["id"]
        suppressed["suppression_reason"] = rule["reason"]
        suppressed["suppression_similarity"] = similarity
        suppressed_items.append(suppressed)
        if db is not None:
            _record_report_suppression_hit(
                db,
                rule_id=int(rule["id"]),
                task_id=int(task_id),
                result_code=result_code,
                item=suppressed,
            )
    return visible_items, suppressed_items


def _matching_report_suppression_rule(item: dict, rules: list[dict]) -> tuple[dict | None, float]:
    description = _report_suppression_description(item)
    if not description:
        return None, 0.0
    best_rule = None
    best_similarity = 0.0
    for rule in rules:
        similarity = _report_description_similarity(description, rule.get("description"))
        if similarity > best_similarity:
            best_rule = rule
            best_similarity = similarity
    if best_similarity < REPORT_SUPPRESSION_DESCRIPTION_SIMILARITY_THRESHOLD:
        return None, best_similarity
    return best_rule, best_similarity


def _report_description_similarity(left, right) -> float:
    left_text = _normalize_report_description(left)
    right_text = _normalize_report_description(right)
    if not left_text or not right_text:
        return 0.0
    if left_text == right_text:
        return 1.0
    shorter, longer = sorted((left_text, right_text), key=len)
    if len(shorter) >= 6 and shorter in longer:
        return 0.95
    if len(shorter) < 4:
        return 0.0
    sequence_score = SequenceMatcher(None, left_text, right_text).ratio()
    character_score = _report_description_dice(set(left_text), set(right_text)) * 0.9
    bigram_score = _report_description_dice(
        _report_description_ngrams(left_text, 2),
        _report_description_ngrams(right_text, 2),
    )
    return max(sequence_score, character_score, bigram_score)


def _normalize_report_description(value) -> str:
    text = str(value or "").strip().lower()
    for source, replacement in REPORT_SUPPRESSION_DESCRIPTION_REPLACEMENTS:
        text = text.replace(source, replacement)
    return re.sub(r"[\W_]+", "", text)


def _report_description_ngrams(text: str, size: int) -> set[str]:
    if len(text) < size:
        return {text} if text else set()
    return {text[index : index + size] for index in range(len(text) - size + 1)}


def _report_description_dice(left: set[str], right: set[str]) -> float:
    if not left or not right:
        return 0.0
    return 2 * len(left & right) / (len(left) + len(right))


def _record_report_suppression_hit(db, *, rule_id: int, task_id: int, result_code: str, item: dict):
    now = now_text()
    cursor = db.execute(
        """
        INSERT OR IGNORE INTO report_suppression_hits(rule_id, task_id, result_code, item_id, item_json, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (
            rule_id,
            task_id,
            result_code,
            str(item.get("id") or ""),
            json.dumps(_report_suppression_item_snapshot(item), ensure_ascii=False),
            now,
        ),
    )
    if cursor.rowcount:
        db.execute(
            """
            UPDATE report_suppression_rules
            SET hit_count = hit_count + 1, last_hit_at = ?, updated_at = ?
            WHERE id = ?
            """,
            (now, now, rule_id),
        )
        db.commit()


def _maybe_create_report_suppression_candidate(
    db,
    *,
    task,
    result_code: str,
    result: dict,
    item_id: str,
    item_type: str,
    acceptance_status: str | None,
    rejection_reason: str,
    rejection_note: str,
) -> bool:
    if (
        item_type != "non_issue"
        or acceptance_status != "rejected"
        or rejection_reason not in REPORT_SUPPRESSION_REJECTION_REASONS
    ):
        return False

    report_item = next((item for item in _result_report_items(result) if item.get("id") == item_id), None)
    if report_item is None:
        return False

    task_type = task["task_type"] or DOCUMENT_TASK_TYPE
    fingerprint = _report_item_suppression_fingerprint(task_type, result_code, report_item)
    now = now_text()
    reason = REPORT_REJECTION_REASONS.get(rejection_reason, "") or rejection_note
    snapshot = _report_suppression_item_snapshot(report_item)
    db.execute(
        """
        INSERT INTO report_suppression_rules(
            task_type, check_code, fingerprint, item_json, reason, enabled,
            source_task_id, source_result_code, source_item_id,
            hit_count, created_at, updated_at
        )
        VALUES (?, ?, ?, ?, ?, 0, ?, ?, ?, 0, ?, ?)
        ON CONFLICT(task_type, check_code, fingerprint) DO UPDATE SET
            item_json = excluded.item_json,
            reason = CASE
                WHEN report_suppression_rules.reason IS NULL OR report_suppression_rules.reason = ''
                THEN excluded.reason
                ELSE report_suppression_rules.reason
            END,
            updated_at = excluded.updated_at
        """,
        (
            task_type,
            result_code,
            fingerprint,
            json.dumps(snapshot, ensure_ascii=False),
            reason,
            task["id"],
            result_code,
            item_id,
            now,
            now,
        ),
    )
    return True


def _report_item_suppression_fingerprint(task_type: str, result_code: str, item: dict) -> str:
    source = {
        "task_type": str(task_type or ""),
        "check_code": str(result_code or ""),
        "description": _report_suppression_description(item),
    }
    raw = json.dumps(source, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _report_suppression_description(item: dict) -> str:
    return _normalize_suppression_text(item.get("description") or item.get("text"))


def _report_suppression_item_snapshot(item: dict) -> dict:
    fields = _report_suppression_item_fields(item)
    return {
        **fields,
        "text": _normalize_suppression_text(item.get("text")),
        "type": _normalize_report_item_type(item.get("type")) or "issue",
    }


def _report_suppression_item_fields(item: dict) -> dict:
    return {
        field: _normalize_suppression_text(item.get(field))
        for field in REPORT_SUPPRESSION_FIELDS
    }


def _normalize_suppression_text(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def _result_report_items(result: dict, structured_report: dict | None = None) -> list[dict]:
    code = str(result.get("code") or "")
    if structured_report is None:
        structured_report = _result_structured_report(result)
    if structured_report is not None:
        return _structured_report_items(code, structured_report)

    text = str(result.get("result") or "").strip()
    if not text:
        return []
    chunks = _extract_report_item_chunks(text) or [text]
    items = []
    for index, chunk in enumerate(chunks, start=1):
        item_text = chunk.strip()
        if not item_text:
            continue
        fields = _legacy_report_item_fields(item_text)
        items.append(
            {
                "id": _report_item_id(code, index, item_text),
                "index": index,
                "text": item_text,
                **fields,
                "type": _infer_report_item_type(item_text),
            }
        )
    return items


def _result_structured_report(result: dict) -> dict | None:
    for key in ("structured_report", "report_json"):
        structured = _normalize_structured_report_payload(result.get(key))
        if structured is not None:
            return structured

    structured_items = result.get("structured_items")
    if isinstance(structured_items, list):
        summary = _first_report_field(result, REPORT_JSON_SUMMARY_KEYS)
        return {"summary": summary, "items": structured_items}

    return _normalize_structured_report_payload(result.get("result"))


def _result_report_summary(result: dict, structured_report: dict | None) -> str:
    if structured_report is not None:
        return str(structured_report.get("summary") or "").strip()
    return ""


def _normalize_structured_report_payload(value) -> dict | None:
    payload = value
    if isinstance(value, str):
        payload = _parse_structured_report_json(value)
    if isinstance(payload, list):
        return {"summary": "", "items": payload}
    if not isinstance(payload, dict):
        return None

    items = None
    for key in REPORT_JSON_ITEM_KEYS:
        candidate = payload.get(key)
        if isinstance(candidate, list):
            items = candidate
            break
        if isinstance(candidate, str):
            parsed_items = _parse_structured_report_json(candidate)
            if isinstance(parsed_items, list):
                items = parsed_items
                break
    summary = _first_report_field(payload, REPORT_JSON_SUMMARY_KEYS)
    if items is None:
        if any(_first_report_field(payload, aliases) for aliases in REPORT_FIELD_ALIASES.values()):
            items = [payload]
        elif summary:
            items = []
        else:
            return None
    return {"summary": summary, "items": items}


def _parse_structured_report_json(text: str, depth: int = 0):
    text = str(text or "").strip()
    if not text:
        return None
    for candidate in _structured_json_candidates(text):
        parsed = _load_structured_json_candidate(candidate, depth)
        if isinstance(parsed, (dict, list)):
            return parsed
    return None


def _load_structured_json_candidate(candidate: str, depth: int):
    if depth > 3:
        return None
    variants = _json_candidate_variants(candidate)
    for variant in variants:
        try:
            parsed = json.loads(variant)
        except (TypeError, json.JSONDecodeError):
            continue
        if isinstance(parsed, str):
            nested = _parse_structured_report_json(parsed, depth + 1)
            if nested is not None:
                return nested
            continue
        return parsed
    for variant in variants:
        parsed = _parse_truncated_structured_report_candidate(variant)
        if parsed is not None:
            return parsed
    return None


def _parse_truncated_structured_report_candidate(candidate: str) -> dict | None:
    text = str(candidate or "").strip()
    object_start = text.find("{")
    if object_start < 0:
        return None
    text = text[object_start:]

    item_array_matches = []
    for key in REPORT_JSON_ITEM_KEYS:
        match = re.search(rf'"{re.escape(key)}"\s*:\s*\[', text)
        if match:
            item_array_matches.append((match.start(), match.end() - 1, key))
    if not item_array_matches:
        return None

    _, array_start, item_key = min(item_array_matches)
    try:
        payload = json.loads(f"{text[:array_start + 1]}]}}")
    except (TypeError, json.JSONDecodeError):
        return None
    if not isinstance(payload, dict):
        return None

    decoder = json.JSONDecoder()
    items = []
    position = array_start + 1
    while position < len(text):
        while position < len(text) and (text[position].isspace() or text[position] == ","):
            position += 1
        if position >= len(text) or text[position] == "]":
            break
        try:
            item, end = decoder.raw_decode(text, position)
        except json.JSONDecodeError:
            break
        if isinstance(item, (dict, str)):
            items.append(item)
        position = end

    if not items:
        return None
    payload[item_key] = items
    return payload


def _json_candidate_variants(candidate: str) -> list[str]:
    raw = str(candidate or "").strip()
    if not raw:
        return []
    variants = [raw]
    repaired = _escape_json_string_control_chars(raw)
    if repaired != raw:
        variants.append(repaired)
    for variant in tuple(variants):
        repaired = _repair_common_json_model_errors(variant)
        if repaired != variant and repaired not in variants:
            variants.append(repaired)
    return variants


def _repair_common_json_model_errors(text: str) -> str:
    status_values = "|".join(re.escape(value) for value in REPORT_ITEM_TYPES)
    status_keys = r"status|classification|item_type|type"
    return re.sub(
        rf'("(?:(?:{status_keys}))"\s*:\s*")({status_values})"\s*:\s*"({status_values})(")',
        r"\1\3\4",
        str(text or ""),
    )


def _escape_json_string_control_chars(text: str) -> str:
    result = []
    in_string = False
    escaped = False
    for char in str(text or ""):
        if escaped:
            result.append(char)
            escaped = False
            continue
        if char == "\\":
            result.append(char)
            escaped = True
            continue
        if char == '"':
            in_string = not in_string
            result.append(char)
            continue
        if in_string and char in {"\n", "\r", "\t"}:
            result.append({"\n": "\\n", "\r": "\\r", "\t": "\\t"}[char])
            continue
        result.append(char)
    return "".join(result)


def _structured_json_candidates(text: str) -> list[str]:
    candidates = [text]
    for match in re.finditer(r"```(?:json)?\s*(.*?)```", text, flags=re.IGNORECASE | re.DOTALL):
        candidates.append(match.group(1).strip())
    object_start = text.find("{")
    object_end = text.rfind("}")
    if object_start >= 0 and object_end > object_start:
        candidates.append(text[object_start : object_end + 1])
    array_start = text.find("[")
    array_end = text.rfind("]")
    if array_start >= 0 and array_end > array_start:
        candidates.append(text[array_start : array_end + 1])

    seen = set()
    unique = []
    for candidate in candidates:
        candidate = str(candidate or "").strip()
        if candidate and candidate not in seen:
            seen.add(candidate)
            unique.append(candidate)
    return unique


def _structured_report_items(result_code: str, structured_report: dict) -> list[dict]:
    raw_items = structured_report.get("items")
    if not isinstance(raw_items, list):
        return []

    items = []
    for raw_item in raw_items:
        fields = _normalize_structured_report_item(raw_item)
        if not fields:
            continue
        index = len(items) + 1
        item_text = _structured_report_item_text(fields)
        item_type = fields.pop("type", "") or _infer_report_item_type(item_text)
        items.append(
            {
                "id": _report_item_id(result_code, index, item_text),
                "index": index,
                "text": item_text,
                **fields,
                "type": item_type,
            }
        )
    return items


def _normalize_structured_report_item(raw_item) -> dict:
    if isinstance(raw_item, str):
        text = raw_item.strip()
        parsed = _parse_structured_report_json(text)
        if isinstance(parsed, dict):
            return _normalize_structured_report_item(parsed)
        return {
            "severity": "",
            "severity_label": "",
            "confidence": "",
            "confidence_label": "",
            "category": "",
            "location": "",
            "excerpt": "",
            "description": text,
            "impact": "",
            "suggestion": "",
            "type": _infer_report_item_type(text),
        } if text else {}
    if not isinstance(raw_item, dict):
        return {}

    status = _normalize_report_item_status(_first_report_field(raw_item, REPORT_STATUS_KEYS))
    fields = {
        field: _first_report_field(raw_item, aliases)
        for field, aliases in REPORT_FIELD_ALIASES.items()
    }
    fields["severity"] = _normalize_report_severity(fields.get("severity"))
    fields["severity_label"] = REPORT_SEVERITY_LABELS.get(fields["severity"], "")
    fields["confidence"] = _normalize_report_confidence(fields.get("confidence"))
    fields["confidence_label"] = REPORT_CONFIDENCE_LABELS.get(fields["confidence"], "")
    if _looks_like_status_only(fields["category"]):
        status = status or _normalize_report_item_status(fields["category"])
        fields["category"] = ""
    if not any(fields.get(field) for field in REPORT_SUPPRESSION_FIELDS):
        fields["description"] = _report_field_text(raw_item)
    item_text = _structured_report_item_text(fields)
    fields["type"] = "non_issue" if _is_no_action_report_item(fields) else status or _infer_report_item_type(item_text)
    return fields


def _normalize_report_severity(value) -> str:
    compact = _compact_report_text(value)
    aliases = {
        "critical": "critical",
        "fatal": "critical",
        "blocker": "critical",
        "致命": "critical",
        "灾难性": "critical",
        "极高": "critical",
        "high": "high",
        "严重": "high",
        "重大": "high",
        "高": "high",
        "medium": "medium",
        "middle": "medium",
        "moderate": "medium",
        "一般": "medium",
        "中": "medium",
        "low": "low",
        "minor": "low",
        "轻微": "low",
        "低": "low",
    }
    return aliases.get(compact, "")


def _normalize_report_confidence(value) -> str:
    compact = _compact_report_text(value)
    aliases = {
        "high": "high",
        "certain": "high",
        "明确": "high",
        "高": "high",
        "medium": "medium",
        "middle": "medium",
        "moderate": "medium",
        "较高": "medium",
        "中": "medium",
        "low": "low",
        "uncertain": "low",
        "较低": "low",
        "低": "low",
    }
    return aliases.get(compact, "")


def _deduplicate_report_items(report_items: list[dict]) -> list[dict]:
    unique_items = []
    items_by_key = {}
    for report_item in report_items:
        key = (
            _normalize_suppression_text(report_item.get("type")),
            _normalize_suppression_text(report_item.get("category")),
            _normalize_suppression_text(report_item.get("excerpt")),
            _normalize_suppression_text(report_item.get("description")),
            _normalize_suppression_text(report_item.get("impact")),
            _normalize_suppression_text(report_item.get("suggestion")),
        )
        if not any(key[1:]):
            unique_items.append(report_item)
            continue
        existing = items_by_key.get(key)
        if existing is None:
            items_by_key[key] = report_item
            unique_items.append(report_item)
            continue
        existing["location"] = _merge_report_field_values(existing.get("location"), report_item.get("location"))
        if _report_item_priority_key(report_item) < _report_item_priority_key(existing):
            existing["severity"] = report_item.get("severity", "")
            existing["severity_label"] = report_item.get("severity_label", "")
            existing["confidence"] = report_item.get("confidence", "")
            existing["confidence_label"] = report_item.get("confidence_label", "")
    return unique_items


def _merge_report_field_values(left, right) -> str:
    left_text = str(left or "").strip()
    right_text = str(right or "").strip()
    if not left_text:
        return right_text
    if not right_text or right_text in left_text:
        return left_text
    return f"{left_text}；{right_text}"


def _report_item_priority_key(report_item: dict) -> tuple[int, int, int, int]:
    item_type_order = {"issue": 0, "suggestion": 1, "non_issue": 2}
    item_type = str(report_item.get("type") or "")
    return (
        item_type_order.get(item_type, len(item_type_order)),
        REPORT_CONFIDENCE_ORDER.get(str(report_item.get("confidence") or ""), len(REPORT_CONFIDENCE_ORDER)),
        REPORT_SEVERITY_ORDER.get(str(report_item.get("severity") or ""), len(REPORT_SEVERITY_ORDER)),
        int(report_item.get("index") or 0),
    )


def _limit_ranked_report_items(
    report_items: list[dict],
    *,
    issue_output_limit,
    original_count: int,
    duplicate_count: int,
) -> tuple[list[dict], dict | None]:
    limit = normalize_issue_output_limit(issue_output_limit)

    before_limit_count = len(report_items)
    omitted_count = 0
    limit_applied = False
    if before_limit_count > limit:
        omitted_count = before_limit_count - limit
        report_items = report_items[:limit]
        limit_applied = True

    if not duplicate_count and not omitted_count:
        return report_items, None
    return report_items, {
        "limit": limit,
        "original_count": original_count,
        "duplicate_count": duplicate_count,
        "deduplicated_count": original_count - duplicate_count,
        "displayed_count": len(report_items),
        "omitted_count": omitted_count,
        "limit_applied": limit_applied,
        "missing_ranking": False,
    }


def _first_report_field(source: dict, aliases: tuple[str, ...]) -> str:
    if not isinstance(source, dict):
        return ""
    for key in aliases:
        if key in source:
            text = _report_field_text(source.get(key))
            if text:
                return text
    return ""


def _report_field_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (int, float, bool)):
        return str(value)
    if isinstance(value, list):
        parts = []
        for item in value:
            text = _report_field_text(item)
            if text:
                parts.append(text)
        return "；".join(parts)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return str(value).strip()


def _structured_report_item_text(fields: dict) -> str:
    lines = []
    for key, label in REPORT_ITEM_FIELDS:
        value = str(fields.get(key) or "").strip()
        if value:
            lines.append(f"{label}：{value}")
    return "\n".join(lines).strip()


def _is_no_action_report_item(fields: dict) -> bool:
    impact = _compact_report_text(fields.get("impact"))
    suggestion = _compact_report_text(fields.get("suggestion"))
    return _has_report_marker(impact, REPORT_NO_ACTION_IMPACT_MARKERS) and _has_report_marker(
        suggestion,
        REPORT_NO_ACTION_SUGGESTION_MARKERS,
    )


def _compact_report_text(value) -> str:
    return re.sub(r"[\s，。；、,.!！?？:：;；/\\|()（）【】\\[\\]\"'“”‘’_-]+", "", str(value or "")).lower()


def _has_report_marker(text: str, markers: tuple[str, ...]) -> bool:
    return any(marker in text for marker in markers)


def _legacy_report_item_fields(text: str) -> dict:
    fields = {key: "" for key, _ in REPORT_ITEM_FIELDS}
    body_lines = []
    for raw_line in str(text or "").splitlines():
        line = _clean_report_item_line(raw_line)
        if not line:
            continue
        match = re.match(r"^([^:：]{1,28})[:：]\s*(.*)$", line)
        if match:
            label = match.group(1).strip()
            value = match.group(2).strip()
            field = REPORT_LEGACY_LABEL_FIELDS.get(label)
            if field and value:
                fields[field] = _append_report_field(fields[field], value)
                continue
        body_lines.append(line)
    if not fields["description"]:
        fields["description"] = "\n".join(body_lines).strip() or fields["suggestion"] or str(text or "").strip()
    return fields


def _clean_report_item_line(line: str) -> str:
    stripped = REPORT_ITEM_PREFIX_RE.sub("", str(line or "").strip()).strip()
    stripped = re.sub(r"^(?:\d{1,3}[.、)]|\(\d{1,3}\)|（\d{1,3}）)\s*", "", stripped)
    return stripped.strip("*_`~ ")


def _append_report_field(current: str, value: str) -> str:
    current = str(current or "").strip()
    value = str(value or "").strip()
    if not current:
        return value
    if not value or value in current.split("；"):
        return current
    return f"{current}；{value}"


def _looks_like_status_only(value: str) -> bool:
    compact = re.sub(r"\s+", "", str(value or "")).strip()
    return compact.lower() in {"issue", "problem", "suggestion", "advice", "non_issue", "nonissue", "not_issue"} or compact in {
        "问题",
        "明确问题",
        "建议",
        "需人工确认",
        "非问题",
        "不是问题",
    }


def _extract_report_item_chunks(text: str) -> list[str]:
    chunks = []
    current = []
    for line in str(text or "").splitlines():
        if _is_report_item_start(line):
            if current:
                chunks.append("\n".join(current).strip())
            current = [line]
            continue
        if current and _is_report_auxiliary_section_start(line):
            chunks.append("\n".join(current).strip())
            current = []
            continue
        if current:
            current.append(line)
    if current:
        chunks.append("\n".join(current).strip())
    return [chunk for chunk in chunks if chunk]


def _is_report_item_start(line: str) -> bool:
    stripped = str(line or "").strip()
    if not stripped:
        return False
    if stripped.startswith(("|", "```", ">")):
        return False
    match_text = REPORT_ITEM_PREFIX_RE.sub("", stripped).strip()
    return bool(REPORT_ITEM_START_RE.match(match_text))


def _is_report_auxiliary_section_start(line: str) -> bool:
    stripped = REPORT_ITEM_PREFIX_RE.sub("", str(line or "").strip()).strip("*_`~ ")
    if not stripped:
        return False
    if stripped in {"总体判断", "明确问题", "需人工确认"}:
        return True
    return stripped.startswith(
        (
            "页面级检查结果",
            "图文联合检查结果",
            "图片检查结果",
            "视频帧检查结果",
            "检查汇总",
            "覆盖图片",
            "覆盖视频帧",
            "已跳过的图片",
            "系统需人工确认",
        )
    )


def _report_item_id(result_code: str, index: int, text: str) -> str:
    source = f"{result_code}\n{index}\n{text}"
    return hashlib.sha1(source.encode("utf-8"), usedforsecurity=False).hexdigest()[:16]


def _infer_report_item_type(text: str) -> str:
    compact = re.sub(r"\s+", "", str(text or ""))
    if any(marker in compact for marker in ("非问题", "未发现", "无明显", "未见明显", "无需修改", "未见异常", "无异常")):
        return "non_issue"
    if any(marker in compact for marker in ("需人工确认", "人工确认", "疑似", "不确定", "证据不足", "建议核实", "建议复核", "看不清", "无法确认")):
        return "suggestion"
    issue_markers = (
        "问题",
        "错误",
        "不一致",
        "矛盾",
        "冲突",
        "缺失",
        "风险",
        "不规范",
        "不匹配",
        "异常",
    )
    if "建议" in compact and not any(marker in compact for marker in issue_markers):
        return "suggestion"
    if any(marker in compact for marker in issue_markers):
        return "issue"
    return "suggestion"


def _normalize_report_item_status(value) -> str | None:
    exact = _normalize_report_item_type(value)
    if exact:
        return exact
    compact = re.sub(r"\s+", "", str(value or "")).strip().lower()
    if not compact:
        return None
    if any(marker in compact for marker in ("non_issue", "nonissue", "not_issue", "非问题", "不是问题", "无需修改", "无问题")):
        return "non_issue"
    if any(marker in compact for marker in ("suggestion", "advise", "manual", "uncertain", "建议", "需人工确认", "人工确认", "疑似", "不确定", "证据不足")):
        return "suggestion"
    if any(marker in compact for marker in ("issue", "problem", "明确问题", "问题", "错误", "不一致", "缺失", "冲突")):
        return "issue"
    return None


def _normalize_report_item_type(value) -> str | None:
    value = str(value or "").strip()
    return value if value in REPORT_ITEM_TYPES else None


def _normalize_report_acceptance_status(value) -> str | None:
    value = str(value or "").strip()
    return value if value in REPORT_ACCEPTANCE_STATUSES else None


def _normalize_report_rejection_reason(value) -> str:
    value = str(value or "").strip()
    return value if value in REPORT_REJECTION_REASONS else ""


def _normalize_report_acceptance(value) -> dict:
    if not isinstance(value, dict):
        status = _normalize_report_acceptance_status(value) or "pending"
        return {
            "acceptance_status": status,
            "acceptance_label": REPORT_ACCEPTANCE_STATUSES[status],
            "rejection_reason": "",
            "rejection_reason_label": "",
            "rejection_note": "",
        }

    status = _normalize_report_acceptance_status(value.get("status")) or "pending"
    reason = _normalize_report_rejection_reason(value.get("rejection_reason")) if status == "rejected" else ""
    note = str(value.get("rejection_note") or "").strip() if status == "rejected" else ""
    return {
        "acceptance_status": status,
        "acceptance_label": REPORT_ACCEPTANCE_STATUSES[status],
        "rejection_reason": reason,
        "rejection_reason_label": REPORT_REJECTION_REASONS.get(reason, ""),
        "rejection_note": note,
    }


def _report_rate_label(numerator: int, denominator: int) -> str:
    if denominator <= 0:
        return "-"
    return f"{numerator / denominator * 100:.1f}%"


def _finalize_report_counts(counts: dict) -> dict:
    counts["total"] = sum(int(counts.get(key) or 0) for key in REPORT_ITEM_TYPE_ORDER)
    confirmed_issues = int(counts.get("accepted_issue") or 0) + int(counts.get("rejected_issue") or 0)
    counts["issue_detection_rate"] = _report_rate_label(int(counts.get("issue") or 0), counts["total"])
    counts["issue_acceptance_rate"] = _report_rate_label(int(counts.get("accepted_issue") or 0), confirmed_issues)
    return counts


def _count_report_items(items: list[dict], *, suppressed_count: int = 0) -> dict:
    counts = {key: 0 for key in REPORT_ITEM_TYPE_ORDER}
    counts.update(
        {
            "accepted_issue": 0,
            "rejected_issue": 0,
            "pending_issue_acceptance": 0,
            "suppressed": max(0, int(suppressed_count or 0)),
        }
    )
    for item in items:
        item_type = _normalize_report_item_type(item.get("type")) or "issue"
        counts[item_type] += 1
        if item_type == "issue":
            acceptance_status = _normalize_report_acceptance_status(item.get("acceptance_status")) or "pending"
            if acceptance_status == "accepted":
                counts["accepted_issue"] += 1
            elif acceptance_status == "rejected":
                counts["rejected_issue"] += 1
            else:
                counts["pending_issue_acceptance"] += 1
    return _finalize_report_counts(counts)


def _report_item_totals(results: list[dict]) -> dict:
    totals = {key: 0 for key in REPORT_ITEM_TYPE_ORDER}
    totals.update(
        {
            "accepted_issue": 0,
            "rejected_issue": 0,
            "pending_issue_acceptance": 0,
            "suppressed": 0,
        }
    )
    for result in results:
        counts = result.get("report_counts") or {}
        for key in tuple(REPORT_ITEM_TYPE_ORDER) + (
            "accepted_issue",
            "rejected_issue",
            "pending_issue_acceptance",
            "suppressed",
        ):
            totals[key] += int(counts.get(key) or 0)
    return _finalize_report_counts(totals)


def _update_report_item_type(task):
    if task["status"] in {"queued", "running"}:
        return {"ok": False, "error": "任务尚未完成，暂不能修改报告条目判定。"}, 409
    data = request.get_json(silent=True) if request.is_json else None
    if not isinstance(data, dict):
        data = request.form
    result_code = str(data.get("result_code") or "").strip()
    item_id = str(data.get("item_id") or "").strip()
    item_type = _normalize_report_item_type(data.get("item_type"))
    if not result_code or not item_id or not item_type:
        return {"ok": False, "error": "报告条目判定数据无效。"}, 400
    acceptance_supplied = "acceptance_status" in data
    acceptance_status = None
    rejection_reason = ""
    rejection_note = ""
    if acceptance_supplied:
        acceptance_status = _normalize_report_acceptance_status(data.get("acceptance_status"))
        rejection_reason = _normalize_report_rejection_reason(data.get("rejection_reason"))
        rejection_note = str(data.get("rejection_note") or "").strip()
        if acceptance_status is None:
            return {"ok": False, "error": "接纳状态数据无效。"}, 400
        if acceptance_status == "rejected":
            if rejection_reason == "other" and not rejection_note:
                return {"ok": False, "error": "选择其他原因时必须填写具体原因。"}, 400

    results = _raw_task_results(task)
    target = None
    valid_item_ids = set()
    for result in results:
        if str(result.get("code") or "") != result_code:
            continue
        target = result
        valid_item_ids = {item["id"] for item in _result_report_items(result)}
        break
    if target is None or item_id not in valid_item_ids:
        return {"ok": False, "error": "报告条目不存在。"}, 404

    classifications = target.get("item_classifications")
    if not isinstance(classifications, dict):
        classifications = {}
    classifications[item_id] = item_type
    target["item_classifications"] = classifications
    if acceptance_supplied:
        acceptances = target.get("item_acceptances")
        if not isinstance(acceptances, dict):
            acceptances = {}
        if acceptance_status == "pending":
            acceptances.pop(item_id, None)
        else:
            record = {"status": acceptance_status}
            if acceptance_status == "rejected":
                record["rejection_reason"] = rejection_reason
                record["rejection_note"] = rejection_note
            acceptances[item_id] = record
        target["item_acceptances"] = acceptances

    db = get_db()
    suppression_candidate_created = _maybe_create_report_suppression_candidate(
        db,
        task=task,
        result_code=result_code,
        result=target,
        item_id=item_id,
        item_type=item_type,
        acceptance_status=acceptance_status,
        rejection_reason=rejection_reason,
        rejection_note=rejection_note,
    )
    db.execute(
        "UPDATE tasks SET result_json = ?, updated_at = ? WHERE id = ?",
        (json.dumps(results, ensure_ascii=False), now_text(), task["id"]),
    )
    db.commit()

    prepared = _prepare_task_results(
        results,
        task_type=task["task_type"] or DOCUMENT_TASK_TYPE,
        task_id=task["id"],
    )
    updated_result = next((item for item in prepared if str(item.get("code") or "") == result_code), None)
    updated_report_item = None
    if updated_result:
        updated_report_item = next(
            (item for item in updated_result.get("report_items", []) if item.get("id") == item_id),
            None,
        )
    return {
        "ok": True,
        "item_id": item_id,
        "item_type": item_type,
        "item_type_label": REPORT_ITEM_TYPES[item_type],
        "acceptance_status": (updated_report_item or {}).get("acceptance_status", "pending"),
        "acceptance_label": (updated_report_item or {}).get("acceptance_label", REPORT_ACCEPTANCE_STATUSES["pending"]),
        "rejection_reason": (updated_report_item or {}).get("rejection_reason", ""),
        "rejection_reason_label": (updated_report_item or {}).get("rejection_reason_label", ""),
        "rejection_note": (updated_report_item or {}).get("rejection_note", ""),
        "result_counts": (updated_result or {}).get("report_counts", {}),
        "totals": _report_item_totals(prepared),
        "suppression_candidate_created": suppression_candidate_created,
    }


def _export_task_report(task):
    static_folder = current_app.static_folder
    if not static_folder:
        raise RuntimeError("静态资源目录未配置，无法导出报告。")
    app_css = (Path(static_folder) / "app.css").read_text(encoding="utf-8")
    results = _task_results(task)
    html = render_template(
        "task_report_export.html",
        task=task,
        results=results,
        report_totals=_report_item_totals(results),
        report_item_types=REPORT_ITEM_TYPES,
        report_item_fields=_report_item_fields_for_task(task["task_type"]),
        media_report=_is_media_report_task_type(task["task_type"]),
        document_groups=_task_document_groups(task),
        app_css=app_css,
    )
    filename = f"document-check-report-{task['id']}.html"
    return Response(
        html,
        mimetype="text/html",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _export_task_report_excel(task):
    results = _task_results(task)
    report_totals = _report_item_totals(results)
    document_groups = _task_document_groups(task)
    workbook = Workbook()
    report_sheet = workbook.active
    report_sheet.title = "报告条目"

    _fill_report_items_sheet(report_sheet, task, results, document_groups)
    _fill_report_totals_sheet(workbook.create_sheet("统计"), report_totals)

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    filename = f"document-check-report-{task['id']}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype=REPORT_EXPORT_MIMETYPE,
    )


def _fill_report_items_sheet(sheet, task, results: list[dict], document_groups: list[dict]) -> None:
    report_item_fields = _report_item_fields_for_task(task["task_type"])
    headers = [
        "任务ID",
        "任务类型",
        "文件名称",
        "检查项",
        "条目",
        *[label for _, label in report_item_fields],
        REPORT_ITEM_TYPE_LABEL,
        "是否接纳",
        "不接纳原因",
        "人工原因",
    ]
    sheet.append(headers)
    context = _excel_task_context(task, document_groups)
    for result in results:
        report_items = result.get("report_items") or []
        if not report_items:
            sheet.append(
                [
                    *context,
                    _excel_cell_text(result.get("name")),
                    "",
                    *["" for _ in report_item_fields],
                    "未拆分",
                    "",
                    "",
                    "",
                ]
            )
            continue
        for item in report_items:
            sheet.append(
                [
                    *context,
                    _excel_cell_text(result.get("name")),
                    f"条目 {item.get('index')}",
                    *[_excel_cell_text(item.get(field)) for field, _ in report_item_fields],
                    _excel_cell_text(item.get("type_label")),
                    _excel_cell_text(item.get("acceptance_label")),
                    _excel_cell_text(item.get("rejection_reason_label")) if item.get("acceptance_status") == "rejected" else "",
                    _excel_cell_text(item.get("rejection_note")) if item.get("acceptance_status") == "rejected" else "",
                ]
            )
    _style_excel_sheet(sheet)


def _fill_report_totals_sheet(sheet, report_totals: dict) -> None:
    sheet.append(["指标", "值"])
    for label, key in REPORT_TOTAL_EXPORT_ROWS:
        sheet.append([label, report_totals.get(key, 0)])
    _style_excel_sheet(sheet)


def _excel_task_context(task, document_groups: list[dict]) -> list:
    return [
        _row_value(task, "id", ""),
        task_type_label(_row_value(task, "task_type", DOCUMENT_TASK_TYPE)),
        _excel_document_names(task, document_groups),
    ]


def _excel_document_names(task, document_groups: list[dict]) -> str:
    names = []
    for group in document_groups:
        for file in group.get("files", []):
            name = str(file.get("original_filename") or "").strip()
            if name:
                names.append(name)
    if names:
        return "\n".join(names)
    return _excel_cell_text(_row_value(task, "original_filename", ""))


def _excel_cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (int, float)):
        return value
    return str(value).strip()


def _style_excel_sheet(sheet) -> None:
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        cell.font = REPORT_EXPORT_HEADER_FONT
        cell.fill = REPORT_EXPORT_HEADER_FILL
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells[:200]) + 2
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(width, 10), 60)


def _page_arg() -> int:
    try:
        page = int(request.args.get("page", "1"))
    except ValueError:
        return 1
    return max(1, page)


def _bounded_page(page: int, total: int, per_page: int) -> int:
    pages = max(1, (total + per_page - 1) // per_page)
    return min(max(1, page), pages)


def _pagination(page: int, total: int, per_page: int) -> dict:
    pages = max(1, (total + per_page - 1) // per_page)
    return {
        "page": page,
        "pages": pages,
        "per_page": per_page,
        "total": total,
        "has_prev": page > 1,
        "has_next": page < pages,
        "prev_page": max(1, page - 1),
        "next_page": min(pages, page + 1),
        "start": 0 if total == 0 else (page - 1) * per_page + 1,
        "end": min(total, page * per_page),
    }


def _task_stats_for_where(where: str, params: tuple) -> dict:
    stats = {"total": 0, "queued": 0, "running": 0, "completed": 0, "failed": 0, "canceled": 0}
    rows = get_db().execute(
        f"SELECT status, COUNT(*) AS total FROM tasks WHERE {where} GROUP BY status",
        params,
    ).fetchall()
    for row in rows:
        count = row["total"]
        stats["total"] += count
        if row["status"] in stats:
            stats[row["status"]] = count
    return stats
